import shutil
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, "src")

from models.yccd_period_map_record import (
    YCCDPeriodMapRecord,
)
from repositories.yccd_period_map_writer import (
    YCCDPeriodMapWriter,
)
from repositories.yccd_repository_v2 import (
    YCCDRepositoryV2,
)
from utils.yccd_period_map_integrity import (
    validate_period_map_integrity,
)


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

BACKUP_DIR = Path(
    "output/backups"
)

# ============================================================
# CHẾ ĐỘ AN TOÀN
#
# True:
#   chỉ kiểm tra, KHÔNG ghi workbook thật.
#
# False:
#   tạo backup rồi mới ghi 5 mapping vào workbook thật.
#
# Chỉ thay đổi đúng dòng này.
# ============================================================

DRY_RUN = True


def build_map_records() -> list[
    YCCDPeriodMapRecord
]:
    common = {
        "lesson_key": "T7_DAI_B03",
        "version": "1.0",
        "status": "draft",
        "updated_at": date(
            2026,
            8,
            8,
        ),
    }

    return [
        YCCDPeriodMapRecord(
            **common,
            map_id="T7_DAI_B03_P01_Y01",
            period_in_lesson=1,
            yccd_id="T7_DAI_B03_Y01",
            role="CHINH",
        ),
        YCCDPeriodMapRecord(
            **common,
            map_id="T7_DAI_B03_P02_Y01",
            period_in_lesson=2,
            yccd_id="T7_DAI_B03_Y01",
            role="CUNG_CO",
        ),
        YCCDPeriodMapRecord(
            **common,
            map_id="T7_DAI_B03_P02_Y02",
            period_in_lesson=2,
            yccd_id="T7_DAI_B03_Y02",
            role="CHINH",
        ),
        YCCDPeriodMapRecord(
            **common,
            map_id="T7_DAI_B03_P03_Y02",
            period_in_lesson=3,
            yccd_id="T7_DAI_B03_Y02",
            role="CUNG_CO",
        ),
        YCCDPeriodMapRecord(
            **common,
            map_id="T7_DAI_B03_P03_Y03",
            period_in_lesson=3,
            yccd_id="T7_DAI_B03_Y03",
            role="CHINH",
        ),
    ]


def read_existing_map_ids() -> set[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        EXCEL_FILE,
        read_only=False,
        data_only=True,
        keep_vba=True,
    )

    try:
        if (
            "YCCD_PERIOD_MAP"
            not in workbook.sheetnames
        ):
            raise ValueError(
                "Không tìm thấy sheet YCCD_PERIOD_MAP."
            )

        worksheet = workbook[
            "YCCD_PERIOD_MAP"
        ]

        ids: set[str] = set()

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_index,
                column=1,
            ).value

            if value not in (
                None,
                "",
            ):
                ids.add(
                    str(value).strip()
                )

        return ids

    finally:
        workbook.close()


def make_backup() -> Path:
    BACKUP_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_file = (
        BACKUP_DIR
        / (
            "LBG-TUYEN_chuan_VBA_macro_"
            "BEFORE_YCCD_PERIOD_MAP_"
            f"{timestamp}.xlsm"
        )
    )

    shutil.copy2(
        EXCEL_FILE,
        backup_file,
    )

    return backup_file


def main() -> None:
    print("=" * 72)
    print(
        "LP-03D.3D.10 - "
        "SAFE FIRST YCCD PERIOD MAP IMPORT"
    )
    print("=" * 72)

    # ========================================================
    # 1. KIỂM TRA WORKBOOK
    # ========================================================

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: "
            f"{EXCEL_FILE}"
        )

    # ========================================================
    # 2. ĐỌC 4 YCCĐ THẬT
    # ========================================================

    repository = YCCDRepositoryV2()

    yccd_records = (
        repository.find_by_lesson_key(
            file_path=EXCEL_FILE,
            lesson_key="T7_DAI_B03",
            status="draft",
        )
    )

    if len(yccd_records) != 4:
        raise ValueError(
            "Phải có đúng 4 YCCĐ draft "
            "của T7_DAI_B03. "
            f"Hiện có: {len(yccd_records)}"
        )

    print(
        "- Đọc đủ 4 YCCĐ draft: PASS"
    )

    # ========================================================
    # 3. TẠO + VALIDATE 5 MAPPING
    # ========================================================

    map_records = build_map_records()

    if len(map_records) != 5:
        raise ValueError(
            "Phải có đúng 5 mapping."
        )

    for record in map_records:
        record.validate()

    print(
        "- Validate 5 mapping: PASS"
    )

    validate_period_map_integrity(
        yccd_records,
        map_records,
    )

    print(
        "- Validate mapping integrity: PASS"
    )

    # ========================================================
    # 4. KIỂM TRA TRÙNG MAP_ID
    # ========================================================

    existing_ids = (
        read_existing_map_ids()
    )

    new_ids = {
        record.map_id
        for record in map_records
    }

    duplicated_ids = (
        existing_ids
        & new_ids
    )

    if duplicated_ids:
        raise ValueError(
            "Workbook đã có các MAP_ID: "
            + ", ".join(
                sorted(
                    duplicated_ids
                )
            )
        )

    print(
        "- Không trùng MAP_ID "
        "với workbook: PASS"
    )

    print(
        "- Số mapping hiện có "
        f"trong workbook: {len(existing_ids)}"
    )

    # ========================================================
    # 5. HIỂN THỊ 5 MAPPING
    # ========================================================

    print(
        "\n5 MAPPING ĐÃ KIỂM TRA"
    )

    for record in map_records:
        print(
            f"- {record.map_id} | "
            f"Tiết {record.period_in_lesson} | "
            f"{record.yccd_id} | "
            f"{record.role} | "
            f"{record.status}"
        )

    # ========================================================
    # 6. DRY RUN
    # ========================================================

    if DRY_RUN is True:
        print(
            "\nDRY_RUN = True"
        )

        print(
            "KHÔNG ghi bất kỳ mapping nào "
            "vào workbook thật."
        )

        print(
            "\nKẾT QUẢ: "
            "PERIOD MAP DRY RUN ACCEPTED"
        )

        return

    # ========================================================
    # 7. CHẾ ĐỘ GHI THẬT
    # ========================================================

    if DRY_RUN is not False:
        raise ValueError(
            "DRY_RUN phải là True hoặc False."
        )

    print(
        "\nDRY_RUN = False"
    )

    print(
        "Bắt đầu quy trình backup "
        "và ghi dữ liệu."
    )

    # ========================================================
    # 8. BACKUP TRƯỚC KHI GHI
    # ========================================================

    backup_file = make_backup()

    if not backup_file.exists():
        raise RuntimeError(
            "Không tạo được backup."
        )

    print(
        f"- Backup: {backup_file}"
    )

    # ========================================================
    # 9. GHI 5 MAPPING
    # ========================================================

    writer = YCCDPeriodMapWriter()

    written = writer.append_records(
        EXCEL_FILE,
        map_records,
    )

    if written != 5:
        raise RuntimeError(
            "Số mapping đã ghi không đúng. "
            f"Expected=5, actual={written}"
        )

    print(
        "- Ghi 5 mapping thật: PASS"
    )

    # ========================================================
    # 10. ĐỌC LẠI ĐỂ XÁC MINH
    # ========================================================

    saved_ids = (
        read_existing_map_ids()
    )

    missing_after_write = (
        new_ids
        - saved_ids
    )

    if missing_after_write:
        raise RuntimeError(
            "Sau khi ghi còn thiếu MAP_ID: "
            + ", ".join(
                sorted(
                    missing_after_write
                )
            )
        )

    print(
        "- Đọc lại đủ 5 MAP_ID: PASS"
    )

    # ========================================================
    # KẾT QUẢ
    # ========================================================

    print(
        "\nKẾT QUẢ: "
        "REAL YCCD PERIOD MAP IMPORT ACCEPTED"
    )


if __name__ == "__main__":
    main()
