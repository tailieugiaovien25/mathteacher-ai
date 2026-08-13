from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(
    r"data\working\LBG-TUYEN_APPLY_FORWARD_TEST.xlsm"
)

SHEET_NAME = "TKB-Q"

FIRST_WEEK_ROW = 2
WEEK_HEIGHT = 60
WEEK_STEP = 67
WEEK_COUNT = 69

SOURCE_WEEK = 20

FIRST_INPUT_COL = 3   # C
LAST_INPUT_COL = 4    # D


def week_start_row(week):
    return (
        FIRST_WEEK_ROW
        + (week - 1) * WEEK_STEP
    )


def week_end_row(week):
    return (
        week_start_row(week)
        + WEEK_HEIGHT
        - 1
    )


def read_week_values(
    ws,
    week,
):
    start_row = week_start_row(
        week
    )

    end_row = week_end_row(
        week
    )

    values = []

    for row in range(
        start_row,
        end_row + 1,
    ):
        row_values = []

        for col in range(
            FIRST_INPUT_COL,
            LAST_INPUT_COL + 1,
        ):
            row_values.append(
                ws.cell(
                    row=row,
                    column=col,
                ).value
            )

        values.append(
            tuple(row_values)
        )

    return values


def main():
    print("=" * 76)
    print(
        "M5-XLS-VBA-REWRITE-01D6 - "
        "VERIFY APPLY SCHEDULE FORWARD RESULT"
    )
    print("=" * 76)

    print(
        "Chế độ: READ ONLY"
    )

    print(
        "Workbook KHÔNG bị thay đổi."
    )

    print()

    if not WORKBOOK.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: "
            f"{WORKBOOK}"
        )

    wb = load_workbook(
        WORKBOOK,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(
                f"Không có sheet: "
                f"{SHEET_NAME}"
            )

        ws = wb[
            SHEET_NAME
        ]

        source_values = (
            read_week_values(
                ws,
                SOURCE_WEEK,
            )
        )

        print(
            f"SOURCE WEEK: "
            f"{SOURCE_WEEK}"
        )

        print(
            f"SOURCE RANGE: "
            f"C{week_start_row(SOURCE_WEEK)}:"
            f"D{week_end_row(SOURCE_WEEK)}"
        )

        print()

        mismatches = []

        matched_weeks = []

        for target_week in range(
            SOURCE_WEEK + 1,
            WEEK_COUNT + 1,
        ):
            target_values = (
                read_week_values(
                    ws,
                    target_week,
                )
            )

            if (
                target_values
                == source_values
            ):
                matched_weeks.append(
                    target_week
                )

                continue

            cell_differences = []

            target_start = (
                week_start_row(
                    target_week
                )
            )

            source_start = (
                week_start_row(
                    SOURCE_WEEK
                )
            )

            for offset in range(
                WEEK_HEIGHT
            ):
                for col_offset in range(
                    2
                ):
                    source_value = (
                        source_values[
                            offset
                        ][
                            col_offset
                        ]
                    )

                    target_value = (
                        target_values[
                            offset
                        ][
                            col_offset
                        ]
                    )

                    if (
                        source_value
                        == target_value
                    ):
                        continue

                    column_letter = (
                        "C"
                        if col_offset == 0
                        else "D"
                    )

                    source_cell = (
                        f"{column_letter}"
                        f"{source_start + offset}"
                    )

                    target_cell = (
                        f"{column_letter}"
                        f"{target_start + offset}"
                    )

                    cell_differences.append(
                        {
                            "source_cell": (
                                source_cell
                            ),
                            "target_cell": (
                                target_cell
                            ),
                            "source_value": (
                                source_value
                            ),
                            "target_value": (
                                target_value
                            ),
                        }
                    )

            mismatches.append(
                {
                    "week": (
                        target_week
                    ),
                    "difference_count": (
                        len(
                            cell_differences
                        )
                    ),
                    "sample": (
                        cell_differences[:20]
                    ),
                }
            )

        print(
            "TARGET WEEKS EXPECTED:",
            WEEK_COUNT - SOURCE_WEEK,
        )

        print(
            "TARGET WEEKS MATCHED:",
            len(
                matched_weeks
            ),
        )

        print(
            "TARGET WEEKS MISMATCHED:",
            len(
                mismatches
            ),
        )

        print()

        print(
            "FIRST MATCHED WEEKS:",
            matched_weeks[:10],
        )

        print(
            "LAST MATCHED WEEKS:",
            matched_weeks[-10:],
        )

        print()

        print("=" * 76)
        print("MISMATCH DETAILS")
        print("=" * 76)

        if not mismatches:
            print(
                "Không phát hiện mismatch."
            )

        else:
            for item in mismatches[:20]:
                print(
                    f"- Week "
                    f"{item['week']:02d}: "
                    f"differences="
                    f"{item['difference_count']}"
                )

                for sample in (
                    item[
                        "sample"
                    ]
                ):
                    print(
                        "   ",
                        sample[
                            "source_cell"
                        ],
                        "=",
                        repr(
                            sample[
                                "source_value"
                            ]
                        ),
                        "->",
                        sample[
                            "target_cell"
                        ],
                        "=",
                        repr(
                            sample[
                                "target_value"
                            ]
                        ),
                    )

        print()

        if (
            len(matched_weeks) == 49
            and len(mismatches) == 0
        ):
            print(
                "VERIFY RESULT: "
                "PASS - ALL 49 TARGET WEEKS "
                "MATCH SOURCE WEEK 20"
            )
        else:
            print(
                "VERIFY RESULT: "
                "FAIL - REVIEW REQUIRED"
            )

        print()
        print(
            "Workbook KHÔNG bị thay đổi."
        )

    finally:
        wb.close()


if __name__ == "__main__":
    main()