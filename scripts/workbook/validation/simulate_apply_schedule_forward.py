from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK = Path(
    r"data\input\LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "TKB-Q"

EXPECTED_WEEK_COUNT = 69
EXPECTED_HEIGHT = 60
EXPECTED_STEP = 67

TEST_SOURCE_WEEKS = [
    1,
    4,
    20,
    68,
]


def ranges_for_column(
    ws,
    column_letter,
):
    result = []

    for dv in ws.data_validations.dataValidation:

        if dv.type != "list":
            continue

        for rng in dv.ranges.ranges:

            if rng.min_col != rng.max_col:
                continue

            actual_column = get_column_letter(
                rng.min_col
            )

            if actual_column != column_letter:
                continue

            result.append(
                {
                    "start": rng.min_row,
                    "end": rng.max_row,
                    "height": (
                        rng.max_row
                        - rng.min_row
                        + 1
                    ),
                    "range": str(rng),
                    "formula1": dv.formula1,
                }
            )

    return sorted(
        result,
        key=lambda item: item["start"],
    )


def build_week_map(ws):
    c_blocks = ranges_for_column(
        ws,
        "C",
    )

    d_blocks = ranges_for_column(
        ws,
        "D",
    )

    if len(c_blocks) != EXPECTED_WEEK_COUNT:
        raise RuntimeError(
            f"C blocks={len(c_blocks)}, "
            f"expected {EXPECTED_WEEK_COUNT}"
        )

    if len(d_blocks) != EXPECTED_WEEK_COUNT:
        raise RuntimeError(
            f"D blocks={len(d_blocks)}, "
            f"expected {EXPECTED_WEEK_COUNT}"
        )

    week_map = []

    for index in range(
        EXPECTED_WEEK_COUNT
    ):
        c_block = c_blocks[index]
        d_block = d_blocks[index]

        if (
            c_block["start"]
            != d_block["start"]
            or
            c_block["end"]
            != d_block["end"]
        ):
            raise RuntimeError(
                f"Week {index + 1}: "
                "C/D range mismatch"
            )

        if (
            c_block["height"]
            != EXPECTED_HEIGHT
            or
            d_block["height"]
            != EXPECTED_HEIGHT
        ):
            raise RuntimeError(
                f"Week {index + 1}: "
                "height mismatch"
            )

        if index > 0:
            previous = week_map[
                index - 1
            ]

            step = (
                c_block["start"]
                - previous["start_row"]
            )

            if step != EXPECTED_STEP:
                raise RuntimeError(
                    f"Week {index + 1}: "
                    f"step={step}, "
                    f"expected {EXPECTED_STEP}"
                )

        week_map.append(
            {
                "week": (
                    index + 1
                ),
                "start_row": (
                    c_block["start"]
                ),
                "end_row": (
                    c_block["end"]
                ),
                "range": (
                    f"C{c_block['start']}:"
                    f"D{c_block['end']}"
                ),
            }
        )

    return week_map


def simulate_apply_forward(
    week_map,
    source_week,
):
    if (
        source_week < 1
        or
        source_week > len(
            week_map
        )
    ):
        raise ValueError(
            f"Tuần nguồn không hợp lệ: "
            f"{source_week}"
        )

    source = week_map[
        source_week - 1
    ]

    targets = [
        item
        for item in week_map
        if item["week"] > source_week
    ]

    return {
        "source": source,
        "targets": targets,
        "target_count": len(
            targets
        ),
    }


def print_simulation(
    simulation,
):
    source = simulation[
        "source"
    ]

    targets = simulation[
        "targets"
    ]

    print()
    print("-" * 72)

    print(
        f"SOURCE WEEK: "
        f"{source['week']}"
    )

    print(
        f"SOURCE RANGE: "
        f"{source['range']}"
    )

    print(
        f"TARGET COUNT: "
        f"{simulation['target_count']}"
    )

    if targets:
        print(
            f"FIRST TARGET: "
            f"Week {targets[0]['week']} "
            f"{targets[0]['range']}"
        )

        print(
            f"LAST TARGET: "
            f"Week {targets[-1]['week']} "
            f"{targets[-1]['range']}"
        )

        print(
            "FIRST 5 TARGETS:"
        )

        for item in targets[:5]:
            print(
                f"  Week "
                f"{item['week']:02d}: "
                f"{item['range']}"
            )

    else:
        print(
            "Không có tuần đích."
        )


def main():
    print("=" * 72)

    print(
        "M5-XLS-VBA-REWRITE-01C - "
        "SIMULATE APPLY SCHEDULE FORWARD"
    )

    print("=" * 72)

    print(
        "Chế độ: READ ONLY / SIMULATION"
    )

    print(
        "Workbook KHÔNG bị thay đổi."
    )

    if not WORKBOOK.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: "
            f"{WORKBOOK}"
        )

    wb = load_workbook(
        WORKBOOK,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(
                f"Không có sheet: "
                f"{SHEET_NAME}"
            )

        ws = wb[
            SHEET_NAME
        ]

        week_map = (
            build_week_map(
                ws
            )
        )

        print()
        print(
            "Week map count:",
            len(week_map),
        )

        print(
            "First week:",
            week_map[0]["range"],
        )

        print(
            "Last week:",
            week_map[-1]["range"],
        )

        for source_week in (
            TEST_SOURCE_WEEKS
        ):
            simulation = (
                simulate_apply_forward(
                    week_map,
                    source_week,
                )
            )

            print_simulation(
                simulation
            )

        print()
        print("=" * 72)

        print(
            "EXPECTED COUNTS"
        )

        print("=" * 72)

        expected_counts = {
            1: 68,
            4: 65,
            20: 49,
            68: 1,
        }

        all_pass = True

        for source_week, expected in (
            expected_counts.items()
        ):
            simulation = (
                simulate_apply_forward(
                    week_map,
                    source_week,
                )
            )

            actual = simulation[
                "target_count"
            ]

            status = (
                "PASS"
                if actual == expected
                else "FAIL"
            )

            if actual != expected:
                all_pass = False

            print(
                f"Week "
                f"{source_week:02d}: "
                f"targets={actual}, "
                f"expected={expected} "
                f"=> {status}"
            )

        print()

        if all_pass:
            print(
                "SIMULATION RESULT: "
                "PASS - APPLY FORWARD "
                "LOGIC VERIFIED"
            )
        else:
            print(
                "SIMULATION RESULT: "
                "FAIL - DO NOT WRITE VBA"
            )

        print()
        print(
            "Workbook KHÔNG bị thay đổi."
        )

    finally:
        wb.close()


if __name__ == "__main__":
    main()