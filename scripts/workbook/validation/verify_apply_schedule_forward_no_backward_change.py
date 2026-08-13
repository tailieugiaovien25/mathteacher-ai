from pathlib import Path

from openpyxl import load_workbook


BASELINE_WORKBOOK = Path(
    r"data\working\LBG-TUYEN_CLEANUP_SAFE_WORKING.xlsm"
)

TEST_WORKBOOK = Path(
    r"data\working\LBG-TUYEN_APPLY_FORWARD_TEST.xlsm"
)

SHEET_NAME = "TKB-Q"

FIRST_WEEK_ROW = 2
WEEK_HEIGHT = 60
WEEK_STEP = 67

SOURCE_WEEK = 20

FIRST_INPUT_COL = 3   # C
LAST_INPUT_COL = 4    # D


def week_start_row(week):
    return (
        FIRST_WEEK_ROW
        + (week - 1) * WEEK_STEP
    )


def week_end_row(week):
    return (
        week_start_row(week)
        + WEEK_HEIGHT
        - 1
    )


def read_week_values(
    ws,
    week,
):
    start_row = week_start_row(
        week
    )

    end_row = week_end_row(
        week
    )

    values = []

    for row in range(
        start_row,
        end_row + 1,
    ):
        row_values = []

        for col in range(
            FIRST_INPUT_COL,
            LAST_INPUT_COL + 1,
        ):
            row_values.append(
                ws.cell(
                    row=row,
                    column=col,
                ).value
            )

        values.append(
            tuple(row_values)
        )

    return values


def main():
    print("=" * 76)
    print(
        "M5-XLS-VBA-REWRITE-01D7 - "
        "VERIFY NO BACKWARD CHANGE"
    )
    print("=" * 76)

    print(
        "Chế độ: READ ONLY"
    )

    print(
        "So sánh TEST với baseline sạch."
    )

    print()

    if not BASELINE_WORKBOOK.exists():
        raise FileNotFoundError(
            f"Không tìm thấy baseline: "
            f"{BASELINE_WORKBOOK}"
        )

    if not TEST_WORKBOOK.exists():
        raise FileNotFoundError(
            f"Không tìm thấy TEST: "
            f"{TEST_WORKBOOK}"
        )

    baseline_wb = load_workbook(
        BASELINE_WORKBOOK,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    test_wb = load_workbook(
        TEST_WORKBOOK,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:
        baseline_ws = baseline_wb[
            SHEET_NAME
        ]

        test_ws = test_wb[
            SHEET_NAME
        ]

        unchanged_weeks = []

        changed_weeks = []

        for week in range(
            1,
            SOURCE_WEEK,
        ):
            baseline_values = (
                read_week_values(
                    baseline_ws,
                    week,
                )
            )

            test_values = (
                read_week_values(
                    test_ws,
                    week,
                )
            )

            if (
                baseline_values
                == test_values
            ):
                unchanged_weeks.append(
                    week
                )

                continue

            differences = []

            start_row = week_start_row(
                week
            )

            for offset in range(
                WEEK_HEIGHT
            ):
                for col_offset in range(
                    2
                ):
                    baseline_value = (
                        baseline_values[
                            offset
                        ][
                            col_offset
                        ]
                    )

                    test_value = (
                        test_values[
                            offset
                        ][
                            col_offset
                        ]
                    )

                    if (
                        baseline_value
                        == test_value
                    ):
                        continue

                    column_letter = (
                        "C"
                        if col_offset == 0
                        else "D"
                    )

                    cell = (
                        f"{column_letter}"
                        f"{start_row + offset}"
                    )

                    differences.append(
                        {
                            "cell": cell,
                            "baseline": (
                                baseline_value
                            ),
                            "test": (
                                test_value
                            ),
                        }
                    )

            changed_weeks.append(
                {
                    "week": week,
                    "difference_count": (
                        len(
                            differences
                        )
                    ),
                    "sample": (
                        differences[:20]
                    ),
                }
            )

        print(
            "WEEKS EXPECTED UNCHANGED:",
            SOURCE_WEEK - 1,
        )

        print(
            "WEEKS UNCHANGED:",
            len(
                unchanged_weeks
            ),
        )

        print(
            "WEEKS CHANGED:",
            len(
                changed_weeks
            ),
        )

        print()

        print(
            "UNCHANGED WEEKS:",
            unchanged_weeks,
        )

        print()

        print("=" * 76)
        print("CHANGE DETAILS")
        print("=" * 76)

        if not changed_weeks:
            print(
                "Không phát hiện thay đổi "
                "trong Week 1 -> Week 19."
            )

        else:
            for item in changed_weeks:
                print(
                    f"- Week "
                    f"{item['week']:02d}: "
                    f"differences="
                    f"{item['difference_count']}"
                )

                for sample in (
                    item[
                        "sample"
                    ]
                ):
                    print(
                        "   ",
                        sample["cell"],
                        "baseline=",
                        repr(
                            sample[
                                "baseline"
                            ]
                        ),
                        "test=",
                        repr(
                            sample[
                                "test"
                            ]
                        ),
                    )

        print()

        if (
            len(unchanged_weeks) == 19
            and len(changed_weeks) == 0
        ):
            print(
                "VERIFY RESULT: "
                "PASS - WEEKS 1-19 "
                "UNCHANGED"
            )
        else:
            print(
                "VERIFY RESULT: "
                "FAIL - BACKWARD CHANGE "
                "DETECTED"
            )

        print()

        print(
            "Lưu ý: Week 20 không được "
            "so với baseline vì đã có "
            "dữ liệu test chủ động."
        )

        print()

        print(
            "Workbook KHÔNG bị thay đổi."
        )

    finally:
        baseline_wb.close()
        test_wb.close()


if __name__ == "__main__":
    main()