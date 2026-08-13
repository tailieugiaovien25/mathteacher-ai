from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK = Path(
    r"data\input\LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "TKB-Q"

EXPECTED_HEIGHT = 60
EXPECTED_STEP = 67


def ranges_for_column(ws, column_letter):
    result = []

    for dv in ws.data_validations.dataValidation:

        if dv.type != "list":
            continue

        for rng in dv.ranges.ranges:

            if rng.min_col != rng.max_col:
                continue

            actual_column = get_column_letter(
                rng.min_col
            )

            if actual_column != column_letter:
                continue

            result.append(
                {
                    "start": rng.min_row,
                    "end": rng.max_row,
                    "height": (
                        rng.max_row
                        - rng.min_row
                        + 1
                    ),
                    "range": str(rng),
                    "formula1": dv.formula1,
                }
            )

    return sorted(
        result,
        key=lambda item: item["start"],
    )


def main():

    print("=" * 72)

    print(
        "M5-XLS-VBA-REWRITE-01B2 - "
        "VERIFY TKB C:D WEEK BLOCKS"
    )

    print("=" * 72)

    print("Chế độ: READ ONLY")
    print("Workbook KHÔNG bị thay đổi.")

    print()

    if not WORKBOOK.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: "
            f"{WORKBOOK}"
        )

    wb = load_workbook(
        WORKBOOK,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:

        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(
                f"Không có sheet: "
                f"{SHEET_NAME}"
            )

        ws = wb[SHEET_NAME]

        c_blocks = ranges_for_column(
            ws,
            "C",
        )

        d_blocks = ranges_for_column(
            ws,
            "D",
        )

        print(
            "C blocks:",
            len(c_blocks),
        )

        print(
            "D blocks:",
            len(d_blocks),
        )

        print()

        anomalies = []

        week_map = []

        max_count = max(
            len(c_blocks),
            len(d_blocks),
        )

        for index in range(
            max_count
        ):

            week = index + 1

            c_block = (
                c_blocks[index]
                if index < len(c_blocks)
                else None
            )

            d_block = (
                d_blocks[index]
                if index < len(d_blocks)
                else None
            )

            if (
                c_block is None
                or d_block is None
            ):

                anomalies.append(
                    {
                        "week": week,
                        "type": (
                            "MISSING_C_OR_D_BLOCK"
                        ),
                        "c_block": c_block,
                        "d_block": d_block,
                    }
                )

                continue

            # ---------------------------------------------
            # Kiểm tra C và D cùng vùng hàng
            # ---------------------------------------------

            same_rows = (
                c_block["start"]
                == d_block["start"]
                and
                c_block["end"]
                == d_block["end"]
            )

            if not same_rows:

                anomalies.append(
                    {
                        "week": week,
                        "type": (
                            "C_D_RANGE_MISMATCH"
                        ),
                        "c_range": (
                            c_block["range"]
                        ),
                        "d_range": (
                            d_block["range"]
                        ),
                    }
                )

            # ---------------------------------------------
            # Kiểm tra chiều cao block
            # ---------------------------------------------

            if (
                c_block["height"]
                != EXPECTED_HEIGHT
                or
                d_block["height"]
                != EXPECTED_HEIGHT
            ):

                anomalies.append(
                    {
                        "week": week,
                        "type": (
                            "HEIGHT_MISMATCH"
                        ),
                        "c_height": (
                            c_block["height"]
                        ),
                        "d_height": (
                            d_block["height"]
                        ),
                        "expected": (
                            EXPECTED_HEIGHT
                        ),
                    }
                )

            # ---------------------------------------------
            # Kiểm tra bước giữa tuần
            # ---------------------------------------------

            if index > 0:

                previous_c = (
                    c_blocks[index - 1]
                )

                previous_d = (
                    d_blocks[index - 1]
                )

                c_step = (
                    c_block["start"]
                    - previous_c["start"]
                )

                d_step = (
                    d_block["start"]
                    - previous_d["start"]
                )

                if (
                    c_step != EXPECTED_STEP
                    or
                    d_step != EXPECTED_STEP
                ):

                    anomalies.append(
                        {
                            "week": week,
                            "type": (
                                "START_STEP_MISMATCH"
                            ),
                            "c_step": c_step,
                            "d_step": d_step,
                            "expected": (
                                EXPECTED_STEP
                            ),
                        }
                    )

            week_map.append(
                {
                    "week": week,
                    "start_row": (
                        c_block["start"]
                    ),
                    "end_row": (
                        c_block["end"]
                    ),
                    "range": (
                        f"C{c_block['start']}:"
                        f"D{c_block['end']}"
                    ),
                    "c_validation_source": (
                        c_block["formula1"]
                    ),
                    "d_validation_source": (
                        d_block["formula1"]
                    ),
                }
            )

        # ==================================================
        # KẾT QUẢ
        # ==================================================

        print("=" * 72)
        print("FIRST 10 WEEK BLOCKS")
        print("=" * 72)

        for item in week_map[:10]:

            print(
                f"Week {item['week']:02d}: "
                f"{item['range']}"
            )

        print()

        print("=" * 72)
        print("LAST 10 WEEK BLOCKS")
        print("=" * 72)

        for item in week_map[-10:]:

            print(
                f"Week {item['week']:02d}: "
                f"{item['range']}"
            )

        print()

        print("=" * 72)
        print("VALIDATION SOURCES")
        print("=" * 72)

        if week_map:

            print(
                "C source:",
                week_map[0][
                    "c_validation_source"
                ],
            )

            print(
                "D source:",
                week_map[0][
                    "d_validation_source"
                ],
            )

        print()

        print("=" * 72)
        print("ANOMALIES")
        print("=" * 72)

        print(
            "ANOMALIES:",
            len(anomalies),
        )

        for anomaly in anomalies[:100]:

            print(
                "-",
                anomaly,
            )

        print()

        # ==================================================
        # FINAL RESULT
        # ==================================================

        expected_same_count = (
            len(c_blocks)
            == len(d_blocks)
        )

        no_anomaly = (
            len(anomalies) == 0
        )

        if (
            expected_same_count
            and no_anomaly
        ):

            print(
                "WEEK MAP RESULT: "
                "PASS - C:D BLOCKS CONSISTENT"
            )

        else:

            print(
                "WEEK MAP RESULT: "
                "REVIEW REQUIRED"
            )

        print()

        print(
            "Tổng số week block hợp lệ:",
            len(week_map),
        )

        print()

        print(
            "Workbook KHÔNG bị thay đổi."
        )

        print()

        print(
            "KẾT QUẢ: "
            "VERIFY TKB C:D WEEK BLOCKS COMPLETE"
        )

    finally:

        wb.close()


if __name__ == "__main__":
    main()