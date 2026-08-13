import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import (
    build_lesson_id,
    extract_lesson_number,
)


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    if value is None:
        return None

    match = re.search(r"([6-9])", str(value))

    if not match:
        return None

    return int(match.group(1))


def classify_name(name: str) -> str:
    text = name.strip().lower()

    if text.startswith("bài "):
        return "BAT_DAU_BANG_BAI"

    if "luyện tập chung" in text:
        return "LUYEN_TAP_CHUNG"

    if "bài tập cuối chương" in text:
        return "BAI_TAP_CUOI_CHUONG"

    if "kiểm tra" in text:
        return "KIEM_TRA"

    if "ôn tập" in text:
        return "ON_TAP"

    if "hđtn" in text or "thực hành" in text:
        return "HD_TRẢI_NGHIEM_THUC_HANH"

    return "KHAC"


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        unresolved = []

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if not lesson_name:
                continue

            grade = detect_grade(subject_grade)

            if grade is None:
                continue

            name = str(lesson_name)

            lesson_id = build_lesson_id(
                grade,
                name,
            )

            if lesson_id is None:
                unresolved.append(
                    {
                        "row": row_number,
                        "grade": grade,
                        "period": period,
                        "name": name,
                        "category": classify_name(name),
                        "number": extract_lesson_number(name),
                    }
                )

        counts = Counter(
            item["category"]
            for item in unresolved
        )

        print("=" * 70)
        print("LP-03C.3 - PPCT KEY DIAGNOSTIC")
        print("=" * 70)

        print(
            f"Tổng unresolved: {len(unresolved)}"
        )

        print("\nPHÂN NHÓM")

        for category, count in counts.most_common():
            print(
                f"- {category}: {count}"
            )

        print(
            "\nCÁC DÒNG BẮT ĐẦU BẰNG 'BÀI' "
            "NHƯNG VẪN KHÔNG SINH ĐƯỢC ID"
        )

        special = [
            item
            for item in unresolved
            if item["category"]
            == "BAT_DAU_BANG_BAI"
        ]

        for item in special[:30]:
            print(
                f"- Row {item['row']} | "
                f"Khối {item['grade']} | "
                f"Tiết {item['period']} | "
                f"number={item['number']} | "
                f"{item['name']!r}"
            )
        print("\n30 DÒNG KHAC ĐẦU TIÊN")

        other_rows = [
            item
            for item in unresolved
            if item["category"] == "KHAC"
        ]

        for item in other_rows[:30]:
            print(
                f"- Row {item['row']} | "
                f"Khối {item['grade']} | "
                f"Tiết {item['period']} | "
                f"{item['name']!r}"
            )
        print("\nKẾT QUẢ: DIAGNOSTIC COMPLETE")

    finally:
        workbook.close()


if __name__ == "__main__":
    main()