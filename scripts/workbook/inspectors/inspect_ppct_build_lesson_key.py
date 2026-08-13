import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import build_lesson_key


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    if value is None:
        return None

    match = re.search(
        r"([6-9])",
        str(value),
    )

    if not match:
        return None

    return int(match.group(1))


def normalize_text(value) -> str:
    if value is None:
        return ""

    text = unicodedata.normalize(
        "NFC",
        str(value),
    )

    return " ".join(
        text.strip().lower().split()
    )


def remove_period_suffix(
    value: str,
) -> str:
    """Bỏ hậu tố tiết để nhận diện cùng một nội dung."""

    text = normalize_text(value)

    patterns = [
        r"\(\s*tiết\s*\d+\s*\)\s*$",
        r"\(\s*t\d+\s*\)\s*$",
    ]

    for pattern in patterns:
        text = re.sub(
            pattern,
            "",
            text,
            flags=re.IGNORECASE,
        ).strip()

    return text


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        records = []

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if (
                subject_grade is None
                or period is None
                or lesson_name is None
            ):
                continue

            grade = detect_grade(
                subject_grade
            )

            if grade is None:
                continue

            lesson_key = build_lesson_key(
                grade=grade,
                subject_grade=str(
                    subject_grade
                ),
                lesson_name=str(
                    lesson_name
                ),
                period=period,
            )

            records.append(
                {
                    "row": row_number,
                    "subject_grade": str(
                        subject_grade
                    ),
                    "grade": grade,
                    "period": period,
                    "lesson_name": str(
                        lesson_name
                    ),
                    "normalized_name": (
                        remove_period_suffix(
                            str(lesson_name)
                        )
                    ),
                    "lesson_key": lesson_key,
                }
            )

        unresolved = [
            item
            for item in records
            if item["lesson_key"] is None
        ]

        resolved_count = (
            len(records) - len(unresolved)
        )

        # =====================================================
        # PHÂN TÍCH KEY LẶP
        # =====================================================

        key_index = defaultdict(list)

        for item in records:
            lesson_key = item["lesson_key"]

            if lesson_key is None:
                continue

            key_index[lesson_key].append(
                item
            )

        duplicate_keys = {
            lesson_key: items
            for lesson_key, items
            in key_index.items()
            if len(items) > 1
        }

        repeat_ok = {}
        conflicts = {}

        for lesson_key, items in (
            duplicate_keys.items()
        ):
            names = {
                item["normalized_name"]
                for item in items
            }

            streams = {
                item["subject_grade"]
                for item in items
            }

            if (
                len(names) == 1
                and len(streams) == 1
            ):
                repeat_ok[lesson_key] = (
                    items
                )
            else:
                conflicts[lesson_key] = (
                    items
                )

        # =====================================================
        # BÁO CÁO
        # =====================================================

        print("=" * 72)
        print(
            "LP-03C.6B - "
            "BUILD LESSON KEY PPCT INSPECTION"
        )
        print("=" * 72)

        print(
            f"Tổng dòng PPCT được xử lý: "
            f"{len(records)}"
        )

        print(
            f"Tạo được LessonKey: "
            f"{resolved_count}"
        )

        print(
            f"Không tạo được LessonKey: "
            f"{len(unresolved)}"
        )

        print(
            f"Tổng LessonKey xuất hiện nhiều dòng: "
            f"{len(duplicate_keys)}"
        )

        print(
            f"Lặp hợp lệ dự kiến: "
            f"{len(repeat_ok)}"
        )

        print(
            f"Xung đột khóa thực sự: "
            f"{len(conflicts)}"
        )

        # =====================================================
        # UNRESOLVED
        # =====================================================

        if unresolved:
            print(
                "\nCÁC DÒNG KHÔNG TẠO ĐƯỢC LESSONKEY"
            )

            for item in unresolved[:30]:
                print(
                    f"- Row {item['row']} | "
                    f"{item['subject_grade']} | "
                    f"Tiết {item['period']} | "
                    f"{item['lesson_name']!r}"
                )

        # =====================================================
        # LẶP HỢP LỆ
        # =====================================================

        print(
            "\n10 LESSONKEY LẶP HỢP LỆ ĐẦU TIÊN"
        )

        for lesson_key, items in list(
            repeat_ok.items()
        )[:10]:
            periods = [
                item["period"]
                for item in items
            ]

            print(
                f"- {lesson_key} | "
                f"Tiết {periods} | "
                f"{items[0]['lesson_name']}"
            )

        # =====================================================
        # XUNG ĐỘT
        # =====================================================

        if conflicts:
            print(
                "\nCÁC XUNG ĐỘT KHÓA"
            )

            for lesson_key, items in (
                conflicts.items()
            ):
                print(
                    f"\n- KEY: {lesson_key}"
                )

                for item in items:
                    print(
                        f"    Row "
                        f"{item['row']} | "
                        f"{item['subject_grade']} | "
                        f"Tiết "
                        f"{item['period']} | "
                        f"{item['lesson_name']!r}"
                    )
        else:
            print(
                "\nKhông phát hiện "
                "xung đột khóa."
            )

        # =====================================================
        # TIÊU CHUẨN CHẤP NHẬN
        # =====================================================

        accepted = (
            len(records) == 571
            and resolved_count == 571
            and len(unresolved) == 0
            and len(conflicts) == 0
        )

        print("\nTIÊU CHUẨN CHẤP NHẬN")

        print(
            "- 571 dòng PPCT: "
            + (
                "PASS"
                if len(records) == 571
                else "FAIL"
            )
        )

        print(
            "- 571 LessonKey: "
            + (
                "PASS"
                if resolved_count == 571
                else "FAIL"
            )
        )

        print(
            "- Unresolved = 0: "
            + (
                "PASS"
                if len(unresolved) == 0
                else "FAIL"
            )
        )

        print(
            "- Conflict = 0: "
            + (
                "PASS"
                if len(conflicts) == 0
                else "FAIL"
            )
        )

        print(
            "\nKẾT QUẢ: "
            + (
                "LESSON KEY API ACCEPTED"
                if accepted
                else "LESSON KEY API NOT ACCEPTED"
            )
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()