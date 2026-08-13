import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

OUTPUT_FILE = Path(
    "output/reports/tkb_luubg_week_map_audit.json"
)

TKB_SHEET = "TKB-Q"
LUUBG_SHEET = "LuuBG"


# ------------------------------------------------------------
# Formula helpers
# ------------------------------------------------------------

LUUBG_REF_PATTERN = re.compile(
    r"(?:'LuuBG'|LuuBG)!"
    r"\$?([A-Z]{1,3})\$?(\d+)"
    r"(?:"
    r":\$?([A-Z]{1,3})\$?(\d+)"
    r")?",
    flags=re.IGNORECASE,
)


def is_formula(value):
    return (
        isinstance(value, str)
        and value.startswith("=")
    )


def formula_references_luubg(value):
    if not is_formula(value):
        return False

    text = value.lower()

    return (
        "luubg!" in text
        or "'luubg'!" in text
    )


def extract_luubg_refs(formula):
    refs = []

    for match in LUUBG_REF_PATTERN.finditer(
        formula
    ):
        start_col = match.group(1)
        start_row = int(
            match.group(2)
        )

        end_col = match.group(3)
        end_row = match.group(4)

        refs.append(
            {
                "start_col": start_col,
                "start_row": start_row,
                "end_col": end_col,
                "end_row": (
                    int(end_row)
                    if end_row
                    else None
                ),
                "raw": match.group(0),
            }
        )

    return refs


# ------------------------------------------------------------
# Group rows into clusters
# ------------------------------------------------------------

def group_rows(
    rows,
    max_gap=3,
):
    rows = sorted(
        set(rows)
    )

    if not rows:
        return []

    groups = []

    current = [
        rows[0]
    ]

    for row in rows[1:]:
        if (
            row
            - current[-1]
            <= max_gap
        ):
            current.append(
                row
            )

        else:
            groups.append(
                current
            )

            current = [
                row
            ]

    groups.append(
        current
    )

    return groups


# ------------------------------------------------------------
# Inspect cells in candidate block
# ------------------------------------------------------------

def inspect_block(
    ws,
    start_row,
    end_row,
):
    formula_count = 0
    luubg_formula_count = 0
    value_count = 0
    blank_count = 0
    unlocked_nonformula = []
    formula_columns = defaultdict(
        int
    )
    value_columns = defaultdict(
        int
    )

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            value = cell.value

            if value is None:
                blank_count += 1
                continue

            if is_formula(
                value
            ):
                formula_count += 1

                formula_columns[
                    cell.column_letter
                ] += 1

                if formula_references_luubg(
                    value
                ):
                    luubg_formula_count += 1

            else:
                value_count += 1

                value_columns[
                    cell.column_letter
                ] += 1

                try:
                    locked = (
                        cell.protection.locked
                    )
                except Exception:
                    locked = True

                if locked is False:
                    unlocked_nonformula.append(
                        cell.coordinate
                    )

    return {
        "formula_count": (
            formula_count
        ),
        "luubg_formula_count": (
            luubg_formula_count
        ),
        "value_count": (
            value_count
        ),
        "blank_count": (
            blank_count
        ),
        "formula_columns": dict(
            sorted(
                formula_columns.items()
            )
        ),
        "value_columns": dict(
            sorted(
                value_columns.items()
            )
        ),
        "unlocked_nonformula_count": (
            len(
                unlocked_nonformula
            )
        ),
        "unlocked_nonformula_sample": (
            unlocked_nonformula[:100]
        ),
    }


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------

def main():
    print("=" * 76)
    print(
        "M5-XLS-VBA-REWRITE-01A - "
        "TKB-Q <-> LuuBG WEEK MAP AUDIT"
    )
    print("=" * 76)

    print(
        "Chế độ: READ ONLY"
    )

    print(
        "Workbook KHÔNG bị thay đổi."
    )

    if not WORKBOOK_FILE.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: "
            f"{WORKBOOK_FILE}"
        )

    wb = load_workbook(
        WORKBOOK_FILE,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:
        if TKB_SHEET not in wb.sheetnames:
            raise RuntimeError(
                f"Không có sheet "
                f"{TKB_SHEET}"
            )

        if LUUBG_SHEET not in wb.sheetnames:
            raise RuntimeError(
                f"Không có sheet "
                f"{LUUBG_SHEET}"
            )

        tkb = wb[
            TKB_SHEET
        ]

        luubg = wb[
            LUUBG_SHEET
        ]

        print()
        print(
            f"TKB-Q dimension: "
            f"{tkb.calculate_dimension()}"
        )

        print(
            f"LuuBG dimension: "
            f"{luubg.calculate_dimension()}"
        )

        # ----------------------------------------------------
        # Scan all TKB-Q formulas referencing LuuBG
        # ----------------------------------------------------

        formula_records = []

        rows_with_luubg = []

        columns_with_luubg = defaultdict(
            int
        )

        referenced_luubg_rows = []

        referenced_luubg_columns = defaultdict(
            int
        )

        for row in tkb.iter_rows():
            for cell in row:
                value = cell.value

                if not formula_references_luubg(
                    value
                ):
                    continue

                refs = extract_luubg_refs(
                    value
                )

                formula_records.append(
                    {
                        "tkb_cell": (
                            cell.coordinate
                        ),
                        "tkb_row": (
                            cell.row
                        ),
                        "tkb_col": (
                            cell.column_letter
                        ),
                        "formula": value,
                        "luubg_refs": refs,
                    }
                )

                rows_with_luubg.append(
                    cell.row
                )

                columns_with_luubg[
                    cell.column_letter
                ] += 1

                for ref in refs:
                    referenced_luubg_rows.append(
                        ref[
                            "start_row"
                        ]
                    )

                    referenced_luubg_columns[
                        ref[
                            "start_col"
                        ]
                    ] += 1

                    if (
                        ref[
                            "end_row"
                        ]
                        is not None
                    ):
                        referenced_luubg_rows.append(
                            ref[
                                "end_row"
                            ]
                        )

                    if ref[
                        "end_col"
                    ]:
                        referenced_luubg_columns[
                            ref[
                                "end_col"
                            ]
                        ] += 1

        # ----------------------------------------------------
        # Candidate blocks
        # ----------------------------------------------------

        row_groups = group_rows(
            rows_with_luubg,
            max_gap=3,
        )

        block_reports = []

        for index, rows in enumerate(
            row_groups,
            start=1,
        ):
            start_row = min(
                rows
            )

            end_row = max(
                rows
            )

            block_formula_records = [
                record
                for record
                in formula_records
                if (
                    start_row
                    <= record[
                        "tkb_row"
                    ]
                    <= end_row
                )
            ]

            referenced_rows = []

            referenced_columns = defaultdict(
                int
            )

            for record in (
                block_formula_records
            ):
                for ref in record[
                    "luubg_refs"
                ]:
                    referenced_rows.append(
                        ref[
                            "start_row"
                        ]
                    )

                    referenced_columns[
                        ref[
                            "start_col"
                        ]
                    ] += 1

                    if (
                        ref[
                            "end_row"
                        ]
                        is not None
                    ):
                        referenced_rows.append(
                            ref[
                                "end_row"
                            ]
                        )

                    if ref[
                        "end_col"
                    ]:
                        referenced_columns[
                            ref[
                                "end_col"
                            ]
                        ] += 1

            block_info = inspect_block(
                tkb,
                start_row,
                end_row,
            )

            block_reports.append(
                {
                    "block_id": (
                        f"BLOCK_{index:03d}"
                    ),
                    "tkb_start_row": (
                        start_row
                    ),
                    "tkb_end_row": (
                        end_row
                    ),
                    "tkb_height": (
                        end_row
                        - start_row
                        + 1
                    ),
                    "luubg_formula_cells": (
                        len(
                            block_formula_records
                        )
                    ),
                    "tkb_columns_with_luubg_formula": (
                        sorted(
                            {
                                record[
                                    "tkb_col"
                                ]
                                for record
                                in block_formula_records
                            }
                        )
                    ),
                    "luubg_reference_min_row": (
                        min(
                            referenced_rows
                        )
                        if referenced_rows
                        else None
                    ),
                    "luubg_reference_max_row": (
                        max(
                            referenced_rows
                        )
                        if referenced_rows
                        else None
                    ),
                    "luubg_reference_columns": (
                        dict(
                            sorted(
                                referenced_columns.items()
                            )
                        )
                    ),
                    "block_statistics": (
                        block_info
                    ),
                    "formula_sample": (
                        block_formula_records[:20]
                    ),
                }
            )

        # ----------------------------------------------------
        # Detect row-distance pattern
        # ----------------------------------------------------

        block_start_rows = [
            item[
                "tkb_start_row"
            ]
            for item in block_reports
        ]

        start_row_differences = []

        for index in range(
            1,
            len(
                block_start_rows
            ),
        ):
            start_row_differences.append(
                block_start_rows[
                    index
                ]
                - block_start_rows[
                    index - 1
                ]
            )

        # ----------------------------------------------------
        # Report
        # ----------------------------------------------------

        report = {
            "audit_id": (
                "M5-XLS-VBA-REWRITE-01A"
            ),
            "mode": (
                "READ_ONLY"
            ),
            "workbook_modified": (
                False
            ),
            "workbook": str(
                WORKBOOK_FILE
            ),
            "sheets": {
                "tkb": (
                    TKB_SHEET
                ),
                "luubg": (
                    LUUBG_SHEET
                ),
            },
            "summary": {
                "tkb_max_row": (
                    tkb.max_row
                ),
                "tkb_max_column": (
                    tkb.max_column
                ),
                "luubg_max_row": (
                    luubg.max_row
                ),
                "luubg_max_column": (
                    luubg.max_column
                ),
                "luubg_formula_cell_count": (
                    len(
                        formula_records
                    )
                ),
                "rows_with_luubg_formula": (
                    len(
                        set(
                            rows_with_luubg
                        )
                    )
                ),
                "candidate_block_count": (
                    len(
                        block_reports
                    )
                ),
                "columns_with_luubg_formula": (
                    dict(
                        sorted(
                            columns_with_luubg.items()
                        )
                    )
                ),
                "luubg_reference_columns": (
                    dict(
                        sorted(
                            referenced_luubg_columns.items()
                        )
                    )
                ),
                "block_start_rows": (
                    block_start_rows
                ),
                "block_start_row_differences": (
                    start_row_differences
                ),
            },
            "candidate_blocks": (
                block_reports
            ),
            "formula_records": (
                formula_records
            ),
        }

        OUTPUT_FILE.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        OUTPUT_FILE.write_text(
            json.dumps(
                report,
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        # ----------------------------------------------------
        # Terminal summary
        # ----------------------------------------------------

        print()
        print("=" * 76)
        print(
            "KẾT QUẢ TỔNG HỢP"
        )
        print("=" * 76)

        print(
            "Số ô công thức trên TKB-Q "
            "tham chiếu LuuBG: "
            f"{len(formula_records)}"
        )

        print(
            "Số hàng TKB-Q có liên kết LuuBG: "
            f"{len(set(rows_with_luubg))}"
        )

        print(
            "Số khối liên kết phát hiện: "
            f"{len(block_reports)}"
        )

        print()
        print(
            "CỘT TKB-Q CÓ CÔNG THỨC "
            "THAM CHIẾU LuuBG"
        )

        for column, count in (
            sorted(
                columns_with_luubg.items()
            )
        ):
            print(
                f"- {column}: "
                f"{count}"
            )

        print()
        print(
            "20 KHỐI ĐẦU TIÊN"
        )

        for block in (
            block_reports[:20]
        ):
            print(
                f"- {block['block_id']} | "
                f"TKB rows "
                f"{block['tkb_start_row']}:"
                f"{block['tkb_end_row']} | "
                f"height="
                f"{block['tkb_height']} | "
                f"LuuBG formulas="
                f"{block['luubg_formula_cells']} | "
                f"LuuBG rows="
                f"{block['luubg_reference_min_row']}:"
                f"{block['luubg_reference_max_row']}"
            )

        print()
        print(
            "BLOCK START ROWS:"
        )

        print(
            block_start_rows[:100]
        )

        print()
        print(
            "KHOẢNG CÁCH GIỮA CÁC "
            "BLOCK START:"
        )

        print(
            start_row_differences[:100]
        )

        print()
        print(
            "Đã tạo báo cáo:"
        )

        print(
            OUTPUT_FILE
        )

        print()
        print(
            "Workbook KHÔNG bị thay đổi."
        )

        print()
        print(
            "KẾT QUẢ: "
            "TKB-Q / LuuBG MAP AUDIT COMPLETE"
        )

    finally:
        wb.close()


if __name__ == "__main__":
    main()