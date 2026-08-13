import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import build_lesson_key
from utils.period_mapping import (
    build_period_mapping,
    is_fallback_key,
)


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    """Lấy khối 6-9 từ giá trị Môn-lớp."""

    if value is None:
        return None

    match = re.search(
        r"([6-9])",
        str(value),
    )

    if not match:
        return None

    return int(match.group(1))


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        records = []

        # =====================================================
        # 1. ĐỌC PPCT + TẠO LESSONKEY
        # =====================================================

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if (
                subject_grade is None
                or period is None
                or lesson_name is None
            ):
                continue

            grade = detect_grade(
                subject_grade
            )

            if grade is None:
                continue

            lesson_key = build_lesson_key(
                grade=grade,
                subject_grade=str(
                    subject_grade
                ),
                lesson_name=str(
                    lesson_name
                ),
                period=period,
            )

            if lesson_key is None:
                continue

            records.append(
                {
                    "row": row_number,
                    "subject_grade": str(
                        subject_grade
                    ),
                    "grade": grade,
                    "period": period,
                    "lesson_name": str(
                        lesson_name
                    ),
                    "lesson_key": lesson_key,
                }
            )

        # =====================================================
        # 2. GỌI MODULE DÙNG CHUNG
        # =====================================================

        mapped_records = build_period_mapping(
            records
        )

        # =====================================================
        # 3. KIỂM TRA KẾT QUẢ
        # =====================================================

        missing_period_in_lesson = [
            item
            for item in mapped_records
            if item.get(
                "period_in_lesson"
            ) is None
        ]

        groups = defaultdict(list)

        for item in mapped_records:
            groups[
                item["lesson_key"]
            ].append(item)

        invalid_groups = []

        for lesson_key, items in groups.items():
            values = [
                item.get(
                    "period_in_lesson"
                )
                for item in items
            ]

            if is_fallback_key(
                lesson_key
            ):
                if any(
                    value != 1
                    for value in values
                ):
                    invalid_groups.append(
                        lesson_key
                    )

                continue

            expected = list(
                range(
                    1,
                    len(items) + 1,
                )
            )

            if sorted(values) != expected:
                invalid_groups.append(
                    lesson_key
                )

        multi_period_groups = {
            lesson_key: items
            for lesson_key, items
            in groups.items()
            if (
                not is_fallback_key(
                    lesson_key
                )
                and len(items) > 1
            )
        }

        fallback_groups = {
            lesson_key: items
            for lesson_key, items
            in groups.items()
            if is_fallback_key(
                lesson_key
            )
        }

        # =====================================================
        # 4. BÁO CÁO
        # =====================================================

        print("=" * 72)
        print(
            "LP-03D.2E - "
            "PPCT PERIOD MAPPING MODULE INSPECTION"
        )
        print("=" * 72)

        print(
            f"Tổng dòng PPCT đọc được: "
            f"{len(records)}"
        )

        print(
            f"Tổng dòng đã mapping: "
            f"{len(mapped_records)}"
        )

        print(
            f"Dòng thiếu TIET_TRONG_BAI: "
            f"{len(missing_period_in_lesson)}"
        )

        print(
            f"Tổng LessonKey: "
            f"{len(groups)}"
        )

        print(
            f"Bài có nhiều tiết: "
            f"{len(multi_period_groups)}"
        )

        print(
            f"Fallback theo tiết: "
            f"{len(fallback_groups)}"
        )

        print(
            f"LessonKey mapping sai: "
            f"{len(invalid_groups)}"
        )

        print(
            "\n10 BÀI NHIỀU TIẾT ĐẦU TIÊN"
        )

        for lesson_key, items in list(
            multi_period_groups.items()
        )[:10]:
            print(
                f"\n- {lesson_key}"
            )

            sorted_items = sorted(
                items,
                key=lambda item: (
                    item[
                        "period_in_lesson"
                    ]
                ),
            )

            for item in sorted_items:
                print(
                    f"    Tiết PPCT "
                    f"{item['period']} "
                    f"-> Tiết trong bài "
                    f"{item['period_in_lesson']} "
                    f"| "
                    f"{item['lesson_name']}"
                )

        if invalid_groups:
            print(
                "\nCÁC LESSONKEY MAPPING SAI"
            )

            for lesson_key in (
                invalid_groups
            ):
                print(
                    f"- {lesson_key}"
                )
        else:
            print(
                "\nKhông phát hiện lỗi mapping."
            )

        # =====================================================
        # 5. TIÊU CHUẨN CHẤP NHẬN
        # =====================================================

        accepted = (
            len(records) == 571
            and len(mapped_records) == 571
            and len(
                missing_period_in_lesson
            ) == 0
            and len(
                invalid_groups
            ) == 0
        )

        print(
            "\nTIÊU CHUẨN CHẤP NHẬN"
        )

        print(
            "- 571 dòng PPCT: "
            + (
                "PASS"
                if len(records) == 571
                else "FAIL"
            )
        )

        print(
            "- 571 dòng đã mapping: "
            + (
                "PASS"
                if len(
                    mapped_records
                ) == 571
                else "FAIL"
            )
        )

        print(
            "- Thiếu TIET_TRONG_BAI = 0: "
            + (
                "PASS"
                if len(
                    missing_period_in_lesson
                ) == 0
                else "FAIL"
            )
        )

        print(
            "- Mapping sai = 0: "
            + (
                "PASS"
                if len(
                    invalid_groups
                ) == 0
                else "FAIL"
            )
        )

        print(
            "\nKẾT QUẢ: "
            + (
                "PERIOD MAPPING MODULE ACCEPTED"
                if accepted
                else "PERIOD MAPPING MODULE NOT ACCEPTED"
            )
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()