from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)
SHEET_NAME = "LuuBG"

ROWS_TO_INSPECT = (81, 82, 83, 150, 151)
MAX_COLUMNS = 20


def normalize(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def main() -> None:
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Không tìm thấy file Excel: {EXCEL_FILE}"
        )

    workbook = load_workbook(
        filename=EXCEL_FILE,
        read_only=False,
        data_only=True,
        keep_vba=True,
    )

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Không tìm thấy worksheet: {SHEET_NAME}"
            )

        worksheet = workbook[SHEET_NAME]

        print("=" * 70)
        print("AI-101.4C - KHẢO SÁT VỊ TRÍ CỘT")
        print("=" * 70)

        for row_number in ROWS_TO_INSPECT:
            print(f"\nHÀNG {row_number}")
            print("-" * 70)

            for column_number in range(
                1,
                min(
                    worksheet.max_column,
                    MAX_COLUMNS,
                )
                + 1,
            ):
                value = worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value

                text = normalize(value)

                if not text:
                    continue

                column_letter = get_column_letter(
                    column_number
                )

                print(
                    f"{column_letter:>2} "
                    f"(cột {column_number:>2}): "
                    f"{text!r}"
                )

        print("\n" + "=" * 70)
        print("KHẢO SÁT HOÀN THÀNH")
        print("=" * 70)

    finally:
        workbook.close()


if __name__ == "__main__":
    main()