import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(
    r"data\working\LBG-TUYEN_CLEANUP_SAFE_WORKING.xlsm"
)

VOLATILE_FUNCTIONS = [
    "INDIRECT",
    "OFFSET",
    "NOW",
    "TODAY",
    "RAND",
    "RANDBETWEEN",
    "CELL",
    "INFO",
]

SPECIAL_FUNCTIONS = [
    "VLOOKUP",
    "HLOOKUP",
    "XLOOKUP",
    "INDEX",
    "MATCH",
    "SUMPRODUCT",
    "COUNTIF",
    "COUNTIFS",
    "SUMIF",
    "SUMIFS",
    "IFERROR",
]


def is_formula(value):
    return (
        isinstance(value, str)
        and value.startswith("=")
    )


def count_function(formula, function_name):
    pattern = (
        r"\b"
        + re.escape(function_name)
        + r"\s*\("
    )

    return len(
        re.findall(
            pattern,
            formula,
            flags=re.IGNORECASE,
        )
    )


def main():
    print("=" * 76)
    print(
        "M5-XLS-DIAG - "
        "FORMULA VOLATILITY INVENTORY"
    )
    print("=" * 76)

    print("Chế độ: READ ONLY")
    print("Workbook KHÔNG bị thay đổi.")
    print()

    if not WORKBOOK.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: {WORKBOOK}"
        )

    wb = load_workbook(
        WORKBOOK,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:
        total_formulas = 0

        formulas_by_sheet = Counter()

        volatile_counts = Counter()
        special_counts = Counter()

        volatile_by_sheet = defaultdict(
            Counter
        )

        volatile_samples = defaultdict(
            list
        )

        for ws in wb.worksheets:

            for row in ws.iter_rows():

                for cell in row:

                    value = cell.value

                    if not is_formula(value):
                        continue

                    total_formulas += 1

                    formulas_by_sheet[
                        ws.title
                    ] += 1

                    for function_name in (
                        VOLATILE_FUNCTIONS
                    ):
                        count = count_function(
                            value,
                            function_name,
                        )

                        if count == 0:
                            continue

                        volatile_counts[
                            function_name
                        ] += count

                        volatile_by_sheet[
                            ws.title
                        ][
                            function_name
                        ] += count

                        if (
                            len(
                                volatile_samples[
                                    function_name
                                ]
                            )
                            < 20
                        ):
                            volatile_samples[
                                function_name
                            ].append(
                                {
                                    "sheet": (
                                        ws.title
                                    ),
                                    "cell": (
                                        cell.coordinate
                                    ),
                                    "formula": value,
                                }
                            )

                    for function_name in (
                        SPECIAL_FUNCTIONS
                    ):
                        count = count_function(
                            value,
                            function_name,
                        )

                        if count:
                            special_counts[
                                function_name
                            ] += count

        print("FORMULA SUMMARY")
        print("-" * 76)

        print(
            "Tổng số công thức:",
            total_formulas,
        )

        print()
        print(
            "Công thức theo sheet:"
        )

        for sheet, count in (
            formulas_by_sheet.most_common()
        ):
            print(
                f"- {sheet}: {count}"
            )

        print()
        print("=" * 76)
        print("VOLATILE FUNCTIONS")
        print("=" * 76)

        volatile_total = sum(
            volatile_counts.values()
        )

        print(
            "Tổng volatile calls:",
            volatile_total,
        )

        for name in VOLATILE_FUNCTIONS:
            print(
                f"{name}: "
                f"{volatile_counts[name]}"
            )

        print()
        print("=" * 76)
        print("VOLATILE BY SHEET")
        print("=" * 76)

        if not volatile_by_sheet:
            print(
                "Không phát hiện volatile formula."
            )

        for sheet, counts in (
            volatile_by_sheet.items()
        ):
            print(
                f"- {sheet}: "
                f"{sum(counts.values())}"
            )

            for name, count in (
                counts.most_common()
            ):
                print(
                    f"    {name}: {count}"
                )

        print()
        print("=" * 76)
        print("OTHER HEAVY FUNCTIONS")
        print("=" * 76)

        for name in SPECIAL_FUNCTIONS:
            print(
                f"{name}: "
                f"{special_counts[name]}"
            )

        print()
        print("=" * 76)
        print("VOLATILE SAMPLES")
        print("=" * 76)

        for name in VOLATILE_FUNCTIONS:

            samples = (
                volatile_samples[
                    name
                ]
            )

            if not samples:
                continue

            print()
            print(
                f"{name}:"
            )

            for item in samples[:10]:
                print(
                    " ",
                    item["sheet"],
                    item["cell"],
                    "=>",
                    item["formula"],
                )

        print()
        print("=" * 76)

        if volatile_total == 0:
            print(
                "RESULT: "
                "NO VOLATILE FORMULAS DETECTED"
            )
        else:
            print(
                "RESULT: "
                "VOLATILE FORMULAS DETECTED - "
                "REVIEW REQUIRED"
            )

        print("=" * 76)

        print()
        print(
            "Workbook KHÔNG bị thay đổi."
        )

    finally:
        wb.close()


if __name__ == "__main__":
    main()