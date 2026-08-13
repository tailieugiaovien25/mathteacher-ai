import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import build_lesson_id


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    """Trích khối từ các giá trị như Đại7, Hình8, Toán9..."""

    if value is None:
        return None

    match = re.search(r"([6-9])", str(value))

    if not match:
        return None

    return int(match.group(1))


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        generated = []
        unresolved = []

        # Theo PPCT hiện tại:
        # B = Môn-lớp
        # C = Tiết
        # D = Tên bài giảng
        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if not lesson_name:
                continue

            grade = detect_grade(subject_grade)

            if grade is None:
                continue

            lesson_id = build_lesson_id(
                grade,
                str(lesson_name),
            )

            record = {
                "row": row_number,
                "grade": grade,
                "period": period,
                "lesson_name": str(lesson_name),
                "lesson_id": lesson_id,
            }

            if lesson_id:
                generated.append(record)
            else:
                unresolved.append(record)

        print("=" * 72)
        print("LP-03C.2 - PPCT LESSON KEY INSPECTION")
        print("=" * 72)

        print(
            f"Tổng dòng có BAI_ID tự động: "
            f"{len(generated)}"
        )

        print(
            f"Tổng dòng chưa sinh được BAI_ID: "
            f"{len(unresolved)}"
        )

        print("\n10 DÒNG SINH ID ĐẦU TIÊN")

        for item in generated[:10]:
            print(
                f"- Row {item['row']} | "
                f"Khối {item['grade']} | "
                f"Tiết {item['period']} | "
                f"{item['lesson_id']} | "
                f"{item['lesson_name']}"
            )

        print("\nCÁC DÒNG CHƯA SINH ĐƯỢC BAI_ID")

        for item in unresolved:
            print(
                f"- Row {item['row']} | "
                f"Khối {item['grade']} | "
                f"Tiết {item['period']} | "
                f"{item['lesson_name']}"
            )

        print("\nKẾT QUẢ: INSPECTION COMPLETE")

    finally:
        workbook.close()


if __name__ == "__main__":
    main()