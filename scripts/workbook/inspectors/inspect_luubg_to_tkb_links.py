from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(
    r"data\input\LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SOURCE_SHEET = "LuuBG"
TARGET_SHEET = "TKB-Q"


print("=" * 72)
print("M5-XLS-DIAG - LuuBG -> TKB-Q FORMULA LINKS")
print("=" * 72)

print("Chế độ: READ ONLY")
print("Workbook KHÔNG bị thay đổi.")
print()

if not WORKBOOK.exists():
    raise FileNotFoundError(
        f"Không tìm thấy workbook: {WORKBOOK}"
    )


wb = load_workbook(
    WORKBOOK,
    data_only=False,
    read_only=False,
    keep_vba=True,
    keep_links=True,
)

try:
    if SOURCE_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"Không có sheet: {SOURCE_SHEET}"
        )

    ws = wb[SOURCE_SHEET]

    matches = []

    for row in ws.iter_rows():
        for cell in row:

            value = cell.value

            if not isinstance(value, str):
                continue

            if not value.startswith("="):
                continue

            formula_lower = value.lower()

            if (
                "tkb-q!" not in formula_lower
                and "'tkb-q'!" not in formula_lower
            ):
                continue

            matches.append(
                {
                    "cell": cell.coordinate,
                    "formula": value,
                }
            )

    print(
        "LuuBG formulas -> TKB-Q =",
        len(matches),
    )

    print()
    print("FIRST 30:")

    for item in matches[:30]:
        print(
            item["cell"],
            "=>",
            item["formula"],
        )

    print()
    print("KẾT QUẢ: INSPECTION COMPLETE")

finally:
    wb.close()