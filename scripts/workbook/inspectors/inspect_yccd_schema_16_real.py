import sys
from pathlib import Path

sys.path.insert(0, "src")

from repositories.yccd_repository_v2 import (
    YCCDRepositoryV2,
)


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)


EXPECTED_HEADERS = [
    "YCCD_ID",
    "LESSON_KEY",
    "MON",
    "KHOI",
    "BAI_ID",
    "TEN_BAI",
    "YCCD_ORDER",
    "YEU_CAU_CAN_DAT",
    "LOAI_YCCD",
    "YCCD_GOC_ID",
    "NGUON",
    "THAM_CHIEU",
    "PHIEN_BAN",
    "TRANG_THAI",
    "NGAY_CAP_NHAT",
    "GHI_CHU",
]


def main() -> None:
    print("=" * 70)
    print(
        "LP-03D.3D.3C - "
        "REAL YCCD 16-COLUMN SCHEMA INSPECTION"
    )
    print("=" * 70)

    repository = YCCDRepositoryV2()

    # ---------------------------------------------------------
    # 1. Repository phải đọc được workbook thật
    # ---------------------------------------------------------

    rows = repository.load_rows(
        EXCEL_FILE
    )

    print(
        f"Tổng dòng dữ liệu YCCD: {len(rows)}"
    )

    # ---------------------------------------------------------
    # 2. Đọc trực tiếp header để kiểm tra thứ tự 16 cột
    # ---------------------------------------------------------

    from openpyxl import load_workbook

    workbook = load_workbook(
        EXCEL_FILE,
        read_only=False,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook["YCCD"]

        actual_headers = []

        for column_index in range(
            1,
            17,
        ):
            value = worksheet.cell(
                row=1,
                column=column_index,
            ).value

            actual_headers.append(
                str(value).strip()
                if value is not None
                else ""
            )

    finally:
        workbook.close()

    # ---------------------------------------------------------
    # 3. Kiểm tra
    # ---------------------------------------------------------

    assert len(actual_headers) == 16

    assert (
        actual_headers
        == EXPECTED_HEADERS
    )

    assert isinstance(
        rows,
        list,
    )

    # Bảng hiện phải chưa có YCCĐ thật.
    assert len(rows) == 0

    print("\nHEADER THỰC TẾ")

    for index, header in enumerate(
        actual_headers,
        start=1,
    ):
        print(
            f"{index:02d}. {header}"
        )

    print(
        "\nTIÊU CHUẨN CHẤP NHẬN"
    )

    print(
        "- Workbook thật đọc được: PASS"
    )
    print(
        "- Đủ 16 cột: PASS"
    )
    print(
        "- Đúng tên 16 header: PASS"
    )
    print(
        "- Đúng thứ tự 16 header: PASS"
    )
    print(
        "- LOAI_YCCD đúng vị trí: PASS"
    )
    print(
        "- YCCD_GOC_ID đúng vị trí: PASS"
    )
    print(
        "- Bảng chưa có dữ liệu thật: PASS"
    )

    print(
        "\nKẾT QUẢ: "
        "REAL YCCD 16-COLUMN SCHEMA ACCEPTED"
    )


if __name__ == "__main__":
    main()