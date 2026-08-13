import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import (
    build_fallback_lesson_id,
    build_lesson_id,
)


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    if value is None:
        return None

    match = re.search(
        r"([6-9])",
        str(value),
    )

    if not match:
        return None

    return int(match.group(1))


def normalize_text(value) -> str:
    """Chuẩn hóa để so sánh tên bài."""

    if value is None:
        return ""

    text = unicodedata.normalize(
        "NFC",
        str(value),
    )

    text = " ".join(
        text.strip().lower().split()
    )

    return text


def remove_period_suffix(
    value: str,
) -> str:
    """Bỏ các hậu tố kiểu (tiết 1), (tiết 2), (t1), (t2)."""

    text = normalize_text(value)

    patterns = [
        r"\(\s*tiết\s*\d+\s*\)\s*$",
        r"\(\s*t\d+\s*\)\s*$",
    ]

    for pattern in patterns:
        text = re.sub(
            pattern,
            "",
            text,
            flags=re.IGNORECASE,
        ).strip()

    return text


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        id_index = defaultdict(list)

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if (
                subject_grade is None
                or period is None
                or lesson_name is None
            ):
                continue

            grade = detect_grade(
                subject_grade
            )

            if grade is None:
                continue

            lesson_id = build_lesson_id(
                grade,
                str(lesson_name),
            )

            key_type = "LESSON"

            if lesson_id is None:
                lesson_id = build_fallback_lesson_id(
                    grade,
                    str(subject_grade),
                    period,
                )

                key_type = "FALLBACK"

            if lesson_id is None:
                continue

            id_index[lesson_id].append(
                {
                    "row": row_number,
                    "subject_grade": str(
                        subject_grade
                    ),
                    "grade": grade,
                    "period": period,
                    "lesson_name": str(
                        lesson_name
                    ),
                    "normalized_name": (
                        remove_period_suffix(
                            str(lesson_name)
                        )
                    ),
                    "key_type": key_type,
                }
            )

        duplicate_ids = {
            lesson_id: items
            for lesson_id, items
            in id_index.items()
            if len(items) > 1
        }

        repeat_ok = {}
        conflicts = {}

        for lesson_id, items in duplicate_ids.items():
            normalized_names = {
                item["normalized_name"]
                for item in items
            }

            subject_grades = {
                item["subject_grade"]
                for item in items
            }

            if (
                len(normalized_names) == 1
                and len(subject_grades) == 1
            ):
                repeat_ok[lesson_id] = items
            else:
                conflicts[lesson_id] = items

        print("=" * 72)
        print(
            "LP-03C.5H - "
            "DUPLICATE LESSON KEY INSPECTION"
        )
        print("=" * 72)

        print(
            f"Tổng ID xuất hiện nhiều dòng: "
            f"{len(duplicate_ids)}"
        )

        print(
            f"Lặp hợp lệ dự kiến: "
            f"{len(repeat_ok)}"
        )

        print(
            f"Xung đột cần kiểm tra: "
            f"{len(conflicts)}"
        )

        print(
            "\n10 ID LẶP HỢP LỆ ĐẦU TIÊN"
        )

        for lesson_id, items in list(
            repeat_ok.items()
        )[:10]:
            periods = [
                item["period"]
                for item in items
            ]

            print(
                f"- {lesson_id} | "
                f"Tiết {periods} | "
                f"{items[0]['lesson_name']}"
            )

        if conflicts:
            print(
                "\nCÁC ID CÓ NGUY CƠ XUNG ĐỘT"
            )

            for lesson_id, items in conflicts.items():
                print(
                    f"\n- ID: {lesson_id}"
                )

                for item in items:
                    print(
                        f"    Row {item['row']} | "
                        f"{item['subject_grade']} | "
                        f"Tiết {item['period']} | "
                        f"{item['lesson_name']!r}"
                    )
        else:
            print(
                "\nKhông phát hiện xung đột khóa."
            )

        print(
            "\nKẾT QUẢ: "
            "DUPLICATE INSPECTION COMPLETE"
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()