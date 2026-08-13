import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import (
    build_fallback_lesson_id,
    build_lesson_id,
)


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    if value is None:
        return None

    match = re.search(
        r"([6-9])",
        str(value),
    )

    if not match:
        return None

    return int(match.group(1))


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        records = []

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if (
                subject_grade is None
                or period is None
                or lesson_name is None
            ):
                continue

            grade = detect_grade(
                subject_grade
            )

            if grade is None:
                continue

            # Ưu tiên khóa theo số bài.
            lesson_id = build_lesson_id(
                grade,
                str(lesson_name),
            )

            key_type = "LESSON"

            # Nếu không có số bài thì dùng fallback.
            if lesson_id is None:
                lesson_id = build_fallback_lesson_id(
                    grade,
                    str(subject_grade),
                    period,
                )

                key_type = "FALLBACK"

            if lesson_id is None:
                key_type = "UNRESOLVED"

            records.append(
                {
                    "row": row_number,
                    "subject_grade": str(
                        subject_grade
                    ),
                    "grade": grade,
                    "period": period,
                    "lesson_name": str(
                        lesson_name
                    ),
                    "lesson_id": lesson_id,
                    "key_type": key_type,
                }
            )

        counts = Counter(
            item["key_type"]
            for item in records
        )

        unresolved = [
            item
            for item in records
            if item["lesson_id"] is None
        ]

        # Kiểm tra ID bị trùng giữa các dòng.
        id_index = {}

        for item in records:
            lesson_id = item["lesson_id"]

            if lesson_id is None:
                continue

            id_index.setdefault(
                lesson_id,
                [],
            ).append(item)

        duplicate_ids = {
            lesson_id: items
            for lesson_id, items
            in id_index.items()
            if len(items) > 1
        }

        print("=" * 72)
        print(
            "LP-03C.5F - "
            "PPCT FINAL KEY INSPECTION"
        )
        print("=" * 72)

        print(
            f"Tổng dòng kiểm tra: "
            f"{len(records)}"
        )

        print(
            f"Khóa theo Bài số: "
            f"{counts['LESSON']}"
        )

        print(
            f"Khóa fallback: "
            f"{counts['FALLBACK']}"
        )

        print(
            f"Chưa tạo được khóa: "
            f"{counts['UNRESOLVED']}"
        )

        print(
            f"Số ID xuất hiện ở nhiều dòng: "
            f"{len(duplicate_ids)}"
        )

        if unresolved:
            print(
                "\nCÁC DÒNG CHƯA TẠO ĐƯỢC KHÓA"
            )

            for item in unresolved[:30]:
                print(
                    f"- Row {item['row']} | "
                    f"{item['subject_grade']} | "
                    f"Tiết {item['period']} | "
                    f"{item['lesson_name']!r}"
                )

        print(
            "\n10 FALLBACK ID ĐẦU TIÊN"
        )

        fallback_rows = [
            item
            for item in records
            if item["key_type"] == "FALLBACK"
        ]

        for item in fallback_rows[:10]:
            print(
                f"- Row {item['row']} | "
                f"{item['lesson_id']} | "
                f"{item['subject_grade']} | "
                f"Tiết {item['period']} | "
                f"{item['lesson_name']}"
            )
        print(
            "\nCÁC GIÁ TRỊ MÔN-LỚP CHƯA ĐƯỢC NHẬN DIỆN"
        )

        unresolved_subjects = sorted(
            {
                item["subject_grade"]
                for item in unresolved
            }
        )

        for value in unresolved_subjects:
            print(f"- {value}")

        print(
            f"\nTổng loại Môn-lớp chưa nhận diện: "
            f"{len(unresolved_subjects)}"
        )
        print(
            "\nKẾT QUẢ: FINAL KEY INSPECTION COMPLETE"
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()