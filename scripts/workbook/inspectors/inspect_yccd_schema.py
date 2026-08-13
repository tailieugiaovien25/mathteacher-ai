import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "YCCD"
TABLE_NAME = "tblYCCD"


EXPECTED_HEADERS = [
    "YCCD_ID",
    "LESSON_KEY",
    "MON",
    "KHOI",
    "BAI_ID",
    "TEN_BAI",
    "YCCD_ORDER",
    "YEU_CAU_CAN_DAT",
    "NGUON",
    "THAM_CHIEU",
    "PHIEN_BAN",
    "TRANG_THAI",
    "NGAY_CAP_NHAT",
    "GHI_CHU",
]


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=False,
        data_only=True,
        keep_vba=True,
    )

    try:
        # =====================================================
        # 1. KIỂM TRA SHEET
        # =====================================================

        assert (
            SHEET_NAME in workbook.sheetnames
        ), (
            f"Không tìm thấy sheet "
            f"{SHEET_NAME}"
        )

        worksheet = workbook[
            SHEET_NAME
        ]

        # =====================================================
        # 2. ĐỌC HEADER
        # =====================================================

        actual_headers = [
            worksheet.cell(
                row=1,
                column=column_index,
            ).value
            for column_index in range(
                1,
                len(EXPECTED_HEADERS) + 1,
            )
        ]

        actual_headers = [
            (
                str(value).strip()
                if value is not None
                else None
            )
            for value in actual_headers
        ]

        # =====================================================
        # 3. KIỂM TRA TABLE
        # =====================================================

        table_names = list(
            worksheet.tables.keys()
        )

        table_exists = (
            TABLE_NAME in table_names
        )

        # =====================================================
        # 4. SO SÁNH SCHEMA
        # =====================================================

        missing_headers = [
            header
            for header in EXPECTED_HEADERS
            if header not in actual_headers
        ]

        extra_headers = [
            header
            for header in actual_headers
            if (
                header is not None
                and header
                not in EXPECTED_HEADERS
            )
        ]

        order_correct = (
            actual_headers
            == EXPECTED_HEADERS
        )

        # =====================================================
        # 5. BÁO CÁO
        # =====================================================

        print("=" * 72)
        print(
            "LP-03D.3A - "
            "YCCD SCHEMA INSPECTION"
        )
        print("=" * 72)

        print(
            f"Sheet {SHEET_NAME}: "
            f"PASS"
        )

        print(
            f"Table {TABLE_NAME}: "
            + (
                "PASS"
                if table_exists
                else "FAIL"
            )
        )

        print(
            f"Số cột mong đợi: "
            f"{len(EXPECTED_HEADERS)}"
        )

        print(
            f"Số header đọc được: "
            f"{len(actual_headers)}"
        )

        print(
            "\nHEADER THỰC TẾ"
        )

        for index, header in enumerate(
            actual_headers,
            start=1,
        ):
            print(
                f"{index:02d}. {header}"
            )

        print(
            "\nKIỂM TRA SCHEMA"
        )

        print(
            "- Đủ 14 cột: "
            + (
                "PASS"
                if (
                    len(actual_headers)
                    == 14
                )
                else "FAIL"
            )
        )

        print(
            "- Không thiếu header: "
            + (
                "PASS"
                if not missing_headers
                else "FAIL"
            )
        )

        print(
            "- Không có header lạ: "
            + (
                "PASS"
                if not extra_headers
                else "FAIL"
            )
        )

        print(
            "- Đúng thứ tự cột: "
            + (
                "PASS"
                if order_correct
                else "FAIL"
            )
        )

        if missing_headers:
            print(
                "\nHEADER BỊ THIẾU"
            )

            for header in missing_headers:
                print(
                    f"- {header}"
                )

        if extra_headers:
            print(
                "\nHEADER KHÔNG MONG ĐỢI"
            )

            for header in extra_headers:
                print(
                    f"- {header}"
                )

        accepted = (
            table_exists
            and not missing_headers
            and not extra_headers
            and order_correct
        )

        print(
            "\nKẾT QUẢ: "
            + (
                "YCCD SCHEMA ACCEPTED"
                if accepted
                else "YCCD SCHEMA NOT ACCEPTED"
            )
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()