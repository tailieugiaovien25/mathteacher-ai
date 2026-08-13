import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK = Path(
    r"data\input\LBG-TUYEN_chuan_VBA_macro.xlsm"
)

OUTPUT_FILE = Path(
    r"output\reports\tkb_input_blocks_audit.json"
)

SHEET_NAME = "TKB-Q"

INPUT_COLUMNS = ("C", "D")


def cell_has_formula(cell):
    return (
        isinstance(cell.value, str)
        and cell.value.startswith("=")
    )


def get_data_validation_ranges(ws):
    result = []

    dvs = getattr(
        ws,
        "data_validations",
        None,
    )

    if not dvs:
        return result

    for dv in dvs.dataValidation:
        ranges = []

        for cell_range in dv.ranges.ranges:
            ranges.append(
                str(cell_range)
            )

        result.append(
            {
                "type": dv.type,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "allow_blank": dv.allow_blank,
                "ranges": ranges,
            }
        )

    return result


def cell_in_data_validation(
    coordinate,
    validation_records,
):
    from openpyxl.worksheet.cell_range import CellRange

    matches = []

    for item in validation_records:
        for range_text in item["ranges"]:
            try:
                cell_range = CellRange(
                    range_text
                )
            except Exception:
                continue

            if coordinate in cell_range:
                matches.append(
                    item
                )
                break

    return matches


def row_signature(
    ws,
    row_number,
):
    values = []

    for column in range(
        1,
        min(
            ws.max_column,
            12,
        )
        + 1,
    ):
        cell = ws.cell(
            row=row_number,
            column=column,
        )

        value = cell.value

        if value is None:
            values.append("")
        else:
            values.append(
                str(value).strip()
            )

    return values


def looks_like_week_start(
    ws,
    row_number,
):
    # Dựa trên cấu trúc nhìn thấy:
    # đầu tuần thường bắt đầu một khối "Thứ hai"
    # và tiết 1 ở cột B.
    a = ws.cell(
        row=row_number,
        column=1,
    ).value

    b = ws.cell(
        row=row_number,
        column=2,
    ).value

    a_text = (
        str(a).strip().lower()
        if a is not None
        else ""
    )

    b_text = (
        str(b).strip()
        if b is not None
        else ""
    )

    return (
        "thứ hai" in a_text
        and b_text == "1"
    )


def inspect_input_range(
    ws,
    start_row,
    end_row,
    validation_records,
):
    formula_cells = []
    populated_cells = []
    validation_missing = []
    validation_present = []

    for row_number in range(
        start_row,
        end_row + 1,
    ):
        for column_letter in INPUT_COLUMNS:
            cell = ws[
                f"{column_letter}{row_number}"
            ]

            if cell_has_formula(cell):
                formula_cells.append(
                    cell.coordinate
                )

            if cell.value not in (
                None,
                "",
            ):
                populated_cells.append(
                    {
                        "cell": cell.coordinate,
                        "value": cell.value,
                    }
                )

            matches = (
                cell_in_data_validation(
                    cell.coordinate,
                    validation_records,
                )
            )

            if matches:
                validation_present.append(
                    cell.coordinate
                )
            else:
                validation_missing.append(
                    cell.coordinate
                )

    return {
        "formula_cell_count": len(
            formula_cells
        ),
        "formula_cells": (
            formula_cells[:100]
        ),
        "populated_cell_count": len(
            populated_cells
        ),
        "populated_sample": (
            populated_cells[:50]
        ),
        "validation_present_count": len(
            validation_present
        ),
        "validation_missing_count": len(
            validation_missing
        ),
        "validation_missing_sample": (
            validation_missing[:100]
        ),
    }


def main():
    print("=" * 76)
    print(
        "M5-XLS-VBA-REWRITE-01B - "
        "TKB-Q INPUT BLOCK AUDIT"
    )
    print("=" * 76)

    print(
        "Chế độ: READ ONLY"
    )

    print(
        "Workbook KHÔNG bị thay đổi."
    )

    if not WORKBOOK.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: "
            f"{WORKBOOK}"
        )

    wb = load_workbook(
        WORKBOOK,
        data_only=False,
        read_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(
                f"Không có sheet: "
                f"{SHEET_NAME}"
            )

        ws = wb[
            SHEET_NAME
        ]

        validation_records = (
            get_data_validation_ranges(
                ws
            )
        )

        print()
        print(
            f"Sheet dimension: "
            f"{ws.calculate_dimension()}"
        )

        print(
            f"Data Validation records: "
            f"{len(validation_records)}"
        )

        # ====================================================
        # Tìm các đầu tuần dựa trên cấu trúc hiển thị
        # ====================================================

        candidate_starts = []

        for row_number in range(
            1,
            ws.max_row + 1,
        ):
            if looks_like_week_start(
                ws,
                row_number,
            ):
                candidate_starts.append(
                    row_number
                )

        start_differences = []

        for index in range(
            1,
            len(candidate_starts),
        ):
            start_differences.append(
                candidate_starts[index]
                - candidate_starts[
                    index - 1
                ]
            )

        gap_counts = Counter(
            start_differences
        )

        dominant_gap = None

        if gap_counts:
            dominant_gap = (
                gap_counts.most_common(
                    1
                )[0][0]
            )

        # ====================================================
        # Tạo block
        # ====================================================

        blocks = []

        for index, start_row in enumerate(
            candidate_starts,
            start=1,
        ):
            if (
                index
                < len(
                    candidate_starts
                )
            ):
                next_start = (
                    candidate_starts[
                        index
                    ]
                )

                end_row = (
                    next_start - 1
                )
            elif dominant_gap:
                end_row = min(
                    ws.max_row,
                    start_row
                    + dominant_gap
                    - 1,
                )
            else:
                end_row = ws.max_row

            input_start = (
                f"C{start_row}"
            )

            input_end = (
                f"D{end_row}"
            )

            block_inspection = (
                inspect_input_range(
                    ws,
                    start_row,
                    end_row,
                    validation_records,
                )
            )

            blocks.append(
                {
                    "week_index_candidate": (
                        index
                    ),
                    "start_row": (
                        start_row
                    ),
                    "end_row": (
                        end_row
                    ),
                    "height": (
                        end_row
                        - start_row
                        + 1
                    ),
                    "input_range": (
                        f"{input_start}:{input_end}"
                    ),
                    "start_row_signature": (
                        row_signature(
                            ws,
                            start_row,
                        )
                    ),
                    "inspection": (
                        block_inspection
                    ),
                }
            )

        heights = [
            block["height"]
            for block in blocks
        ]

        height_counts = Counter(
            heights
        )

        dominant_height = None

        if height_counts:
            dominant_height = (
                height_counts.most_common(
                    1
                )[0][0]
            )

        anomalies = []

        for block in blocks:
            if (
                dominant_height
                and block["height"]
                != dominant_height
            ):
                anomalies.append(
                    {
                        "week": (
                            block[
                                "week_index_candidate"
                            ]
                        ),
                        "type": (
                            "BLOCK_HEIGHT_MISMATCH"
                        ),
                        "height": (
                            block[
                                "height"
                            ]
                        ),
                        "expected": (
                            dominant_height
                        ),
                    }
                )

            inspection = (
                block[
                    "inspection"
                ]
            )

            if (
                inspection[
                    "formula_cell_count"
                ]
                > 0
            ):
                anomalies.append(
                    {
                        "week": (
                            block[
                                "week_index_candidate"
                            ]
                        ),
                        "type": (
                            "FORMULA_INSIDE_INPUT_RANGE"
                        ),
                        "count": (
                            inspection[
                                "formula_cell_count"
                            ]
                        ),
                        "sample": (
                            inspection[
                                "formula_cells"
                            ]
                        ),
                    }
                )

            if (
                inspection[
                    "validation_missing_count"
                ]
                > 0
            ):
                anomalies.append(
                    {
                        "week": (
                            block[
                                "week_index_candidate"
                            ]
                        ),
                        "type": (
                            "DATA_VALIDATION_MISSING"
                        ),
                        "count": (
                            inspection[
                                "validation_missing_count"
                            ]
                        ),
                        "sample": (
                            inspection[
                                "validation_missing_sample"
                            ]
                        ),
                    }
                )

        report = {
            "audit_id": (
                "M5-XLS-VBA-REWRITE-01B"
            ),
            "mode": (
                "READ_ONLY"
            ),
            "workbook_modified": False,
            "workbook": str(
                WORKBOOK
            ),
            "sheet": (
                SHEET_NAME
            ),
            "input_columns": (
                list(
                    INPUT_COLUMNS
                )
            ),
            "summary": {
                "sheet_max_row": (
                    ws.max_row
                ),
                "sheet_max_column": (
                    ws.max_column
                ),
                "candidate_week_count": (
                    len(
                        candidate_starts
                    )
                ),
                "candidate_week_starts": (
                    candidate_starts
                ),
                "start_row_differences": (
                    start_differences
                ),
                "dominant_gap": (
                    dominant_gap
                ),
                "dominant_block_height": (
                    dominant_height
                ),
                "data_validation_record_count": (
                    len(
                        validation_records
                    )
                ),
                "anomaly_count": (
                    len(
                        anomalies
                    )
                ),
            },
            "blocks": blocks,
            "anomalies": anomalies,
            "data_validations": (
                validation_records
            ),
        }

        OUTPUT_FILE.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        OUTPUT_FILE.write_text(
            json.dumps(
                report,
                ensure_ascii=False,
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )

        print()
        print("=" * 76)
        print(
            "KẾT QUẢ TỔNG HỢP"
        )
        print("=" * 76)

        print(
            "Số đầu tuần phát hiện: "
            f"{len(candidate_starts)}"
        )

        print(
            "Dominant gap: "
            f"{dominant_gap}"
        )

        print(
            "Dominant block height: "
            f"{dominant_height}"
        )

        print()
        print(
            "20 WEEK START ĐẦU TIÊN:"
        )

        print(
            candidate_starts[:20]
        )

        print()
        print(
            "20 KHOẢNG CÁCH ĐẦU TIÊN:"
        )

        print(
            start_differences[:20]
        )

        print()
        print(
            "20 BLOCK ĐẦU TIÊN"
        )

        for block in (
            blocks[:20]
        ):
            inspection = (
                block[
                    "inspection"
                ]
            )

            print(
                f"- Week?"
                f"{block['week_index_candidate']:02d} | "
                f"rows "
                f"{block['start_row']}:"
                f"{block['end_row']} | "
                f"height="
                f"{block['height']} | "
                f"range="
                f"{block['input_range']} | "
                f"formula="
                f"{inspection['formula_cell_count']} | "
                f"DV missing="
                f"{inspection['validation_missing_count']}"
            )

        print()
        print(
            f"ANOMALIES: "
            f"{len(anomalies)}"
        )

        for anomaly in (
            anomalies[:30]
        ):
            print(
                "-",
                anomaly
            )

        print()
        print(
            "Đã tạo báo cáo:"
        )

        print(
            OUTPUT_FILE
        )

        print()
        print(
            "Workbook KHÔNG bị thay đổi."
        )

        print()
        print(
            "KẾT QUẢ: "
            "TKB INPUT BLOCK AUDIT COMPLETE"
        )

    finally:
        wb.close()


if __name__ == "__main__":
    main()