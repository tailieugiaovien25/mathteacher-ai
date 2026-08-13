import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import (
    build_fallback_lesson_id,
    build_lesson_id_v2,
)


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    if value is None:
        return None

    match = re.search(
        r"([6-9])",
        str(value),
    )

    if not match:
        return None

    return int(match.group(1))


def normalize_text(value) -> str:
    """Chuẩn hóa chuỗi để so sánh tên bài."""

    if value is None:
        return ""

    text = unicodedata.normalize(
        "NFC",
        str(value),
    )

    return " ".join(
        text.strip().lower().split()
    )


def remove_period_suffix(
    value: str,
) -> str:
    """Bỏ hậu tố tiết để so sánh cùng nội dung."""

    text = normalize_text(value)

    patterns = [
        r"\(\s*tiết\s*\d+\s*\)\s*$",
        r"\(\s*t\d+\s*\)\s*$",
    ]

    for pattern in patterns:
        text = re.sub(
            pattern,
            "",
            text,
            flags=re.IGNORECASE,
        ).strip()

    return text


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        records = []

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if (
                subject_grade is None
                or period is None
                or lesson_name is None
            ):
                continue

            grade = detect_grade(
                subject_grade
            )

            if grade is None:
                continue

            # -------------------------------------------------
            # 1. Ưu tiên LessonKey V2 cho bài có số.
            # -------------------------------------------------

            lesson_id = build_lesson_id_v2(
                grade=grade,
                subject_grade=str(
                    subject_grade
                ),
                lesson_name=str(
                    lesson_name
                ),
            )

            key_type = "LESSON_V2"

            # -------------------------------------------------
            # 2. Không có số bài -> fallback.
            # -------------------------------------------------

            if lesson_id is None:
                lesson_id = (
                    build_fallback_lesson_id(
                        grade=grade,
                        subject_grade=str(
                            subject_grade
                        ),
                        period=period,
                    )
                )

                key_type = "FALLBACK"

            if lesson_id is None:
                key_type = "UNRESOLVED"

            records.append(
                {
                    "row": row_number,
                    "subject_grade": str(
                        subject_grade
                    ),
                    "grade": grade,
                    "period": period,
                    "lesson_name": str(
                        lesson_name
                    ),
                    "normalized_name": (
                        remove_period_suffix(
                            str(lesson_name)
                        )
                    ),
                    "lesson_id": lesson_id,
                    "key_type": key_type,
                }
            )

        # =====================================================
        # THỐNG KÊ BAO PHỦ
        # =====================================================

        lesson_v2_count = sum(
            1
            for item in records
            if item["key_type"]
            == "LESSON_V2"
        )

        fallback_count = sum(
            1
            for item in records
            if item["key_type"]
            == "FALLBACK"
        )

        unresolved = [
            item
            for item in records
            if item["lesson_id"] is None
        ]

        # =====================================================
        # PHÂN TÍCH ID LẶP
        # =====================================================

        id_index = defaultdict(list)

        for item in records:
            lesson_id = item["lesson_id"]

            if lesson_id is None:
                continue

            id_index[lesson_id].append(
                item
            )

        duplicate_ids = {
            lesson_id: items
            for lesson_id, items
            in id_index.items()
            if len(items) > 1
        }

        repeat_ok = {}
        conflicts = {}

        for lesson_id, items in (
            duplicate_ids.items()
        ):
            normalized_names = {
                item["normalized_name"]
                for item in items
            }

            subject_grades = {
                item["subject_grade"]
                for item in items
            }

            if (
                len(normalized_names) == 1
                and len(subject_grades) == 1
            ):
                repeat_ok[lesson_id] = (
                    items
                )
            else:
                conflicts[lesson_id] = (
                    items
                )

        # =====================================================
        # BÁO CÁO
        # =====================================================

        print("=" * 72)
        print(
            "LP-03C.5K - "
            "PPCT LESSON KEY V2 INSPECTION"
        )
        print("=" * 72)

        print(
            f"Tổng dòng kiểm tra: "
            f"{len(records)}"
        )

        print(
            f"Khóa Lesson V2: "
            f"{lesson_v2_count}"
        )

        print(
            f"Khóa fallback: "
            f"{fallback_count}"
        )

        print(
            f"Chưa tạo được khóa: "
            f"{len(unresolved)}"
        )

        print(
            f"Tổng ID xuất hiện nhiều dòng: "
            f"{len(duplicate_ids)}"
        )

        print(
            f"Lặp hợp lệ dự kiến: "
            f"{len(repeat_ok)}"
        )

        print(
            f"Xung đột khóa cần kiểm tra: "
            f"{len(conflicts)}"
        )

        # =====================================================
        # UNRESOLVED
        # =====================================================

        if unresolved:
            print(
                "\nCÁC DÒNG CHƯA TẠO ĐƯỢC KHÓA"
            )

            for item in unresolved[:30]:
                print(
                    f"- Row {item['row']} | "
                    f"{item['subject_grade']} | "
                    f"Tiết {item['period']} | "
                    f"{item['lesson_name']!r}"
                )

        # =====================================================
        # REPEAT OK
        # =====================================================

        print(
            "\n10 ID LẶP HỢP LỆ ĐẦU TIÊN"
        )

        for lesson_id, items in list(
            repeat_ok.items()
        )[:10]:
            periods = [
                item["period"]
                for item in items
            ]

            print(
                f"- {lesson_id} | "
                f"Tiết {periods} | "
                f"{items[0]['lesson_name']}"
            )

        # =====================================================
        # CONFLICTS
        # =====================================================

        if conflicts:
            print(
                "\nCÁC XUNG ĐỘT KHÓA"
            )

            for lesson_id, items in (
                conflicts.items()
            ):
                print(
                    f"\n- ID: {lesson_id}"
                )

                for item in items:
                    print(
                        f"    Row "
                        f"{item['row']} | "
                        f"{item['subject_grade']} | "
                        f"Tiết "
                        f"{item['period']} | "
                        f"{item['lesson_name']!r}"
                    )
        else:
            print(
                "\nKhông phát hiện "
                "xung đột khóa."
            )

        print(
            "\nKẾT QUẢ: "
            "V2 INSPECTION COMPLETE"
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()