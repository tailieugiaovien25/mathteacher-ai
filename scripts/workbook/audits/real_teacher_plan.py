import os

from openpyxl import load_workbook

from educational_planning_v2 import EducationalPlanningFacade
from educational_planning_v2.adapters import (
    PPCTPlanItemAdapter,
    PPCTRow,
)
from educational_planning_v2.builders.teacher_educational_plan_builder import (
    TeacherEducationalPlanBuilder,
)
from educational_planning_v2.products import (
    TeacherPlanContext,
)


FILE_PATH = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
GRADE = 6
STREAMS = ("Đại6", "Hình6")
ACADEMIC_YEAR = "2026-2027"


# ------------------------------------------------------------
# 1. Read real Math-6 PPCT
# ------------------------------------------------------------

wb = load_workbook(
    FILE_PATH,
    read_only=True,
    data_only=True,
    keep_vba=True,
)

ws = wb["PPCT"]

rows = []

for row_number in range(1, ws.max_row + 1):
    subject_grade = ws.cell(
        row=row_number,
        column=2,
    ).value

    period = ws.cell(
        row=row_number,
        column=3,
    ).value

    lesson_name = ws.cell(
        row=row_number,
        column=4,
    ).value

    if subject_grade is None or lesson_name is None:
        continue

    subject_grade = str(subject_grade).strip()

    if subject_grade not in STREAMS:
        continue

    rows.append(
        PPCTRow(
            subject_grade=subject_grade,
            period=int(period),
            lesson_name=str(lesson_name),
        )
    )

wb.close()


# ------------------------------------------------------------
# 2. PPCT -> PlanItemDraft
# ------------------------------------------------------------

adapter = PPCTPlanItemAdapter()

drafts = adapter.adapt(
    grade=GRADE,
    rows=tuple(rows),
)


# ------------------------------------------------------------
# 3. Build real EducationalPlan
# ------------------------------------------------------------

planning = EducationalPlanningFacade()

educational_plan = planning.build_plan(
    educational_plan_id="EP-MATH6-2026-2027",
    academic_year=ACADEMIC_YEAR,
    subject="Toán",
    grade=GRADE,
    curriculum_ref="CURRICULUM-MATH-2018",
    item_drafts=drafts,
    status="DRAFT",
)


# ------------------------------------------------------------
# 4. Administrative context
# ------------------------------------------------------------

context = TeacherPlanContext(
    school_name=os.environ["WR001_SCHOOL"],
    professional_team=os.environ["WR001_TEAM"],
    teacher_name=os.environ["WR001_TEACHER"],
    academic_year=ACADEMIC_YEAR,
)


# ------------------------------------------------------------
# 5. Build TeacherEducationalPlan
# ------------------------------------------------------------

builder = TeacherEducationalPlanBuilder()

product = builder.build(
    product_id="TEP-MATH6-2026-2027",
    context=context,
    educational_plan=educational_plan,
    metadata={
        "source": "real_ppct",
        "workbook": FILE_PATH,
        "working_release": "WR-001",
    },
)


# ------------------------------------------------------------
# 6. Report
# ------------------------------------------------------------

validation = planning.validate_plan(
    product.educational_plan
)

print("=" * 72)
print("WR-001C.14 - REAL TEACHER EDUCATIONAL PLAN REPORT")
print("=" * 72)

print()
print("ADMINISTRATIVE CONTEXT")
print(f"SCHOOL                 : {product.context.school_name}")
print(f"PROFESSIONAL TEAM      : {product.context.professional_team}")
print(f"TEACHER                : {product.context.teacher_name}")
print(f"ACADEMIC YEAR          : {product.context.academic_year}")

print()
print("PRODUCT")
print(f"PRODUCT ID             : {product.product_id}")
print(
    f"DOMAIN PLAN ID         : "
    f"{product.educational_plan.educational_plan_id}"
)
print(f"SUBJECT                : {product.educational_plan.subject}")
print(f"GRADE                  : {product.educational_plan.grade}")
print(f"ITEMS                  : {len(product.educational_plan.items)}")
print(
    f"TOTAL PERIODS          : "
    f"{sum(item.periods for item in product.educational_plan.items)}"
)
print(f"OTHER DUTIES           : {len(product.other_duties)}")

print()
print("DOMAIN INTEGRITY")
print(
    "DOMAIN PLAN VALID      : "
    f"{validation.is_valid}"
)
print(
    "DOMAIN VIOLATIONS      : "
    f"{len(validation.violations)}"
)

print()
print("KHGD FIELD COVERAGE")

items = product.educational_plan.items

planned_time_count = sum(
    1
    for item in items
    if item.planned_time
)

equipment_count = sum(
    1
    for item in items
    if item.teaching_equipment
)

location_count = sum(
    1
    for item in items
    if item.teaching_location
)

print(
    f"LESSON / PERIOD DATA   : "
    f"{len(items)}/{len(items)}"
)
print(
    f"PLANNED TIME FILLED    : "
    f"{planned_time_count}/{len(items)}"
)
print(
    f"EQUIPMENT FILLED       : "
    f"{equipment_count}/{len(items)}"
)
print(
    f"LOCATION FILLED        : "
    f"{location_count}/{len(items)}"
)

print()
print("SINGLE SOURCE OF TRUTH")
print(
    "PRODUCT USES DOMAIN PLAN DIRECTLY: "
    f"{product.educational_plan is educational_plan}"
)

print()
print("=" * 72)

passed = (
    len(rows) == 140
    and len(drafts) == 100
    and len(product.educational_plan.items) == 100
    and sum(
        item.periods
        for item in product.educational_plan.items
    ) == 140
    and validation.is_valid
    and product.educational_plan is educational_plan
)

if passed:
    print(
        "RESULT: PASS - REAL TEACHER "
        "EDUCATIONAL PLAN PRODUCT VERIFIED"
    )
else:
    print(
        "RESULT: REVIEW REQUIRED - "
        "TEACHER PRODUCT NEEDS ATTENTION"
    )

print("=" * 72)
