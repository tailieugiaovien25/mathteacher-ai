from openpyxl import load_workbook


FILE_PATH = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"

wb = load_workbook(
    FILE_PATH,
    read_only=False,
    data_only=False,
    keep_vba=True,
)

print("=" * 78)
print("WR-001D.1 - REAL WEEK / SCHEDULE SOURCE AUDIT")
print("=" * 78)

print()
print("WORKBOOK SHEETS")
for index, name in enumerate(wb.sheetnames, start=1):
    print(f"{index:02d}. {name}")

keywords = (
    "PPCT",
    "TKB",
    "LBG",
    "LUUBG",
    "LƯUBG",
    "LICH",
    "LỊCH",
    "TUAN",
    "TUẦN",
)

candidate_sheets = []

for name in wb.sheetnames:
    upper = name.upper()

    if any(keyword in upper for keyword in keywords):
        candidate_sheets.append(name)

print()
print("CANDIDATE SHEETS")
if candidate_sheets:
    for name in candidate_sheets:
        print(f"- {name}")
else:
    print("NONE")

print()
print("=" * 78)
print("CANDIDATE SHEET STRUCTURES")
print("=" * 78)

for name in candidate_sheets:
    ws = wb[name]

    print()
    print("-" * 78)
    print(f"SHEET: {name}")
    print(
        f"MAX_ROW={ws.max_row} | "
        f"MAX_COLUMN={ws.max_column}"
    )
    print("-" * 78)

    max_rows = min(ws.max_row, 15)
    max_cols = min(ws.max_column, 15)

    for row in range(1, max_rows + 1):
        values = []

        for col in range(1, max_cols + 1):
            value = ws.cell(
                row=row,
                column=col,
            ).value

            if value not in (None, ""):
                values.append(
                    f"C{col}={value!r}"
                )

        if values:
            print(
                f"ROW {row:02d}: "
                + " | ".join(values)
            )

print()
print("=" * 78)
print("SEARCH WEEK-LIKE CELLS")
print("=" * 78)

matches = []

for name in wb.sheetnames:
    ws = wb[name]

    for row in range(
        1,
        min(ws.max_row, 100) + 1,
    ):
        for col in range(
            1,
            min(ws.max_column, 30) + 1,
        ):
            value = ws.cell(
                row=row,
                column=col,
            ).value

            if value is None:
                continue

            text = str(value).strip().upper()

            if (
                "TUẦN" in text
                or "TUAN" in text
                or text == "WEEK"
            ):
                matches.append(
                    (
                        name,
                        row,
                        col,
                        value,
                    )
                )

for (
    name,
    row,
    col,
    value,
) in matches[:100]:
    print(
        f"{name} | "
        f"R{row}C{col} | "
        f"{value!r}"
    )

print()
print(
    f"TOTAL WEEK-LIKE CELLS: "
    f"{len(matches)}"
)

wb.close()

print()
print("=" * 78)
print("WR-001D.1 AUDIT COMPLETE")
print("=" * 78)
