from openpyxl import load_workbook

from educational_planning_v2.adapters import (
    PPCTPlanItemAdapter,
    PPCTRow,
)
from educational_planning_v2 import EducationalPlanningFacade


FILE_PATH = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
GRADE = 6
STREAMS = ("Đại6", "Hình6")


# ------------------------------------------------------------
# 1. Read real Math-6 PPCT
# ------------------------------------------------------------

wb = load_workbook(
    FILE_PATH,
    read_only=True,
    data_only=True,
    keep_vba=True,
)

ws = wb["PPCT"]

rows = []

for row_number in range(1, ws.max_row + 1):
    subject_grade = ws.cell(
        row=row_number,
        column=2,
    ).value

    period = ws.cell(
        row=row_number,
        column=3,
    ).value

    lesson_name = ws.cell(
        row=row_number,
        column=4,
    ).value

    if subject_grade is None or lesson_name is None:
        continue

    subject_grade = str(subject_grade).strip()

    if subject_grade not in STREAMS:
        continue

    rows.append(
        PPCTRow(
            subject_grade=subject_grade,
            period=int(period),
            lesson_name=str(lesson_name),
        )
    )

wb.close()


# ------------------------------------------------------------
# 2. Adapt PPCT -> PlanItemDraft
# ------------------------------------------------------------

adapter = PPCTPlanItemAdapter()

drafts = adapter.adapt(
    grade=GRADE,
    rows=tuple(rows),
)


# ------------------------------------------------------------
# 3. Build real EducationalPlan through existing facade
# ------------------------------------------------------------

facade = EducationalPlanningFacade()

plan = facade.build_plan(
    educational_plan_id="EP-MATH6-2026-2027",
    academic_year="2026-2027",
    subject="Toán",
    grade=6,
    curriculum_ref="CURRICULUM-MATH-2018",
    item_drafts=drafts,
    status="DRAFT",
)


# ------------------------------------------------------------
# 4. Validate again through public facade
# ------------------------------------------------------------

validation = facade.validate_plan(plan)


# ------------------------------------------------------------
# 5. Report
# ------------------------------------------------------------

print("=" * 72)
print("WR-001C.11 - REAL MATH 6 EDUCATIONAL PLAN REPORT")
print("=" * 72)

print()
print("SOURCE")
print(f"PPCT ROWS              : {len(rows)}")
print(f"PLAN ITEM DRAFTS       : {len(drafts)}")

print()
print("EDUCATIONAL PLAN")
print(f"PLAN ID                : {plan.educational_plan_id}")
print(f"ACADEMIC YEAR          : {plan.academic_year}")
print(f"SUBJECT                : {plan.subject}")
print(f"GRADE                  : {plan.grade}")
print(f"ITEMS                  : {len(plan.items)}")
print(f"TOTAL PERIODS          : {sum(item.periods for item in plan.items)}")
print(f"STATUS                 : {plan.status}")

print()
print("VALIDATION")
print(f"IS VALID               : {validation.is_valid}")
print(f"VIOLATIONS             : {len(validation.violations)}")

if validation.violations:
    for violation in validation.violations[:30]:
        print(
            f"- {violation.code}: "
            f"{violation.message}"
        )

print()
print("FIRST 20 PLAN ITEMS")

for item in plan.items[:20]:
    print(
        f"{item.sequence:03d} | "
        f"PERIODS={item.periods} | "
        f"{item.title!r}"
    )

print()
print("=" * 72)

if (
    len(rows) == 140
    and len(drafts) == 100
    and len(plan.items) == 100
    and sum(item.periods for item in plan.items) == 140
    and validation.is_valid
):
    print(
        "RESULT: PASS - REAL MATH 6 "
        "EDUCATIONAL PLAN VERIFIED"
    )
else:
    print(
        "RESULT: REVIEW REQUIRED - "
        "REAL EDUCATIONAL PLAN NEEDS ATTENTION"
    )

print("=" * 72)
