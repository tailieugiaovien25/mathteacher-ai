from openpyxl import load_workbook

path = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"

wb = load_workbook(
    path,
    read_only=True,
    data_only=True,
    keep_vba=True,
)

ws = wb["PPCT"]

count = 0

for row in range(1, ws.max_row + 1):
    subject_grade = ws.cell(row=row, column=2).value
    period = ws.cell(row=row, column=3).value
    lesson_name = ws.cell(row=row, column=4).value

    if subject_grade is None or lesson_name is None:
        continue

    text = str(subject_grade).strip().lower()

    if "6" not in text:
        continue

    print(
        f"ROW={row} | "
        f"SUBJECT_GRADE={subject_grade!r} | "
        f"PERIOD={period!r} | "
        f"LESSON={lesson_name!r}"
    )

    count += 1

    if count >= 15:
        break

wb.close()

print(f"SAMPLE_ROWS={count}")
