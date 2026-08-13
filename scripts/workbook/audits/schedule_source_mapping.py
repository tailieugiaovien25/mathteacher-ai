from openpyxl import load_workbook


FILE_PATH = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"

wb_formula = load_workbook(
    FILE_PATH,
    read_only=False,
    data_only=False,
    keep_vba=True,
)

wb_value = load_workbook(
    FILE_PATH,
    read_only=False,
    data_only=True,
    keep_vba=True,
)


def show_region(sheet_name, rows, cols):
    wf = wb_formula[sheet_name]
    wv = wb_value[sheet_name]

    print()
    print("=" * 78)
    print(f"SHEET: {sheet_name}")
    print("=" * 78)

    for r in rows:
        values = []

        for c in cols:
            formula = wf.cell(r, c).value
            cached = wv.cell(r, c).value

            if formula not in (None, "") or cached not in (None, ""):
                values.append(
                    f"C{c}: FORMULA={formula!r} | VALUE={cached!r}"
                )

        if values:
            print(f"ROW {r}")
            for value in values:
                print("  " + value)


print("=" * 78)
print("WR-001D.2 - REAL SCHEDULE SOURCE MAPPING REPORT")
print("=" * 78)

# ------------------------------------------------------------
# 1. Week calendar
# ------------------------------------------------------------

show_region(
    "lichtuan",
    range(1, 15),
    range(1, 9),
)

# ------------------------------------------------------------
# 2. TKB structure
# ------------------------------------------------------------

show_region(
    "TKB-Q",
    range(1, 20),
    range(1, 18),
)

# ------------------------------------------------------------
# 3. LuuBG structure
# ------------------------------------------------------------

show_region(
    "LuuBG",
    range(1, 20),
    range(1, 23),
)

# ------------------------------------------------------------
# 4. PPCT header + first Math-6 rows
# ------------------------------------------------------------

print()
print("=" * 78)
print("PPCT MATH-6 SOURCE")
print("=" * 78)

wf = wb_formula["PPCT"]
wv = wb_value["PPCT"]

print()
print("PPCT HEADER ROWS")

for r in range(1, 4):
    values = []

    for c in range(1, min(wf.max_column, 30) + 1):
        value = wf.cell(r, c).value

        if value not in (None, ""):
            values.append(f"C{c}={value!r}")

    if values:
        print(
            f"ROW {r}: "
            + " | ".join(values)
        )

print()
print("FIRST 20 MATH-6 ROWS WITH COLUMNS 1-15")

count = 0

for r in range(1, wf.max_row + 1):
    subject_grade = wv.cell(r, 2).value

    if subject_grade is None:
        continue

    subject_grade = str(subject_grade).strip()

    if subject_grade not in ("Đại6", "Hình6"):
        continue

    values = []

    for c in range(1, 16):
        value = wv.cell(r, c).value

        if value not in (None, ""):
            values.append(f"C{c}={value!r}")

    print(
        f"ROW {r}: "
        + " | ".join(values)
    )

    count += 1

    if count >= 20:
        break


# ------------------------------------------------------------
# 5. Detect actual Math-6 schedule references
# ------------------------------------------------------------

print()
print("=" * 78)
print("MATH-6 REFERENCES IN TKB-Q")
print("=" * 78)

ws = wb_value["TKB-Q"]

matches = []

for r in range(1, ws.max_row + 1):
    for c in range(1, min(ws.max_column, 30) + 1):
        value = ws.cell(r, c).value

        if value is None:
            continue

        text = str(value).strip().upper()

        if (
            "6A" in text
            or "ĐẠI6" in text
            or "HÌNH6" in text
            or "DAI6" in text
            or "HINH6" in text
        ):
            matches.append(
                (r, c, value)
            )

for r, c, value in matches[:100]:
    print(
        f"R{r}C{c} = {value!r}"
    )

print()
print(
    f"TOTAL MATH-6 TKB REFERENCES: "
    f"{len(matches)}"
)


# ------------------------------------------------------------
# 6. Summary
# ------------------------------------------------------------

print()
print("=" * 78)
print("SOURCE CAPABILITY SUMMARY")
print("=" * 78)

print("WEEK SOURCE              : lichtuan")
print("TIMETABLE SOURCE         : TKB-Q")
print("LESSON/PERIOD SOURCE     : PPCT")
print("LESSON-SCHEDULE BRIDGE   : LuuBG")
print("EQUIPMENT CANDIDATE      : PPCT / LuuBG")
print()

wb_formula.close()
wb_value.close()

print("=" * 78)
print("WR-001D.2 SOURCE MAPPING COMPLETE")
print("=" * 78)
