import argparse
import json
import sys
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

DEFAULT_OUTPUT_FILE = Path(
    "output/reports/workbook_structure_audit.json"
)


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Audit cấu trúc workbook XLSM "
            "theo chế độ chỉ đọc."
        )
    )

    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_EXCEL_FILE,
        help=(
            "Workbook cần audit. "
            "Nếu bỏ qua sẽ dùng workbook gốc."
        ),
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help="File JSON lưu kết quả audit.",
    )

    return parser.parse_args()


def count_hidden_rows(worksheet) -> int:
    return sum(
        1
        for row_index in range(
            1,
            worksheet.max_row + 1,
        )
        if worksheet.row_dimensions[
            row_index
        ].hidden
    )


def count_hidden_columns(worksheet) -> int:
    count = 0

    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        if worksheet.column_dimensions[
            column_letter
        ].hidden:
            count += 1

    return count


def count_formulas(worksheet) -> int:
    count = 0

    for row in worksheet.iter_rows():
        for cell in row:
            if (
                isinstance(cell.value, str)
                and cell.value.startswith("=")
            ):
                count += 1

    return count


def get_defined_names(workbook) -> list[dict]:
    result = []

    for name, defined_name in (
        workbook.defined_names.items()
    ):
        result.append(
            {
                "name": name,
                "value": getattr(
                    defined_name,
                    "attr_text",
                    None,
                ),
                "hidden": getattr(
                    defined_name,
                    "hidden",
                    None,
                ),
                "local_sheet_id": getattr(
                    defined_name,
                    "localSheetId",
                    None,
                ),
            }
        )

    return result


def inspect_package(
    file_path: Path,
) -> dict:
    with zipfile.ZipFile(
        file_path,
        "r",
    ) as archive:
        names = archive.namelist()

        bad_zip_entry = archive.testzip()

        vba_size = None

        if "xl/vbaProject.bin" in names:
            vba_size = len(
                archive.read(
                    "xl/vbaProject.bin"
                )
            )

    def matching(prefix: str) -> list[str]:
        return sorted(
            name
            for name in names
            if name.startswith(prefix)
        )

    vml_files = sorted(
        name
        for name in names
        if "vmlDrawing" in name
    )

    control_files = sorted(
        name
        for name in names
        if (
            name.startswith("xl/controls/")
            or name.startswith(
                "xl/ctrlProps/"
            )
        )
    )

    activex_files = matching(
        "xl/activeX/"
    )

    drawing_files = matching(
        "xl/drawings/"
    )

    return {
        "zip_integrity_ok": (
            bad_zip_entry is None
        ),
        "bad_zip_entry": (
            bad_zip_entry
        ),
        "has_vba_project": (
            "xl/vbaProject.bin"
            in names
        ),
        "vba_project_size_bytes": (
            vba_size
        ),
        "drawing_parts_count": len(
            drawing_files
        ),
        "drawing_parts": drawing_files,
        "vml_parts_count": len(
            vml_files
        ),
        "vml_parts": vml_files,
        "control_parts_count": len(
            control_files
        ),
        "control_parts": control_files,
        "activex_parts_count": len(
            activex_files
        ),
        "activex_parts": activex_files,
    }


def inspect_sheet(
    worksheet,
) -> dict:
    tables = sorted(
        list(
            worksheet.tables.keys()
        )
    )

    if worksheet.data_validations:
        data_validation_count = len(
            worksheet
            .data_validations
            .dataValidation
        )
    else:
        data_validation_count = 0

    print_area = None

    try:
        if worksheet.print_area:
            print_area = str(
                worksheet.print_area
            )
    except Exception:
        print_area = None

    print_titles = None

    try:
        if worksheet.print_title_rows:
            print_titles = str(
                worksheet.print_title_rows
            )
    except Exception:
        print_titles = None

    return {
        "title": worksheet.title,
        "state": worksheet.sheet_state,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "dimension": (
            worksheet.calculate_dimension()
        ),
        "formula_count": (
            count_formulas(
                worksheet
            )
        ),
        "table_count": len(
            tables
        ),
        "tables": tables,
        "merged_range_count": len(
            worksheet.merged_cells.ranges
        ),
        "data_validation_count": (
            data_validation_count
        ),
        "hidden_row_count": (
            count_hidden_rows(
                worksheet
            )
        ),
        "hidden_column_count": (
            count_hidden_columns(
                worksheet
            )
        ),
        "image_count": len(
            getattr(
                worksheet,
                "_images",
                [],
            )
        ),
        "chart_count": len(
            getattr(
                worksheet,
                "_charts",
                [],
            )
        ),
        "freeze_panes": (
            str(
                worksheet.freeze_panes
            )
            if worksheet.freeze_panes
            else None
        ),
        "auto_filter": (
            worksheet.auto_filter.ref
            if worksheet.auto_filter
            else None
        ),
        "print_area": (
            print_area
        ),
        "print_title_rows": (
            print_titles
        ),
    }


def main() -> None:
    args = parse_args()

    excel_file = args.workbook
    output_file = args.output

    if not excel_file.exists():
        raise FileNotFoundError(
            f"Không tìm thấy workbook: "
            f"{excel_file}"
        )

    print("=" * 72)
    print(
        "M5-XLS-AUDIT-01 - "
        "WORKBOOK STRUCTURE AUDIT"
    )
    print("=" * 72)

    print(
        "Chế độ: READ / AUDIT ONLY"
    )

    print(
        "Workbook sẽ KHÔNG bị thay đổi."
    )

    print(
        f"Workbook đang audit: "
        f"{excel_file}"
    )

    workbook = load_workbook(
        filename=excel_file,
        read_only=False,
        data_only=False,
        keep_vba=True,
        keep_links=True,
    )

    try:
        sheet_reports = []

        for worksheet in (
            workbook.worksheets
        ):
            print(
                f"Đang kiểm kê: "
                f"{worksheet.title}"
            )

            sheet_reports.append(
                inspect_sheet(
                    worksheet
                )
            )

        package_report = (
            inspect_package(
                excel_file
            )
        )

        defined_names = (
            get_defined_names(
                workbook
            )
        )

        total_formulas = sum(
            item["formula_count"]
            for item in sheet_reports
        )

        total_tables = sum(
            item["table_count"]
            for item in sheet_reports
        )

        total_validations = sum(
            item[
                "data_validation_count"
            ]
            for item in sheet_reports
        )

        total_hidden_rows = sum(
            item["hidden_row_count"]
            for item in sheet_reports
        )

        total_hidden_columns = sum(
            item[
                "hidden_column_count"
            ]
            for item in sheet_reports
        )

        report = {
            "audit": {
                "audit_id": (
                    "M5-XLS-AUDIT-01"
                ),
                "created_at": (
                    datetime.now()
                    .isoformat(
                        timespec="seconds"
                    )
                ),
                "mode": (
                    "READ_ONLY_AUDIT"
                ),
                "workbook_modified": (
                    False
                ),
            },

            "workbook": {
                "file": str(
                    excel_file
                ),
                "file_size_bytes": (
                    excel_file
                    .stat()
                    .st_size
                ),
                "sheet_count": len(
                    workbook.sheetnames
                ),
                "sheet_names": (
                    workbook.sheetnames
                ),
                "defined_name_count": len(
                    defined_names
                ),
                "defined_names": (
                    defined_names
                ),
            },

            "summary": {
                "sheet_count": len(
                    workbook.sheetnames
                ),
                "defined_name_count": len(
                    defined_names
                ),
                "table_count": (
                    total_tables
                ),
                "formula_count": (
                    total_formulas
                ),
                "data_validation_count": (
                    total_validations
                ),
                "hidden_row_count": (
                    total_hidden_rows
                ),
                "hidden_column_count": (
                    total_hidden_columns
                ),
            },

            "sheets": (
                sheet_reports
            ),

            "package": (
                package_report
            ),
        }

        output_file.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        output_file.write_text(
            json.dumps(
                report,
                ensure_ascii=False,
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )

        print("\n" + "=" * 72)
        print("KẾT QUẢ TỔNG HỢP")
        print("=" * 72)

        print(
            f"Số worksheet: "
            f"{len(workbook.sheetnames)}"
        )

        print(
            f"Defined Names: "
            f"{len(defined_names)}"
        )

        print(
            f"Tổng Excel Tables: "
            f"{total_tables}"
        )

        print(
            f"Tổng công thức: "
            f"{total_formulas}"
        )

        print(
            f"Tổng Data Validation: "
            f"{total_validations}"
        )

        print(
            f"Tổng hàng ẩn: "
            f"{total_hidden_rows}"
        )

        print(
            f"Tổng cột ẩn: "
            f"{total_hidden_columns}"
        )

        print(
            "ZIP integrity: "
            + (
                "PASS"
                if package_report[
                    "zip_integrity_ok"
                ]
                else "FAIL"
            )
        )

        print(
            "VBA project: "
            + (
                "YES"
                if package_report[
                    "has_vba_project"
                ]
                else "NO"
            )
        )

        print(
            "VBA size: "
            f"{package_report['vba_project_size_bytes']}"
        )

        print(
            "Drawing parts: "
            f"{package_report['drawing_parts_count']}"
        )

        print(
            "VML parts: "
            f"{package_report['vml_parts_count']}"
        )

        print(
            "Control parts: "
            f"{package_report['control_parts_count']}"
        )

        print(
            "ActiveX parts: "
            f"{package_report['activex_parts_count']}"
        )

        print(
            "\nĐã tạo báo cáo:"
        )

        print(
            output_file
        )

        print(
            "\nWorkbook KHÔNG bị thay đổi."
        )

        print(
            "\nKẾT QUẢ: "
            "WORKBOOK STRUCTURE AUDIT COMPLETE"
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()