from openpyxl import load_workbook

path = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"

wb = load_workbook(
    path,
    read_only=True,
    data_only=True,
    keep_vba=True,
)

ws = wb["PPCT"]

streams = ("Đại6", "Hình6")

print("=" * 72)
print("WR-001C.7 - MATH 6 PPCT STREAM REPORT")
print("=" * 72)

for stream in streams:
    rows = []

    for row in range(1, ws.max_row + 1):
        subject_grade = ws.cell(row=row, column=2).value
        period = ws.cell(row=row, column=3).value
        lesson_name = ws.cell(row=row, column=4).value

        if subject_grade is None:
            continue

        if str(subject_grade).strip() != stream:
            continue

        if lesson_name is None:
            continue

        rows.append(
            (
                row,
                subject_grade,
                period,
                lesson_name,
            )
        )

    print()
    print(f"=== STREAM: {stream} ===")

    for (
        row,
        subject_grade,
        period,
        lesson_name,
    ) in rows[:20]:
        print(
            f"ROW={row} | "
            f"SUBJECT_GRADE={subject_grade!r} | "
            f"PERIOD={period!r} | "
            f"LESSON={lesson_name!r}"
        )

    print(f"TOTAL_{stream}={len(rows)}")

wb.close()

print()
print("=" * 72)
print("WR-001C.7 REPORT COMPLETE")
print("=" * 72)
