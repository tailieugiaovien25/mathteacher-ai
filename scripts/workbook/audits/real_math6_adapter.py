from collections import Counter

from openpyxl import load_workbook

from educational_planning_v2.adapters import (
    PPCTPlanItemAdapter,
    PPCTRow,
)
from utils.lesson_key import build_lesson_key


PATH = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
STREAMS = ("Đại6", "Hình6")
GRADE = 6


wb = load_workbook(
    PATH,
    read_only=True,
    data_only=True,
    keep_vba=True,
)

ws = wb["PPCT"]

source_rows = []

for row_number in range(1, ws.max_row + 1):
    subject_grade = ws.cell(
        row=row_number,
        column=2,
    ).value

    period = ws.cell(
        row=row_number,
        column=3,
    ).value

    lesson_name = ws.cell(
        row=row_number,
        column=4,
    ).value

    if subject_grade is None or lesson_name is None:
        continue

    subject_grade = str(subject_grade).strip()

    if subject_grade not in STREAMS:
        continue

    source_rows.append(
        PPCTRow(
            subject_grade=subject_grade,
            period=int(period),
            lesson_name=str(lesson_name),
        )
    )

wb.close()

adapter = PPCTPlanItemAdapter()

drafts = adapter.adapt(
    grade=GRADE,
    rows=tuple(source_rows),
)

lesson_keys = []

for row in source_rows:
    key = build_lesson_key(
        GRADE,
        row.subject_grade,
        row.lesson_name,
        row.period,
    )

    lesson_keys.append(key)

key_counts = Counter(lesson_keys)

invalid_keys = [
    key
    for key in lesson_keys
    if not key
]

print("=" * 72)
print("WR-001C.9 - REAL MATH 6 PPCT ADAPTER REPORT")
print("=" * 72)

print()
print("SOURCE")
print(f"TOTAL PPCT ROWS       : {len(source_rows)}")
print(
    "ĐẠI6 ROWS             : "
    f"{sum(1 for r in source_rows if r.subject_grade == 'Đại6')}"
)
print(
    "HÌNH6 ROWS            : "
    f"{sum(1 for r in source_rows if r.subject_grade == 'Hình6')}"
)

print()
print("ADAPTER")
print(f"TOTAL PLAN ITEM DRAFTS: {len(drafts)}")
print(f"TOTAL GENERATED KEYS  : {len(lesson_keys)}")
print(f"INVALID LESSON KEYS   : {len(invalid_keys)}")
print(f"UNIQUE LESSON KEYS    : {len(set(lesson_keys))}")

print()
print("FIRST 20 DRAFTS")

for index, draft in enumerate(
    drafts[:20],
    start=1,
):
    print(
        f"{index:03d} | "
        f"PERIODS={draft.periods} | "
        f"{draft.title!r}"
    )

print()
print("MULTI-PERIOD LESSON KEYS")

multi_period = [
    (key, count)
    for key, count in key_counts.items()
    if key and count > 1
]

for key, count in multi_period[:20]:
    print(
        f"{key} | "
        f"PPCT_ROWS={count}"
    )

print()
print(f"MULTI-PERIOD KEY COUNT: {len(multi_period)}")

print()
print("=" * 72)

if (
    len(source_rows) == 140
    and len(invalid_keys) == 0
    and len(drafts) > 0
):
    print(
        "RESULT: PASS - REAL MATH 6 PPCT "
        "ADAPTER RUN VERIFIED"
    )
else:
    print(
        "RESULT: REVIEW REQUIRED - "
        "REAL PPCT DATA NEEDS ATTENTION"
    )

print("=" * 72)
