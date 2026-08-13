from collections import Counter, defaultdict

from openpyxl import load_workbook

from curriculum_v2.canonical_curriculum import (
    get_canonical_curriculum,
)
from repositories.yccd_period_map_repository import (
    YCCDPeriodMapRepository,
)
from utils.lesson_key import build_lesson_key


FILE_PATH = r"data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
GRADE = 6
STREAMS = ("Đại6", "Hình6")


# ------------------------------------------------------------
# 1. Build real Math-6 lesson keys from PPCT
# ------------------------------------------------------------

wb = load_workbook(
    FILE_PATH,
    read_only=True,
    data_only=True,
    keep_vba=True,
)

ws = wb["PPCT"]

lesson_keys = []
seen = set()

for row in range(1, ws.max_row + 1):
    subject_grade = ws.cell(row=row, column=2).value
    period = ws.cell(row=row, column=3).value
    lesson_name = ws.cell(row=row, column=4).value

    if subject_grade is None or lesson_name is None:
        continue

    subject_grade = str(subject_grade).strip()

    if subject_grade not in STREAMS:
        continue

    key = build_lesson_key(
        GRADE,
        subject_grade,
        str(lesson_name),
        period,
    )

    if key and key not in seen:
        seen.add(key)
        lesson_keys.append(key)

wb.close()


# ------------------------------------------------------------
# 2. Read YCCD_PERIOD_MAP through existing repository
# ------------------------------------------------------------

repository = YCCDPeriodMapRepository()

rows = repository.load_rows(FILE_PATH)

mapping_by_key = defaultdict(list)
status_counter = Counter()

for row in rows:
    key = str(row.get("LESSON_KEY") or "").strip()
    status = str(row.get("TRANG_THAI") or "").strip().lower()

    if not key:
        continue

    mapping_by_key[key].append(row)

    if status:
        status_counter[status] += 1


# ------------------------------------------------------------
# 3. Compare PPCT lesson keys with mapping
# ------------------------------------------------------------

mapped_keys = []
missing_keys = []

for key in lesson_keys:
    if mapping_by_key.get(key):
        mapped_keys.append(key)
    else:
        missing_keys.append(key)


# ------------------------------------------------------------
# 4. Validate mapped YCCD IDs against canonical curriculum
# ------------------------------------------------------------

canonical = get_canonical_curriculum()

all_yccd_ids = []
invalid_yccd_ids = []
valid_yccd_ids = []

for key in mapped_keys:
    for row in mapping_by_key[key]:
        yccd_id = str(row.get("YCCD_ID") or "").strip()

        if not yccd_id:
            continue

        all_yccd_ids.append(yccd_id)

        requirement = canonical.requirement_by_id(yccd_id)

        if requirement is None:
            invalid_yccd_ids.append(
                (key, yccd_id)
            )
        else:
            valid_yccd_ids.append(
                (key, yccd_id)
            )


# ------------------------------------------------------------
# 5. Report
# ------------------------------------------------------------

print("=" * 72)
print("WR-001C.10C - REAL MATH 6 YCCD COVERAGE REPORT")
print("=" * 72)

print()
print("PPCT")
print(f"TOTAL UNIQUE LESSON KEYS : {len(lesson_keys)}")

print()
print("YCCD PERIOD MAP")
print(f"MAPPED LESSON KEYS       : {len(mapped_keys)}")
print(f"MISSING LESSON KEYS      : {len(missing_keys)}")
print(f"TOTAL MAPPING ROWS        : {sum(len(mapping_by_key[k]) for k in mapped_keys)}")

print()
print("MAPPING STATUS")
for status, count in sorted(status_counter.items()):
    print(f"{status:<24}: {count}")

print()
print("CANONICAL YCCD VALIDATION")
print(f"YCCD REFERENCES CHECKED  : {len(all_yccd_ids)}")
print(f"VALID YCCD REFERENCES    : {len(valid_yccd_ids)}")
print(f"INVALID YCCD REFERENCES  : {len(invalid_yccd_ids)}")
print(
    f"UNIQUE VALID YCCD IDs    : "
    f"{len(set(y for _, y in valid_yccd_ids))}"
)

print()
print("FIRST 30 MISSING LESSON KEYS")

if missing_keys:
    for key in missing_keys[:30]:
        print(key)
else:
    print("NONE")

print()
print("INVALID CANONICAL REFERENCES")

if invalid_yccd_ids:
    for key, yccd_id in invalid_yccd_ids[:30]:
        print(
            f"{key} -> {yccd_id}"
        )
else:
    print("NONE")

print()
print("=" * 72)

if (
    len(lesson_keys) == 100
    and len(invalid_yccd_ids) == 0
):
    print(
        "RESULT: PASS - REAL MATH 6 YCCD "
        "COVERAGE AUDIT COMPLETED"
    )
else:
    print(
        "RESULT: REVIEW REQUIRED - "
        "YCCD COVERAGE NEEDS ATTENTION"
    )

print("=" * 72)
