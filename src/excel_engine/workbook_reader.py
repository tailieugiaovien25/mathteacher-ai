from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class WorkbookReader:
    """Chỉ chịu trách nhiệm đọc thông tin của Workbook."""

    def load(self, file_path: str):
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy file: {file_path}"
            )

        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
            keep_vba=path.suffix.lower() == ".xlsm",
        )

        return workbook

    def get_info(self, workbook) -> dict[str, Any]:
        return {
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
        }

    def close(self, workbook):
        workbook.close()