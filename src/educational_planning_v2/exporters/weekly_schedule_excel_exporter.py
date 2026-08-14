from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from educational_planning_v2.models import WeeklyTeachingSchedule


@dataclass(frozen=True)
class WeeklyScheduleExcelExport:
    file_name: str
    content: bytes
    mime_type: str = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


class WeeklyScheduleExcelExporter:
    """Render a canonical weekly schedule as a teacher-facing workbook."""

    _HEADERS = (
        "Ngày dạy",
        "Thứ",
        "Tiết TKB",
        "Lớp",
        "Môn học",
        "Phân môn",
        "Tiết PPCT",
        "Mã bài",
        "Tên bài học",
        "Tiết trong bài",
        "Thiết bị dạy học",
    )
    _WEEKDAY_LABELS = {
        1: "Thứ 2",
        2: "Thứ 3",
        3: "Thứ 4",
        4: "Thứ 5",
        5: "Thứ 6",
        6: "Thứ 7",
        7: "Chủ nhật",
    }

    def export(self, schedule: WeeklyTeachingSchedule) -> WeeklyScheduleExcelExport:
        if not isinstance(schedule, WeeklyTeachingSchedule):
            raise TypeError("schedule must be a WeeklyTeachingSchedule")
        if not schedule.entries:
            raise ValueError("Không thể xuất lịch không có tiết dạy.")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Lich_bao_giang"
        self._write_heading(sheet, schedule)
        self._write_table(sheet, schedule)
        self._write_summary(sheet, schedule)
        self._configure_sheet(sheet, len(schedule.entries))

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return WeeklyScheduleExcelExport(
            file_name=self._file_name(schedule),
            content=output.getvalue(),
        )

    def _write_heading(self, sheet, schedule):
        week = schedule.academic_week
        profile = schedule.metadata.get("teacher_profile", {})
        sheet.merge_cells("A1:K1")
        sheet["A1"] = f"LỊCH BÁO GIẢNG TUẦN {week.week_number}"
        sheet["A1"].font = Font(name="Times New Roman", size=16, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="16324F")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 30

        sheet.merge_cells("A2:K2")
        teacher_value = schedule.teacher_id
        if profile.get("show_teacher_name") and profile.get("full_name"):
            teacher_value = f"{profile['full_name']} ({schedule.teacher_id})"
        heading_parts = [
            f"Giáo viên: {teacher_value}",
            f"Năm học: {week.academic_year}",
        ]
        if profile.get("show_school_name") and profile.get("school_name"):
            heading_parts.insert(1, f"Trường: {profile['school_name']}")
        sheet["A2"] = "  |  ".join(heading_parts)
        sheet["A2"].font = Font(name="Times New Roman", size=12, bold=True)
        sheet["A2"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("A3:K3")
        sheet["A3"] = (
            f"Từ ngày {week.start_date.strftime('%d/%m/%Y')} "
            f"đến ngày {week.end_date.strftime('%d/%m/%Y')}"
        )
        sheet["A3"].font = Font(name="Times New Roman", size=11, italic=True)
        sheet["A3"].alignment = Alignment(horizontal="center")

    def _write_table(self, sheet, schedule):
        header_row = 5
        data_start = header_row + 1
        header_fill = PatternFill("solid", fgColor="2364AA")
        header_font = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        light_side = Side(style="thin", color="C8D6E5")
        cell_border = Border(
            left=light_side,
            right=light_side,
            top=light_side,
            bottom=light_side,
        )

        for column, header in enumerate(self._HEADERS, start=1):
            cell = sheet.cell(row=header_row, column=column, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = cell_border
        sheet.row_dimensions[header_row].height = 32

        for row_number, entry in enumerate(schedule.entries, start=data_start):
            values = (
                entry.teaching_date,
                self._WEEKDAY_LABELS[entry.weekday],
                entry.timetable_period,
                entry.class_id,
                entry.subject_ref,
                entry.component_ref or "",
                entry.curriculum_period,
                entry.lesson_id,
                entry.lesson_title,
                f"{entry.period_in_lesson}/{entry.total_lesson_periods}",
                "; ".join(entry.teaching_equipment),
            )
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_number, column=column, value=value)
                cell.font = Font(name="Times New Roman", size=11)
                cell.border = cell_border
                cell.alignment = Alignment(
                    horizontal="center" if column not in (9, 11) else "left",
                    vertical="center",
                    wrap_text=column in (9, 11),
                )
            sheet.cell(row=row_number, column=1).number_format = "dd/mm/yyyy"
            sheet.row_dimensions[row_number].height = 30

        end_row = data_start + len(schedule.entries) - 1
        table = Table(displayName="WeeklyScheduleTable", ref=f"A5:K{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    @staticmethod
    def _write_summary(sheet, schedule):
        summary_row = 7 + len(schedule.entries)
        classes = {entry.class_id for entry in schedule.entries}
        subjects = {
            (entry.subject_ref, entry.component_ref or "")
            for entry in schedule.entries
        }
        sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=11)
        cell = sheet.cell(row=summary_row, column=1)
        cell.value = (
            f"Tổng số tiết: {len(schedule.entries)}  |  "
            f"Số lớp: {len(classes)}  |  "
            f"Số môn/phân môn: {len(subjects)}"
        )
        cell.fill = PatternFill("solid", fgColor="EAF2F8")
        cell.font = Font(name="Times New Roman", size=11, bold=True, color="16324F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[summary_row].height = 26

    @staticmethod
    def _configure_sheet(sheet, entry_count):
        widths = {
            "A": 13,
            "B": 11,
            "C": 11,
            "D": 10,
            "E": 16,
            "F": 16,
            "G": 12,
            "H": 16,
            "I": 34,
            "J": 14,
            "K": 30,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A6"
        sheet.sheet_view.showGridLines = False
        sheet.auto_filter.ref = f"A5:K{5 + entry_count}"
        sheet.print_title_rows = "1:5"
        sheet.print_area = f"A1:K{7 + entry_count}"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.oddFooter.center.text = "Trang &P / &N"
        sheet.oddFooter.center.size = 10
        sheet.oddFooter.center.font = "Times New Roman"
        sheet.page_margins.left = 0.3
        sheet.page_margins.right = 0.3
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

    @staticmethod
    def _file_name(schedule):
        safe_teacher = re.sub(r"[^A-Za-z0-9_-]+", "-", schedule.teacher_id).strip("-")
        safe_year = re.sub(r"[^A-Za-z0-9_-]+", "-", schedule.academic_week.academic_year).strip("-")
        return (
            f"lich-bao-giang-{safe_teacher or 'giao-vien'}-"
            f"{safe_year or 'nam-hoc'}-tuan-{schedule.academic_week.week_number:02d}.xlsx"
        )
