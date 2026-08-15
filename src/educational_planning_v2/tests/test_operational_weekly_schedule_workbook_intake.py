from __future__ import annotations

from io import BytesIO
import inspect

from openpyxl import Workbook

from educational_planning_v2.adapters.operational_weekly_schedule_workbook_intake import (
    WeeklyScheduleWorkbookIntakeAdapter,
    WeeklyScheduleWorkbookIntakeResult,
)
from educational_planning_v2.adapters.weekly_schedule_excel_adapter import (
    WeeklyScheduleSourceData,
    WeeklyScheduleWorkbookError,
    WeeklyScheduleWorkbookSchema,
)
from educational_planning_v2.models.operational_data_io import (
    OperationalInputLocation,
    OperationalInputReference,
)
from educational_planning_v2.models.operational_data_source import (
    OperationalDataOrigin,
    OperationalDataSource,
    OperationalDataStatus,
    OperationalDataType,
)
from educational_planning_v2.services.operational_input_selection_service import (
    OperationalInputSelection,
)


def _write_table(
    *,
    workbook,
    sheet_name: str,
    header_row: int,
    columns,
    logical_values: dict,
    include_data: bool = True,
) -> None:
    sheet = workbook.create_sheet(
        sheet_name
    )

    physical_columns = tuple(
        (
            logical_name,
            physical_name,
        )
        for logical_name, physical_name
        in columns.items()
        if physical_name is not None
    )

    for column_number, (_, physical_name) in enumerate(
        physical_columns,
        start=1,
    ):
        sheet.cell(
            row=header_row,
            column=column_number,
            value=physical_name,
        )

    if not include_data:
        return

    for column_number, (logical_name, _) in enumerate(
        physical_columns,
        start=1,
    ):
        sheet.cell(
            row=header_row + 1,
            column=column_number,
            value=logical_values.get(
                logical_name
            ),
        )


def _workbook_bytes(
    schema: WeeklyScheduleWorkbookSchema | None = None,
) -> bytes:
    schema = (
        schema
        or WeeklyScheduleWorkbookSchema()
    )

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _write_table(
        workbook=workbook,
        sheet_name=schema.academic_weeks_sheet,
        header_row=schema.academic_weeks_header_row,
        columns=schema.academic_week_columns,
        logical_values={
            "academic_year": "2026-2027",
            "week_number": 5,
            "start_date": "2026-09-28",
            "end_date": "2026-10-04",
        },
    )

    _write_table(
        workbook=workbook,
        sheet_name=schema.timetable_sheet,
        header_row=schema.timetable_header_row,
        columns=schema.timetable_columns,
        logical_values={
            "teacher_id": "GV001",
            "class_id": "6A1",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "weekday": 2,
            "timetable_period": 1,
            "effective_from": "2026-09-01",
            "effective_to": "2027-05-31",
        },
    )

    _write_table(
        workbook=workbook,
        sheet_name=schema.curriculum_sheet,
        header_row=schema.curriculum_header_row,
        columns=schema.curriculum_columns,
        logical_values={
            "class_id": "6A1",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "period_number": 9,
            "lesson_id": "LESSON-009",
            "lesson_title": "Bai hoc thu 9",
            "period_in_lesson": 1,
            "total_lesson_periods": 1,
            "teaching_equipment": None,
        },
    )

    _write_table(
        workbook=workbook,
        sheet_name=schema.executions_sheet,
        header_row=schema.executions_header_row,
        columns=schema.execution_columns,
        logical_values={},
        include_data=False,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    return buffer.getvalue()


def _local_selection() -> OperationalInputSelection:
    return OperationalInputSelection(
        reference=OperationalInputReference(
            location=(
                OperationalInputLocation.LOCAL_UPLOAD
            ),
        ),
        source=None,
    )


def _library_selection() -> OperationalInputSelection:
    source = OperationalDataSource(
        source_id="PPCT-2026",
        data_type=OperationalDataType.PPCT,
        origin=OperationalDataOrigin.FILE_IMPORTED,
        owner_id="GV001",
        academic_year="2026-2027",
        status=OperationalDataStatus.ACTIVE,
    )

    return OperationalInputSelection(
        reference=OperationalInputReference(
            location=(
                OperationalInputLocation.SYSTEM_LIBRARY
            ),
            source_id=source.source_id,
            source_academic_year=(
                source.academic_year
            ),
        ),
        source=source,
    )


def expect_error(
    error_type,
    action,
) -> bool:
    try:
        action()
    except error_type:
        return True
    except Exception:
        return False

    return False


def run_contract() -> bool:
    print("=" * 72)
    print(
        "MVP-OPS-002E - "
        "LOCAL WEEKLY SCHEDULE WORKBOOK INTAKE TEST"
    )
    print("=" * 72)

    tests = []

    selection = _local_selection()

    adapter = (
        WeeklyScheduleWorkbookIntakeAdapter()
    )

    content = _workbook_bytes()

    result = adapter.load(
        selection=selection,
        workbook_bytes=content,
    )

    tests.append((
        "OWI1 Local workbook accepted",
        isinstance(
            result,
            WeeklyScheduleWorkbookIntakeResult,
        ),
    ))

    tests.append((
        "OWI2 Input selection preserved",
        result.selection is selection,
    ))

    tests.append((
        "OWI3 Canonical source data produced",
        isinstance(
            result.source_data,
            WeeklyScheduleSourceData,
        ),
    ))

    tests.append((
        "OWI4 Academic week parsed",
        (
            len(
                result.source_data.academic_weeks
            )
            == 1
            and
            result.source_data.academic_weeks[
                0
            ].week_number
            == 5
        ),
    ))

    tests.append((
        "OWI5 Timetable parsed",
        (
            len(
                result.source_data.timetable_slots
            )
            == 1
            and
            result.source_data.timetable_slots[
                0
            ].teacher_id
            == "GV001"
        ),
    ))

    tests.append((
        "OWI6 Curriculum period parsed",
        (
            len(
                result.source_data.curriculum_periods
            )
            == 1
            and
            result.source_data.curriculum_periods[
                0
            ].period_number
            == 9
        ),
    ))

    tests.append((
        "OWI7 Empty execution history accepted",
        (
            result.source_data.execution_records
            == ()
        ),
    ))

    week = result.source_data.week(
        5,
        "2026-2027",
    )

    tests.append((
        "OWI8 Parsed source supports week lookup",
        week.start_date.isoformat()
        == "2026-09-28",
    ))

    tests.append((
        "OWI9 Empty workbook blocked",
        expect_error(
            ValueError,
            lambda: adapter.load(
                selection=selection,
                workbook_bytes=b"",
            ),
        ),
    ))

    tests.append((
        "OWI10 Wrong workbook type blocked",
        expect_error(
            TypeError,
            lambda: adapter.load(
                selection=selection,
                workbook_bytes="not-bytes",
            ),
        ),
    ))

    tests.append((
        "OWI11 Invalid Excel bytes blocked",
        expect_error(
            WeeklyScheduleWorkbookError,
            lambda: adapter.load(
                selection=selection,
                workbook_bytes=b"not-an-xlsx",
            ),
        ),
    ))

    tests.append((
        "OWI12 SYSTEM_LIBRARY cannot enter local upload adapter",
        expect_error(
            ValueError,
            lambda: adapter.load(
                selection=_library_selection(),
                workbook_bytes=content,
            ),
        ),
    ))

    tests.append((
        "OWI13 Wrong selection type blocked",
        expect_error(
            TypeError,
            lambda: adapter.load(
                selection="LOCAL_UPLOAD",
                workbook_bytes=content,
            ),
        ),
    ))

    tests.append((
        "OWI14 Wrong schema type blocked",
        expect_error(
            TypeError,
            lambda: (
                WeeklyScheduleWorkbookIntakeAdapter(
                    schema="bad-schema"
                )
            ),
        ),
    ))

    custom_schema = (
        WeeklyScheduleWorkbookSchema()
    )

    custom_result = (
        WeeklyScheduleWorkbookIntakeAdapter(
            custom_schema
        ).load(
            selection=selection,
            workbook_bytes=_workbook_bytes(
                custom_schema
            ),
        )
    )

    tests.append((
        "OWI15 Explicit schema supported",
        (
            custom_result
            .source_data
            .curriculum_periods[0]
            .lesson_id
            == "LESSON-009"
        ),
    ))

    adapter_source = inspect.getsource(
        WeeklyScheduleWorkbookIntakeAdapter
    )

    tests.append((
        "OWI16 Existing Excel adapter reused",
        "WeeklyScheduleExcelAdapter"
        in adapter_source,
    ))

    tests.append((
        "OWI17 Workbook parser not duplicated",
        "load_workbook"
        not in adapter_source,
    ))

    tests.append((
        "OWI18 Intake owns no persistence dependency",
        not any(
            token
            in adapter_source.lower()
            for token in (
                "supabase",
                "sqlite3",
                "googleapiclient",
                "streamlit",
            )
        ),
    ))

    tests.append((
        "OWI19 Intake does not create educational rules",
        not any(
            token
            in adapter_source
            for token in (
                "140",
                "105",
                "70",
                "35",
                "KNTT",
                "Toán 6",
            )
        ),
    ))

    tests.append((
        "OWI20 Workbook remains transport container",
        not any(
            token
            in adapter_source
            for token in (
                "OperationalDataType.PPCT",
                "OperationalDataType.TIMETABLE",
                "OperationalDataType.ACADEMIC_WEEK",
            )
        ),
    ))

    results = []

    for label, passed in tests:
        results.append(passed)

        print(
            f"{label}: "
            f"{'PASS' if passed else 'FAIL'}"
        )

    print()

    if all(results):
        print(
            "RESULT: PASS - LOCAL WEEKLY SCHEDULE "
            "WORKBOOK INTAKE VERIFIED"
        )
        return True

    print(
        "RESULT: FAIL - LOCAL WEEKLY SCHEDULE "
        "WORKBOOK INTAKE VIOLATED"
    )

    return False


def test_operational_weekly_schedule_workbook_intake():
    assert run_contract()


def main():
    raise SystemExit(
        0 if run_contract() else 1
    )


if __name__ == "__main__":
    main()
