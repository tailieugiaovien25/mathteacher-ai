from dataclasses import replace
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from educational_planning_v2.adapters import WeeklyScheduleExcelAdapter
from educational_planning_v2.exporters import WeeklyScheduleExcelExporter
from educational_planning_v2.services import WeeklyTeachingScheduleService


PROJECT_ROOT = Path(__file__).resolve().parents[3]
SOURCE_TEMPLATE = (
    PROJECT_ROOT
    / "templates"
    / "weekly_schedule"
    / "mau_du_lieu_lich_bao_giang_v2.xlsx"
)


def _schedule():
    data = WeeklyScheduleExcelAdapter().load(SOURCE_TEMPLATE)
    return WeeklyTeachingScheduleService().build(
        schedule_id="GV001-2026-2027-W05",
        teacher_id="GV001",
        academic_week=data.week(5, "2026-2027"),
        timetable_slots=data.timetable_slots,
        curriculum_periods=data.curriculum_periods,
        execution_records=data.execution_records,
    )


def test_export_contains_typed_schedule_data_and_safe_filename():
    result = WeeklyScheduleExcelExporter().export(_schedule())

    assert result.file_name == "lich-bao-giang-GV001-2026-2027-tuan-05.xlsx"
    assert result.content.startswith(b"PK")
    workbook = load_workbook(BytesIO(result.content), data_only=False)
    sheet = workbook["Lich_bao_giang"]
    assert sheet["A1"].value == "LỊCH BÁO GIẢNG TUẦN 5"
    assert sheet["A6"].value.date().isoformat() == "2026-09-28"
    assert sheet["G6"].value == 3
    assert sheet["I6"].value == "Phần tử của tập hợp"
    assert sheet["K8"].value == "Máy chiếu"
    assert sheet["A6"].number_format == "dd/mm/yyyy"
    workbook.close()


def test_export_is_formatted_for_review_and_printing():
    result = WeeklyScheduleExcelExporter().export(_schedule())
    workbook = load_workbook(BytesIO(result.content), data_only=False)
    sheet = workbook["Lich_bao_giang"]

    assert "A1:K1" in {str(item) for item in sheet.merged_cells.ranges}
    assert sheet.freeze_panes == "A6"
    assert sheet.sheet_view.showGridLines is False
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.tables["WeeklyScheduleTable"].ref == "A5:K8"
    assert sheet["A10"].value == "Tổng số tiết: 3  |  Số lớp: 2  |  Số môn/phân môn: 2"
    workbook.close()


def test_export_rejects_empty_schedule():
    empty = replace(_schedule(), entries=())

    with pytest.raises(ValueError, match="không có tiết dạy"):
        WeeklyScheduleExcelExporter().export(empty)


def test_export_rejects_non_schedule_value():
    with pytest.raises(TypeError, match="WeeklyTeachingSchedule"):
        WeeklyScheduleExcelExporter().export(object())
