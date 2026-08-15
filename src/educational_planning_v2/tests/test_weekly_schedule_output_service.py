from __future__ import annotations

from io import BytesIO
import inspect

from openpyxl import Workbook, load_workbook

from educational_planning_v2.adapters.operational_weekly_schedule_workbook_intake import (
    WeeklyScheduleWorkbookIntakeAdapter,
)
from educational_planning_v2.adapters.weekly_schedule_excel_adapter import (
    WeeklyScheduleWorkbookSchema,
)
from educational_planning_v2.exporters.weekly_schedule_excel_exporter import (
    WeeklyScheduleExcelExport,
    WeeklyScheduleExcelExporter,
)
from educational_planning_v2.models.operational_data_io import (
    OperationalInputLocation,
    OperationalInputReference,
)
from educational_planning_v2.services.local_weekly_schedule_generation_service import (
    LocalWeeklyScheduleGenerationService,
    WeeklyScheduleGenerationRequest,
    WeeklyScheduleGenerationResult,
)
from educational_planning_v2.services.operational_input_selection_service import (
    OperationalInputSelection,
)
from educational_planning_v2.services.weekly_schedule_output_service import (
    WeeklyScheduleOutputResult,
    WeeklyScheduleOutputService,
)


def _write_headers(
    sheet,
    *,
    row_number: int,
    columns,
):
    physical_columns = tuple(
        (
            logical_name,
            physical_name,
        )
        for logical_name, physical_name
        in columns.items()
        if physical_name is not None
    )

    for column_number, (_, physical_name) in enumerate(
        physical_columns,
        start=1,
    ):
        sheet.cell(
            row=row_number,
            column=column_number,
            value=physical_name,
        )

    return physical_columns


def _append_row(
    sheet,
    *,
    row_number: int,
    physical_columns,
    values: dict,
):
    for column_number, (logical_name, _) in enumerate(
        physical_columns,
        start=1,
    ):
        sheet.cell(
            row=row_number,
            column=column_number,
            value=values.get(logical_name),
        )


def _workbook_bytes() -> bytes:
    schema = WeeklyScheduleWorkbookSchema()

    workbook = Workbook()
    workbook.remove(workbook.active)

    # WEEK
    sheet = workbook.create_sheet(
        schema.academic_weeks_sheet
    )

    columns = _write_headers(
        sheet,
        row_number=schema.academic_weeks_header_row,
        columns=schema.academic_week_columns,
    )

    _append_row(
        sheet,
        row_number=schema.academic_weeks_header_row + 1,
        physical_columns=columns,
        values={
            "academic_year": "2026-2027",
            "week_number": 5,
            "start_date": "2026-09-28",
            "end_date": "2026-10-04",
        },
    )

    # TIMETABLE
    sheet = workbook.create_sheet(
        schema.timetable_sheet
    )

    columns = _write_headers(
        sheet,
        row_number=schema.timetable_header_row,
        columns=schema.timetable_columns,
    )

    for offset, values in enumerate(
        (
            {
                "teacher_id": "GV001",
                "class_id": "6A1",
                "subject_ref": "MATHEMATICS",
                "component_ref": None,
                "weekday": 2,
                "timetable_period": 1,
                "effective_from": "2026-09-01",
                "effective_to": "2027-05-31",
            },
            {
                "teacher_id": "GV001",
                "class_id": "6A1",
                "subject_ref": "MATHEMATICS",
                "component_ref": None,
                "weekday": 4,
                "timetable_period": 2,
                "effective_from": "2026-09-01",
                "effective_to": "2027-05-31",
            },
        ),
        start=1,
    ):
        _append_row(
            sheet,
            row_number=schema.timetable_header_row + offset,
            physical_columns=columns,
            values=values,
        )

    # PPCT
    sheet = workbook.create_sheet(
        schema.curriculum_sheet
    )

    columns = _write_headers(
        sheet,
        row_number=schema.curriculum_header_row,
        columns=schema.curriculum_columns,
    )

    for offset, period in enumerate(
        (9, 10),
        start=1,
    ):
        _append_row(
            sheet,
            row_number=schema.curriculum_header_row + offset,
            physical_columns=columns,
            values={
                "class_id": "6A1",
                "subject_ref": "MATHEMATICS",
                "component_ref": None,
                "period_number": period,
                "lesson_id": f"LESSON-{period:03d}",
                "lesson_title": f"Bai hoc thu {period}",
                "period_in_lesson": 1,
                "total_lesson_periods": 1,
                "teaching_equipment": None,
            },
        )

    # EXECUTION HISTORY
    sheet = workbook.create_sheet(
        schema.executions_sheet
    )

    columns = _write_headers(
        sheet,
        row_number=schema.executions_header_row,
        columns=schema.execution_columns,
    )

    completed_dates = (
        "2026-09-01",
        "2026-09-03",
        "2026-09-08",
        "2026-09-10",
        "2026-09-15",
        "2026-09-17",
        "2026-09-22",
        "2026-09-24",
    )

    for offset, teaching_date in enumerate(
        completed_dates,
        start=1,
    ):
        _append_row(
            sheet,
            row_number=schema.executions_header_row + offset,
            physical_columns=columns,
            values={
                "teacher_id": "GV001",
                "class_id": "6A1",
                "subject_ref": "MATHEMATICS",
                "component_ref": None,
                "teaching_date": teaching_date,
                "curriculum_period": offset,
                "status": "COMPLETED",
            },
        )

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    return buffer.getvalue()


def _generation() -> WeeklyScheduleGenerationResult:
    selection = OperationalInputSelection(
        reference=OperationalInputReference(
            location=OperationalInputLocation.LOCAL_UPLOAD,
        ),
        source=None,
    )

    intake = WeeklyScheduleWorkbookIntakeAdapter().load(
        selection=selection,
        workbook_bytes=_workbook_bytes(),
    )

    return LocalWeeklyScheduleGenerationService().generate(
        intake=intake,
        request=WeeklyScheduleGenerationRequest(
            schedule_id="GV001-2026-2027-W05",
            teacher_id="GV001",
            academic_year="2026-2027",
            week_number=5,
        ),
    )


def expect_error(
    error_type,
    action,
) -> bool:
    try:
        action()
    except error_type:
        return True
    except Exception:
        return False

    return False


def run_contract() -> bool:
    print("=" * 72)
    print(
        "MVP-OPS-002G - "
        "WEEKLY SCHEDULE OUTPUT INTEGRATION TEST"
    )
    print("=" * 72)

    tests = []

    generation = _generation()

    service = WeeklyScheduleOutputService()

    output = service.export_excel(
        generation=generation,
    )

    tests.append((
        "WOI1 Valid generation result accepted",
        isinstance(
            output,
            WeeklyScheduleOutputResult,
        ),
    ))

    tests.append((
        "WOI2 WeeklyTeachingSchedule preserved",
        output.generation.schedule
        is generation.schedule,
    ))

    tests.append((
        "WOI3 Existing exporter reused",
        isinstance(
            service._exporter,
            WeeklyScheduleExcelExporter,
        ),
    ))

    tests.append((
        "WOI4 Excel artifact produced",
        isinstance(
            output.artifact,
            WeeklyScheduleExcelExport,
        ),
    ))

    tests.append((
        "WOI5 File name preserved",
        (
            output.artifact.file_name
            == "lich-bao-giang-GV001-2026-2027-tuan-05.xlsx"
        ),
    ))

    expected_mime = (
        "application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet"
    )

    tests.append((
        "WOI6 MIME type preserved",
        output.artifact.mime_type
        == expected_mime,
    ))

    tests.append((
        "WOI7 Workbook bytes produced",
        (
            isinstance(
                output.artifact.content,
                bytes,
            )
            and
            output.artifact.content.startswith(
                b"PK"
            )
        ),
    ))

    workbook = load_workbook(
        BytesIO(
            output.artifact.content
        ),
        data_only=False,
    )

    tests.append((
        "WOI8 Workbook contains schedule sheet",
        "Lich_bao_giang"
        in workbook.sheetnames,
    ))

    sheet = workbook[
        "Lich_bao_giang"
    ]

    tests.append((
        "WOI9 Workbook contains generated entries",
        (
            sheet.max_row >= 7
        ),
    ))

    workbook.close()

    tests.append((
        "WOI10 Wrong generation result blocked",
        expect_error(
            TypeError,
            lambda: service.export_excel(
                generation="bad-generation",
            ),
        ),
    ))

    tests.append((
        "WOI11 Wrong exporter blocked",
        expect_error(
            TypeError,
            lambda: WeeklyScheduleOutputService(
                exporter=object()
            ),
        ),
    ))

    tests.append((
        "WOI12 Output result requires valid generation",
        expect_error(
            TypeError,
            lambda: WeeklyScheduleOutputResult(
                generation="bad",
                artifact=output.artifact,
            ),
        ),
    ))

    tests.append((
        "WOI13 Output result requires valid artifact",
        expect_error(
            TypeError,
            lambda: WeeklyScheduleOutputResult(
                generation=generation,
                artifact="bad",
            ),
        ),
    ))

    service_source = inspect.getsource(
        WeeklyScheduleOutputService
    )

    tests.append((
        "WOI14 Output service owns no workbook rendering logic",
        not any(
            token
            in service_source
            for token in (
                "Workbook(",
                "load_workbook",
                "merge_cells",
                "PatternFill",
                "Font(",
                "Alignment(",
            )
        ),
    ))

    tests.append((
        "WOI15 Output service owns no physical file writing",
        not any(
            token
            in service_source
            for token in (
                "open(",
                "Path(",
                "write_bytes",
                "write_text",
            )
        ),
    ))

    lower_source = service_source.lower()

    tests.append((
        "WOI16 Output service owns no delivery dependency",
        not any(
            token
            in lower_source
            for token in (
                "streamlit",
                "supabase",
                "googleapiclient",
                "google drive",
                "vtsmas",
            )
        ),
    ))

    tests.append((
        "WOI17 Generation and output responsibilities remain separate",
        (
            "WeeklyTeachingScheduleService"
            not in service_source
            and
            "WeeklyScheduleWorkbookIntakeAdapter"
            not in service_source
        ),
    ))

    tests.append((
        "WOI18 Existing WeeklyScheduleExcelExporter is reused",
        (
            "_exporter.export"
            in service_source
        ),
    ))

    results = []

    for label, passed in tests:
        results.append(passed)

        print(
            f"{label}: "
            f"{'PASS' if passed else 'FAIL'}"
        )

    print()

    print("OUTPUT SUMMARY")

    print(
        "FILE NAME :",
        output.artifact.file_name,
    )

    print(
        "MIME TYPE :",
        output.artifact.mime_type,
    )

    print(
        "BYTES     :",
        len(output.artifact.content),
    )

    print(
        "SCHEDULE  :",
        output.generation.schedule.schedule_id,
    )

    print()

    if all(results):
        print(
            "RESULT: PASS - WEEKLY SCHEDULE "
            "OUTPUT INTEGRATION VERIFIED"
        )

        return True

    print(
        "RESULT: FAIL - WEEKLY SCHEDULE "
        "OUTPUT INTEGRATION VIOLATED"
    )

    return False


def test_weekly_schedule_output_service():
    assert run_contract()


def main():
    raise SystemExit(
        0 if run_contract() else 1
    )


if __name__ == "__main__":
    main()
