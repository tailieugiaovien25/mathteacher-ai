from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

from educational_planning_v2.adapters import (
    WeeklyScheduleExcelAdapter,
    WeeklyScheduleWorkbookError,
    WeeklyScheduleWorkbookSchema,
)
from educational_planning_v2.services import WeeklyTeachingScheduleService


def _workbook_bytes(*, missing_week_column: bool = False) -> BytesIO:
    workbook = Workbook()
    week = workbook.active
    week.title = "Tuan_hoc"
    week.append(
        ["nam_hoc", "tuan", "tu_ngay"]
        if missing_week_column
        else ["nam_hoc", "tuan", "tu_ngay", "den_ngay"]
    )
    week.append(
        ["2026-2027", 5, date(2026, 9, 28)]
        if missing_week_column
        else ["2026-2027", 5, date(2026, 9, 28), date(2026, 10, 4)]
    )

    timetable = workbook.create_sheet("Thoi_khoa_bieu")
    timetable.append(
        [
            "ma_giao_vien", "lop", "mon_hoc", "phan_mon", "thu",
            "tiet_hoc", "hieu_luc_tu", "hieu_luc_den",
        ]
    )
    timetable.append(
        ["GV001", "6A1", "Toan", None, "Thứ 2", 1,
         date(2026, 9, 1), date(2027, 1, 15)]
    )
    timetable.append(
        ["GV001", "6A1", "Toan", None, "Thứ 4", 2,
         date(2026, 9, 1), date(2027, 1, 15)]
    )

    curriculum = workbook.create_sheet("PPCT")
    curriculum.append(
        [
            "lop", "mon_hoc", "phan_mon", "tiet_ppct", "ma_bai_hoc",
            "ten_bai_hoc", "tiet_trong_bai", "tong_tiet_cua_bai",
            "thiet_bi_day_hoc",
        ]
    )
    for period in range(1, 11):
        curriculum.append(
            ["6A1", "Toan", None, period, f"TOAN6-{period:03}",
             f"Bai {period}", 1, 1, "May chieu;Phieu hoc tap"]
        )

    executions = workbook.create_sheet("Tiet_da_day")
    executions.append(
        [
            "ma_giao_vien", "lop", "mon_hoc", "phan_mon", "ngay_day",
            "tiet_ppct", "trang_thai",
        ]
    )
    for period in range(1, 9):
        executions.append(
            ["GV001", "6A1", "Toan", None, date(2026, 9, period + 1),
             period, "Đã dạy"]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_loads_all_four_source_tables():
    data = WeeklyScheduleExcelAdapter().load(_workbook_bytes())

    assert data.week(5, "2026-2027").start_date == date(2026, 9, 28)
    assert len(data.timetable_slots) == 2
    assert len(data.curriculum_periods) == 10
    assert len(data.execution_records) == 8
    assert data.execution_records[0].is_completed
    assert data.curriculum_periods[0].teaching_equipment == (
        "May chieu",
        "Phieu hoc tap",
    )


def test_loaded_data_builds_week_five_schedule():
    data = WeeklyScheduleExcelAdapter().load(_workbook_bytes())

    schedule = WeeklyTeachingScheduleService().build(
        schedule_id="GV001-2026-2027-W05",
        teacher_id="GV001",
        academic_week=data.week(5, "2026-2027"),
        timetable_slots=data.timetable_slots,
        curriculum_periods=data.curriculum_periods,
        execution_records=data.execution_records,
    )

    assert [entry.curriculum_period for entry in schedule.entries] == [9, 10]
    assert [entry.weekday for entry in schedule.entries] == [1, 3]


def test_missing_column_reports_sheet_and_column():
    with pytest.raises(WeeklyScheduleWorkbookError) as error:
        WeeklyScheduleExcelAdapter().load(
            _workbook_bytes(missing_week_column=True)
        )

    assert "Tuan_hoc" in str(error.value)
    assert "den_ngay" in str(error.value)


def test_sheet_names_can_change_without_changing_the_adapter():
    source = _workbook_bytes()
    from openpyxl import load_workbook

    workbook = load_workbook(source)
    workbook["Tuan_hoc"].title = "DM_Tuan"
    renamed = BytesIO()
    workbook.save(renamed)
    renamed.seek(0)

    schema = WeeklyScheduleWorkbookSchema(academic_weeks_sheet="DM_Tuan")
    data = WeeklyScheduleExcelAdapter(schema).load(renamed)

    assert data.week(5).end_date == date(2026, 10, 4)
