from __future__ import annotations

from io import BytesIO
import inspect

from openpyxl import Workbook

from educational_planning_v2.adapters.operational_weekly_schedule_workbook_intake import (
    WeeklyScheduleWorkbookIntakeAdapter,
)
from educational_planning_v2.adapters.weekly_schedule_excel_adapter import (
    WeeklyScheduleWorkbookSchema,
)
from educational_planning_v2.models.operational_data_io import (
    OperationalInputLocation,
    OperationalInputReference,
)
from educational_planning_v2.models.weekly_teaching_schedule import (
    WeeklyTeachingSchedule,
)
from educational_planning_v2.services.local_weekly_schedule_generation_service import (
    LocalWeeklyScheduleGenerationService,
    WeeklyScheduleGenerationRequest,
    WeeklyScheduleGenerationResult,
)
from educational_planning_v2.services.operational_input_selection_service import (
    OperationalInputSelection,
)


def _write_headers(
    sheet,
    *,
    row_number: int,
    columns,
):
    physical_columns = tuple(
        (
            logical_name,
            physical_name,
        )
        for logical_name, physical_name
        in columns.items()
        if physical_name is not None
    )

    for column_number, (_, physical_name) in enumerate(
        physical_columns,
        start=1,
    ):
        sheet.cell(
            row=row_number,
            column=column_number,
            value=physical_name,
        )

    return physical_columns


def _append_logical_row(
    sheet,
    *,
    row_number: int,
    physical_columns,
    values: dict,
):
    for column_number, (logical_name, _) in enumerate(
        physical_columns,
        start=1,
    ):
        sheet.cell(
            row=row_number,
            column=column_number,
            value=values.get(
                logical_name
            ),
        )


def _workbook_bytes() -> bytes:
    schema = WeeklyScheduleWorkbookSchema()

    workbook = Workbook()
    workbook.remove(
        workbook.active
    )

    # --------------------------------------------------------
    # ACADEMIC WEEK
    # --------------------------------------------------------

    week_sheet = workbook.create_sheet(
        schema.academic_weeks_sheet
    )

    week_columns = _write_headers(
        week_sheet,
        row_number=schema.academic_weeks_header_row,
        columns=schema.academic_week_columns,
    )

    _append_logical_row(
        week_sheet,
        row_number=(
            schema.academic_weeks_header_row
            + 1
        ),
        physical_columns=week_columns,
        values={
            "academic_year": "2026-2027",
            "week_number": 5,
            "start_date": "2026-09-28",
            "end_date": "2026-10-04",
        },
    )

    # --------------------------------------------------------
    # TIMETABLE
    # --------------------------------------------------------

    timetable_sheet = workbook.create_sheet(
        schema.timetable_sheet
    )

    timetable_columns = _write_headers(
        timetable_sheet,
        row_number=schema.timetable_header_row,
        columns=schema.timetable_columns,
    )

    _append_logical_row(
        timetable_sheet,
        row_number=(
            schema.timetable_header_row
            + 1
        ),
        physical_columns=timetable_columns,
        values={
            "teacher_id": "GV001",
            "class_id": "6A1",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "weekday": 2,
            "timetable_period": 1,
            "effective_from": "2026-09-01",
            "effective_to": "2027-05-31",
        },
    )

    _append_logical_row(
        timetable_sheet,
        row_number=(
            schema.timetable_header_row
            + 2
        ),
        physical_columns=timetable_columns,
        values={
            "teacher_id": "GV001",
            "class_id": "6A1",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "weekday": 4,
            "timetable_period": 2,
            "effective_from": "2026-09-01",
            "effective_to": "2027-05-31",
        },
    )

    # Another teacher must not enter GV001 schedule.
    _append_logical_row(
        timetable_sheet,
        row_number=(
            schema.timetable_header_row
            + 3
        ),
        physical_columns=timetable_columns,
        values={
            "teacher_id": "GV002",
            "class_id": "6A2",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "weekday": 3,
            "timetable_period": 1,
            "effective_from": "2026-09-01",
            "effective_to": "2027-05-31",
        },
    )

    # --------------------------------------------------------
    # PPCT / CURRICULUM
    # --------------------------------------------------------

    curriculum_sheet = workbook.create_sheet(
        schema.curriculum_sheet
    )

    curriculum_columns = _write_headers(
        curriculum_sheet,
        row_number=schema.curriculum_header_row,
        columns=schema.curriculum_columns,
    )

    _append_logical_row(
        curriculum_sheet,
        row_number=(
            schema.curriculum_header_row
            + 1
        ),
        physical_columns=curriculum_columns,
        values={
            "class_id": "6A1",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "period_number": 9,
            "lesson_id": "LESSON-009",
            "lesson_title": "Bai hoc thu 9",
            "period_in_lesson": 1,
            "total_lesson_periods": 1,
            "teaching_equipment": None,
        },
    )

    _append_logical_row(
        curriculum_sheet,
        row_number=(
            schema.curriculum_header_row
            + 2
        ),
        physical_columns=curriculum_columns,
        values={
            "class_id": "6A1",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "period_number": 10,
            "lesson_id": "LESSON-010",
            "lesson_title": "Bai hoc thu 10",
            "period_in_lesson": 1,
            "total_lesson_periods": 1,
            "teaching_equipment": None,
        },
    )

    _append_logical_row(
        curriculum_sheet,
        row_number=(
            schema.curriculum_header_row
            + 3
        ),
        physical_columns=curriculum_columns,
        values={
            "class_id": "6A2",
            "subject_ref": "MATHEMATICS",
            "component_ref": None,
            "period_number": 9,
            "lesson_id": "LESSON-6A2-009",
            "lesson_title": "Other teacher lesson",
            "period_in_lesson": 1,
            "total_lesson_periods": 1,
            "teaching_equipment": None,
        },
    )

    # --------------------------------------------------------
    # EXECUTION HISTORY
    # --------------------------------------------------------

    execution_sheet = workbook.create_sheet(
        schema.executions_sheet
    )

    execution_columns = _write_headers(
        execution_sheet,
        row_number=schema.executions_header_row,
        columns=schema.execution_columns,
    )

    # Eight lessons were completed before week 5.
    # Therefore the next two curriculum periods are 9 and 10.
    completed_dates = (
        "2026-09-01",
        "2026-09-03",
        "2026-09-08",
        "2026-09-10",
        "2026-09-15",
        "2026-09-17",
        "2026-09-22",
        "2026-09-24",
    )

    for offset, teaching_date in enumerate(
        completed_dates,
        start=1,
    ):
        _append_logical_row(
            execution_sheet,
            row_number=(
                schema.executions_header_row
                + offset
            ),
            physical_columns=execution_columns,
            values={
                "teacher_id": "GV001",
                "class_id": "6A1",
                "subject_ref": "MATHEMATICS",
                "component_ref": None,
                "teaching_date": teaching_date,
                "curriculum_period": offset,
                "status": "COMPLETED",
            },
        )

    buffer = BytesIO()

    workbook.save(
        buffer
    )

    workbook.close()

    return buffer.getvalue()


def _selection() -> OperationalInputSelection:
    return OperationalInputSelection(
        reference=OperationalInputReference(
            location=OperationalInputLocation.LOCAL_UPLOAD,
        ),
        source=None,
    )


def expect_error(
    error_type,
    action,
) -> bool:
    try:
        action()
    except error_type:
        return True
    except Exception:
        return False

    return False


def run_contract() -> bool:
    print("=" * 72)
    print(
        "MVP-OPS-002F - "
        "LOCAL UPLOAD TO WEEKLY SCHEDULE E2E TEST"
    )
    print("=" * 72)

    tests = []

    # ========================================================
    # STEP 1 - LOCAL UPLOAD INTAKE
    # ========================================================

    intake = (
        WeeklyScheduleWorkbookIntakeAdapter()
        .load(
            selection=_selection(),
            workbook_bytes=_workbook_bytes(),
        )
    )

    tests.append((
        "OWG1 Local workbook entered operational pipeline",
        (
            len(
                intake.source_data.academic_weeks
            )
            == 1
        ),
    ))

    tests.append((
        "OWG2 Timetable became canonical source data",
        (
            len(
                intake.source_data.timetable_slots
            )
            == 3
        ),
    ))

    tests.append((
        "OWG3 PPCT became canonical curriculum data",
        (
            len(
                intake.source_data.curriculum_periods
            )
            == 3
        ),
    ))

    # ========================================================
    # STEP 2 - GENERATE WEEKLY SCHEDULE
    # ========================================================

    request = WeeklyScheduleGenerationRequest(
        schedule_id=" GV001-2026-2027-W05 ",
        teacher_id=" GV001 ",
        academic_year=" 2026-2027 ",
        week_number=5,
    )

    tests.append((
        "OWG4 Generation request normalized",
        (
            request.schedule_id
            == "GV001-2026-2027-W05"
            and
            request.teacher_id
            == "GV001"
            and
            request.academic_year
            == "2026-2027"
        ),
    ))

    result = (
        LocalWeeklyScheduleGenerationService()
        .generate(
            intake=intake,
            request=request,
        )
    )

    tests.append((
        "OWG5 Weekly schedule generated",
        isinstance(
            result,
            WeeklyScheduleGenerationResult,
        ),
    ))

    tests.append((
        "OWG6 Generated product is WeeklyTeachingSchedule",
        isinstance(
            result.schedule,
            WeeklyTeachingSchedule,
        ),
    ))

    tests.append((
        "OWG7 Schedule identity preserved",
        result.schedule.schedule_id
        == "GV001-2026-2027-W05",
    ))

    tests.append((
        "OWG8 Teacher identity preserved",
        result.schedule.teacher_id
        == "GV001",
    ))

    tests.append((
        "OWG9 Academic week preserved",
        (
            result.schedule
            .academic_week
            .week_number
            == 5
            and
            result.schedule
            .academic_week
            .academic_year
            == "2026-2027"
        ),
    ))

    tests.append((
        "OWG10 Teacher timetable produced two entries",
        len(
            result.schedule.entries
        )
        == 2,
    ))

    tests.append((
        "OWG11 PPCT progression preserved",
        tuple(
            entry.curriculum_period
            for entry
            in result.schedule.entries
        )
        == (
            9,
            10,
        ),
    ))

    tests.append((
        "OWG12 Timetable weekdays preserved",
        tuple(
            entry.weekday
            for entry
            in result.schedule.entries
        )
        == (
            1,
            3,
        ),
    ))

    tests.append((
        "OWG13 Other teacher excluded",
        all(
            entry.teacher_id
            == "GV001"
            for entry
            in result.schedule.entries
        ),
    ))

    tests.append((
        "OWG14 Class identity preserved",
        all(
            entry.class_id
            == "6A1"
            for entry
            in result.schedule.entries
        ),
    ))

    tests.append((
        "OWG15 Subject identity preserved",
        all(
            entry.subject_ref
            == "MATHEMATICS"
            for entry
            in result.schedule.entries
        ),
    ))

    # ========================================================
    # VALIDATION
    # ========================================================

    tests.append((
        "OWG16 Empty teacher blocked",
        expect_error(
            ValueError,
            lambda: WeeklyScheduleGenerationRequest(
                schedule_id="S",
                teacher_id=" ",
                academic_year="2026-2027",
                week_number=5,
            ),
        ),
    ))

    tests.append((
        "OWG17 Empty academic year blocked",
        expect_error(
            ValueError,
            lambda: WeeklyScheduleGenerationRequest(
                schedule_id="S",
                teacher_id="GV001",
                academic_year=" ",
                week_number=5,
            ),
        ),
    ))

    tests.append((
        "OWG18 Invalid week number blocked",
        expect_error(
            ValueError,
            lambda: WeeklyScheduleGenerationRequest(
                schedule_id="S",
                teacher_id="GV001",
                academic_year="2026-2027",
                week_number=0,
            ),
        ),
    ))

    tests.append((
        "OWG19 Missing week blocked",
        expect_error(
            ValueError,
            lambda: (
                LocalWeeklyScheduleGenerationService()
                .generate(
                    intake=intake,
                    request=WeeklyScheduleGenerationRequest(
                        schedule_id="GV001-W99",
                        teacher_id="GV001",
                        academic_year="2026-2027",
                        week_number=99,
                    ),
                )
            ),
        ),
    ))

    tests.append((
        "OWG20 Wrong intake type blocked",
        expect_error(
            TypeError,
            lambda: (
                LocalWeeklyScheduleGenerationService()
                .generate(
                    intake="bad-intake",
                    request=request,
                )
            ),
        ),
    ))

    tests.append((
        "OWG21 Wrong request type blocked",
        expect_error(
            TypeError,
            lambda: (
                LocalWeeklyScheduleGenerationService()
                .generate(
                    intake=intake,
                    request="bad-request",
                )
            ),
        ),
    ))

    # ========================================================
    # ARCHITECTURE
    # ========================================================

    service_source = inspect.getsource(
        LocalWeeklyScheduleGenerationService
    )

    tests.append((
        "OWG22 Generation consumes canonical source data",
        "source_data"
        in service_source,
    ))

    tests.append((
        "OWG23 Generation never parses workbook bytes",
        not any(
            token
            in service_source
            for token in (
                "workbook_bytes",
                "load_workbook",
                "BytesIO",
                "openpyxl",
            )
        ),
    ))

    tests.append((
        "OWG24 Existing weekly schedule engine reused",
        "WeeklyTeachingScheduleService"
        in service_source,
    ))

    tests.append((
        "OWG25 No persistence dependency introduced",
        not any(
            token
            in service_source.lower()
            for token in (
                "supabase",
                "sqlite3",
                "googleapiclient",
                "streamlit",
            )
        ),
    ))

    tests.append((
        "OWG26 No fixed educational values encoded",
        not any(
            token
            in service_source
            for token in (
                "140",
                "105",
                "70",
                "35",
                "KNTT",
                "Toán 6",
            )
        ),
    ))

    results = []

    for label, passed in tests:
        results.append(passed)

        print(
            f"{label}: "
            f"{'PASS' if passed else 'FAIL'}"
        )

    print()

    print("E2E SUMMARY")

    print(
        "SCHEDULE ID :",
        result.schedule.schedule_id,
    )

    print(
        "TEACHER     :",
        result.schedule.teacher_id,
    )

    print(
        "ACADEMIC YEAR:",
        result.schedule.academic_week.academic_year,
    )

    print(
        "WEEK        :",
        result.schedule.academic_week.week_number,
    )

    print(
        "ENTRIES     :",
        len(result.schedule.entries),
    )

    print(
        "PPCT        :",
        [
            entry.curriculum_period
            for entry
            in result.schedule.entries
        ],
    )

    print()

    if all(results):
        print(
            "RESULT: PASS - LOCAL UPLOAD TO "
            "WEEKLY SCHEDULE E2E VERIFIED"
        )

        return True

    print(
        "RESULT: FAIL - LOCAL UPLOAD TO "
        "WEEKLY SCHEDULE E2E VIOLATED"
    )

    return False


def test_local_weekly_schedule_generation_e2e():
    assert run_contract()


def main():
    raise SystemExit(
        0 if run_contract() else 1
    )


if __name__ == "__main__":
    main()
