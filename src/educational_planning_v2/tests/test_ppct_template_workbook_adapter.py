from io import BytesIO

from openpyxl import load_workbook

from educational_planning_v2.adapters.ppct_template_workbook_adapter import (
    PPCTTemplateWorkbookAdapter,
)
from educational_planning_v2.adapters.ppct_workbook_upload_adapter import (
    PPCTWorkbookUploadAdapter,
)


def test_ppct_template_v1_contract():
    content = (
        PPCTTemplateWorkbookAdapter()
        .build()
    )

    assert isinstance(content, bytes)
    assert content

    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
        data_only=True,
    )

    worksheet = workbook["PPCT"]

    headers = tuple(
        cell.value
        for cell in next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    )

    workbook.close()

    assert headers == (
        "M\u00f4n/L\u1edbp",
        "Ph\u00e2n m\u00f4n",
        "Ti\u1ebft",
        "T\u00ean b\u00e0i h\u1ecdc",
    )


def test_ppct_parser_accepts_optional_sub_subject():
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(
        [
            "M\u00f4n/L\u1edbp",
            "Ph\u00e2n m\u00f4n",
            "Ti\u1ebft",
            "T\u00ean b\u00e0i h\u1ecdc",
        ]
    )

    worksheet.append(
        [
            "To\u00e1n 6",
            "S\u1ed1 h\u1ecdc",
            1,
            "B\u00e0i 1",
        ]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    rows = PPCTWorkbookUploadAdapter().parse(
        workbook_bytes=buffer.getvalue(),
    )

    assert len(rows) == 1
    assert rows[0].sub_subject == "S\u1ed1 h\u1ecdc"
