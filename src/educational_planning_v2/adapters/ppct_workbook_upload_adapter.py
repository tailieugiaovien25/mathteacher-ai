from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
import re
import unicodedata

from openpyxl import load_workbook

from educational_planning_v2.adapters.ppct_plan_item_adapter import (
    PPCTRow,
)
from educational_planning_v2.models.operational_data_source import (
    OperationalDataType,
)
from educational_planning_v2.models.operational_payload import (
    OperationalPayloadEnvelope,
    OperationalPayloadReference,
)


class PPCTWorkbookUploadAdapter:
    """
    Application-boundary adapter for teacher-uploaded PPCT workbooks.

    Owns physical workbook reading and source-column recognition only.
    It does not persist payloads, mutate lifecycle state, or own
    curriculum/YCCD business rules.
    """

    _SUBJECT_GRADE_ALIASES = {
        "mon lop",
        "subject grade",
    }

    _SUB_SUBJECT_ALIASES = {
        "phan mon",
        "sub subject",
        "subsubject",
        "subject component",
    }

    _PERIOD_ALIASES = {
        "tiet",
        "tiet ppct",
        "period",
        "period number",
    }

    _LESSON_ALIASES = {
        "ten bai",
        "ten bai hoc",
        "bai hoc",
        "noi dung",
        "lesson",
        "lesson name",
    }

    def parse(
        self,
        *,
        workbook_bytes: bytes,
    ) -> tuple[PPCTRow, ...]:
        if not isinstance(workbook_bytes, bytes):
            raise TypeError("workbook_bytes must be bytes")

        if not workbook_bytes:
            raise ValueError("workbook_bytes must not be empty")

        workbook = load_workbook(
            BytesIO(workbook_bytes),
            read_only=True,
            data_only=True,
        )

        try:
            for worksheet in workbook.worksheets:
                parsed = self._parse_worksheet(
                    worksheet
                )

                if parsed:
                    return parsed
        finally:
            workbook.close()

        raise ValueError(
            "could not locate PPCT columns in workbook"
        )

    def build_envelope(
        self,
        *,
        workbook_bytes: bytes,
        source_id: str,
        payload_version: str | None = None,
    ) -> OperationalPayloadEnvelope:
        rows = self.parse(
            workbook_bytes=workbook_bytes,
        )

        payload = tuple(
            asdict(row)
            for row in rows
        )

        return OperationalPayloadEnvelope(
            reference=OperationalPayloadReference(
                source_id=source_id,
                data_type=OperationalDataType.PPCT,
                payload_version=payload_version,
            ),
            payload=payload,
        )

    def _parse_worksheet(
        self,
        worksheet,
    ) -> tuple[PPCTRow, ...]:
        values = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        if not values:
            return ()

        header_index, columns = (
            self._find_header(values)
        )

        if header_index is None:
            return ()

        rows = []

        for raw_row in values[
            header_index + 1:
        ]:
            subject_grade = self._cell_text(
                raw_row,
                columns["subject_grade"],
            )

            sub_subject = None

            if "sub_subject" in columns:
                sub_subject = self._cell_text(
                    raw_row,
                    columns["sub_subject"],
                )

            lesson_name = self._cell_text(
                raw_row,
                columns["lesson_name"],
            )

            period_value = self._cell_value(
                raw_row,
                columns["period"],
            )

            if (
                subject_grade is None
                and lesson_name is None
                and period_value is None
            ):
                continue

            if (
                subject_grade is None
                or lesson_name is None
                or period_value is None
            ):
                continue

            period = self._period_number(
                period_value
            )

            rows.append(
                PPCTRow(
                    subject_grade=subject_grade,
                    period=period,
                    lesson_name=lesson_name,
                    sub_subject=sub_subject,
                )
            )

        if not rows:
            raise ValueError(
                "PPCT worksheet contains no valid rows"
            )

        return tuple(rows)

    def _find_header(
        self,
        rows,
    ) -> tuple[
        int | None,
        dict[str, int],
    ]:
        search_limit = min(
            len(rows),
            30,
        )

        for index in range(search_limit):
            mapping = {}

            for column_index, value in enumerate(
                rows[index]
            ):
                normalized = self._normalize_header(
                    value
                )

                if (
                    normalized
                    in self._SUBJECT_GRADE_ALIASES
                ):
                    mapping["subject_grade"] = (
                        column_index
                    )

                if (
                    normalized
                    in self._SUB_SUBJECT_ALIASES
                ):
                    mapping["sub_subject"] = (
                        column_index
                    )

                if (
                    normalized
                    in self._PERIOD_ALIASES
                ):
                    mapping["period"] = (
                        column_index
                    )

                if (
                    normalized
                    in self._LESSON_ALIASES
                ):
                    mapping["lesson_name"] = (
                        column_index
                    )

            required_columns = {
                "subject_grade",
                "period",
                "lesson_name",
            }

            if required_columns.issubset(
                mapping
            ):
                return index, mapping

        return None, {}

    @staticmethod
    def _normalize_header(
        value,
    ) -> str:
        if value is None:
            return ""

        text = str(value).strip().lower()

        text = "".join(
            character
            for character in unicodedata.normalize(
                "NFD",
                text,
            )
            if unicodedata.category(character)
            != "Mn"
        )

        text = text.replace(
            "?",
            "d",
        )

        text = re.sub(
            r"[^a-z0-9]+",
            " ",
            text,
        )

        text = re.sub(
            r"\s+",
            " ",
            text,
        ).strip()

        return text

    @staticmethod
    def _cell_value(
        row,
        index: int,
    ):
        if index >= len(row):
            return None

        return row[index]

    @classmethod
    def _cell_text(
        cls,
        row,
        index: int,
    ) -> str | None:
        value = cls._cell_value(
            row,
            index,
        )

        if value is None:
            return None

        normalized = str(value).strip()

        return normalized or None

    @staticmethod
    def _period_number(
        value,
    ) -> int:
        if isinstance(value, bool):
            raise ValueError(
                "PPCT period must be an integer"
            )

        if isinstance(value, int):
            period = value

        elif (
            isinstance(value, float)
            and value.is_integer()
        ):
            period = int(value)

        elif isinstance(value, str):
            normalized = value.strip()

            if not normalized.isdigit():
                raise ValueError(
                    "PPCT period must be an integer"
                )

            period = int(normalized)

        else:
            raise ValueError(
                "PPCT period must be an integer"
            )

        if period <= 0:
            raise ValueError(
                "PPCT period must be greater than 0"
            )

        return period
