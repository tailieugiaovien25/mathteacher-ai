from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from os import PathLike
from typing import BinaryIO, Mapping

from openpyxl import load_workbook

from educational_planning_v2.models.weekly_teaching_schedule import (
    AcademicWeek,
    CurriculumPeriod,
    LessonExecutionRecord,
    TimetableSlot,
)


@dataclass(frozen=True)
class WeeklyScheduleWorkbookSchema:
    """Map a changeable workbook layout to stable domain field names."""

    academic_weeks_sheet: str = "Tuan_hoc"
    timetable_sheet: str = "Thoi_khoa_bieu"
    curriculum_sheet: str = "PPCT"
    executions_sheet: str = "Tiet_da_day"
    academic_week_columns: Mapping[str, str] = field(
        default_factory=lambda: {
            "academic_year": "nam_hoc",
            "week_number": "tuan",
            "start_date": "tu_ngay",
            "end_date": "den_ngay",
        }
    )
    timetable_columns: Mapping[str, str] = field(
        default_factory=lambda: {
            "teacher_id": "ma_giao_vien",
            "class_id": "lop",
            "subject_ref": "mon_hoc",
            "component_ref": "phan_mon",
            "weekday": "thu",
            "timetable_period": "tiet_hoc",
            "effective_from": "hieu_luc_tu",
            "effective_to": "hieu_luc_den",
        }
    )
    curriculum_columns: Mapping[str, str] = field(
        default_factory=lambda: {
            "class_id": "lop",
            "subject_ref": "mon_hoc",
            "component_ref": "phan_mon",
            "period_number": "tiet_ppct",
            "lesson_id": "ma_bai_hoc",
            "lesson_title": "ten_bai_hoc",
            "period_in_lesson": "tiet_trong_bai",
            "total_lesson_periods": "tong_tiet_cua_bai",
            "teaching_equipment": "thiet_bi_day_hoc",
        }
    )
    execution_columns: Mapping[str, str] = field(
        default_factory=lambda: {
            "teacher_id": "ma_giao_vien",
            "class_id": "lop",
            "subject_ref": "mon_hoc",
            "component_ref": "phan_mon",
            "teaching_date": "ngay_day",
            "curriculum_period": "tiet_ppct",
            "status": "trang_thai",
        }
    )


@dataclass(frozen=True)
class WeeklyScheduleSourceData:
    academic_weeks: tuple[AcademicWeek, ...]
    timetable_slots: tuple[TimetableSlot, ...]
    curriculum_periods: tuple[CurriculumPeriod, ...]
    execution_records: tuple[LessonExecutionRecord, ...]

    def week(
        self,
        week_number: int,
        academic_year: str | None = None,
    ) -> AcademicWeek:
        matches = tuple(
            item
            for item in self.academic_weeks
            if item.week_number == week_number
            and (academic_year is None or item.academic_year == academic_year.strip())
        )
        if not matches:
            raise ValueError(f"khong tim thay tuan {week_number}")
        if len(matches) > 1:
            raise ValueError(f"tuan {week_number} bi trung trong du lieu")
        return matches[0]


class WeeklyScheduleWorkbookError(ValueError):
    """Workbook error with a user-facing sheet/row/column location."""

    def __init__(
        self,
        message: str,
        *,
        sheet: str | None = None,
        row: int | None = None,
        column: str | None = None,
    ) -> None:
        location = ", ".join(
            part
            for part in (
                f"sheet={sheet!r}" if sheet else None,
                f"row={row}" if row else None,
                f"column={column!r}" if column else None,
            )
            if part
        )
        super().__init__(f"{location}: {message}" if location else message)
        self.sheet = sheet
        self.row = row
        self.column = column


class WeeklyScheduleExcelAdapter:
    """Read a workbook and emit canonical weekly-schedule domain objects."""

    _DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y")
    _WEEKDAY_ALIASES = {
        "2": 1, "thu 2": 1, "thứ 2": 1, "thu hai": 1, "thứ hai": 1,
        "3": 2, "thu 3": 2, "thứ 3": 2, "thu ba": 2, "thứ ba": 2,
        "4": 3, "thu 4": 3, "thứ 4": 3, "thu tu": 3, "thứ tư": 3,
        "5": 4, "thu 5": 4, "thứ 5": 4, "thu nam": 4, "thứ năm": 4,
        "6": 5, "thu 6": 5, "thứ 6": 5, "thu sau": 5, "thứ sáu": 5,
        "7": 6, "thu 7": 6, "thứ 7": 6, "thu bay": 6, "thứ bảy": 6,
        "chu nhat": 7, "chủ nhật": 7, "cn": 7,
    }
    _COMPLETED_ALIASES = {
        "completed", "da day", "đã dạy", "hoan thanh", "hoàn thành"
    }

    def __init__(self, schema: WeeklyScheduleWorkbookSchema | None = None) -> None:
        self._schema = schema or WeeklyScheduleWorkbookSchema()

    def load(
        self, source: str | PathLike[str] | BinaryIO
    ) -> WeeklyScheduleSourceData:
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            weeks = self._read_rows(
                workbook, self._schema.academic_weeks_sheet,
                self._schema.academic_week_columns, self._academic_week,
            )
            slots = self._read_rows(
                workbook, self._schema.timetable_sheet,
                self._schema.timetable_columns, self._timetable_slot,
            )
            curriculum = self._read_rows(
                workbook, self._schema.curriculum_sheet,
                self._schema.curriculum_columns, self._curriculum_period,
            )
            executions = self._read_rows(
                workbook, self._schema.executions_sheet,
                self._schema.execution_columns, self._execution_record,
            )
        finally:
            workbook.close()
        return WeeklyScheduleSourceData(
            academic_weeks=weeks,
            timetable_slots=slots,
            curriculum_periods=curriculum,
            execution_records=executions,
        )

    def _read_rows(self, workbook, sheet_name, columns, factory):
        if sheet_name not in workbook.sheetnames:
            raise WeeklyScheduleWorkbookError(
                "khong tim thay bang du lieu", sheet=sheet_name
            )
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            raise WeeklyScheduleWorkbookError("bang du lieu rong", sheet=sheet_name)
        headers = {
            self._text(value).casefold(): index
            for index, value in enumerate(header_row)
            if self._text(value)
        }
        missing = tuple(
            physical for physical in columns.values()
            if physical.casefold() not in headers
        )
        if missing:
            raise WeeklyScheduleWorkbookError(
                "thieu cot bat buoc: " + ", ".join(missing), sheet=sheet_name
            )
        result = []
        for row_number, values in enumerate(iterator, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            logical = {
                name: values[headers[physical.casefold()]]
                if headers[physical.casefold()] < len(values) else None
                for name, physical in columns.items()
            }
            try:
                result.append(factory(logical))
            except WeeklyScheduleWorkbookError as error:
                raise WeeklyScheduleWorkbookError(
                    str(error), sheet=sheet_name, row=row_number,
                    column=error.column,
                ) from error
            except (TypeError, ValueError) as error:
                raise WeeklyScheduleWorkbookError(
                    str(error), sheet=sheet_name, row=row_number
                ) from error
        return tuple(result)

    def _academic_week(self, row):
        return AcademicWeek(
            academic_year=self._required_text(row["academic_year"], "nam_hoc"),
            week_number=self._positive_int(row["week_number"], "tuan"),
            start_date=self._date(row["start_date"], "tu_ngay"),
            end_date=self._date(row["end_date"], "den_ngay"),
        )

    def _timetable_slot(self, row):
        return TimetableSlot(
            teacher_id=self._required_text(row["teacher_id"], "ma_giao_vien"),
            class_id=self._required_text(row["class_id"], "lop"),
            subject_ref=self._required_text(row["subject_ref"], "mon_hoc"),
            component_ref=self._optional_text(row["component_ref"]),
            weekday=self._weekday(row["weekday"]),
            timetable_period=self._positive_int(row["timetable_period"], "tiet_hoc"),
            effective_from=self._date(row["effective_from"], "hieu_luc_tu"),
            effective_to=self._date(row["effective_to"], "hieu_luc_den"),
        )

    def _curriculum_period(self, row):
        return CurriculumPeriod(
            class_id=self._required_text(row["class_id"], "lop"),
            subject_ref=self._required_text(row["subject_ref"], "mon_hoc"),
            component_ref=self._optional_text(row["component_ref"]),
            period_number=self._positive_int(row["period_number"], "tiet_ppct"),
            lesson_id=self._required_text(row["lesson_id"], "ma_bai_hoc"),
            lesson_title=self._required_text(row["lesson_title"], "ten_bai_hoc"),
            period_in_lesson=self._positive_int(
                row["period_in_lesson"] or 1, "tiet_trong_bai"
            ),
            total_lesson_periods=self._positive_int(
                row["total_lesson_periods"] or 1, "tong_tiet_cua_bai"
            ),
            teaching_equipment=self._equipment(row["teaching_equipment"]),
        )

    def _execution_record(self, row):
        status = self._required_text(row["status"], "trang_thai")
        if status.casefold() in self._COMPLETED_ALIASES:
            status = "COMPLETED"
        return LessonExecutionRecord(
            teacher_id=self._required_text(row["teacher_id"], "ma_giao_vien"),
            class_id=self._required_text(row["class_id"], "lop"),
            subject_ref=self._required_text(row["subject_ref"], "mon_hoc"),
            component_ref=self._optional_text(row["component_ref"]),
            teaching_date=self._date(row["teaching_date"], "ngay_day"),
            curriculum_period=self._positive_int(
                row["curriculum_period"], "tiet_ppct"
            ),
            status=status,
        )

    @classmethod
    def _date(cls, value, column):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = cls._required_text(value, column)
        for date_format in cls._DATE_FORMATS:
            try:
                return datetime.strptime(text, date_format).date()
            except ValueError:
                continue
        raise WeeklyScheduleWorkbookError(
            "ngay khong hop le; dung yyyy-mm-dd hoac dd/mm/yyyy", column=column
        )

    @classmethod
    def _weekday(cls, value):
        if isinstance(value, int) and not isinstance(value, bool):
            if value == 1:
                return 1
            if 2 <= value <= 7:
                return value - 1
        normalized = cls._text(value).casefold()
        if normalized in cls._WEEKDAY_ALIASES:
            return cls._WEEKDAY_ALIASES[normalized]
        raise WeeklyScheduleWorkbookError(
            "thu khong hop le; dung 2-7 hoac Chu nhat", column="thu"
        )

    @staticmethod
    def _positive_int(value, column):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        if not isinstance(value, int) or isinstance(value, bool) or value <= 0:
            raise WeeklyScheduleWorkbookError(
                "phai la so nguyen duong", column=column
            )
        return value

    @classmethod
    def _required_text(cls, value, column):
        text = cls._text(value)
        if not text:
            raise WeeklyScheduleWorkbookError("khong duoc de trong", column=column)
        return text

    @staticmethod
    def _optional_text(value):
        return WeeklyScheduleExcelAdapter._text(value) or None

    @staticmethod
    def _text(value):
        return "" if value is None else str(value).strip()

    @classmethod
    def _equipment(cls, value):
        return tuple(
            item.strip()
            for item in cls._text(value).replace(",", ";").split(";")
            if item.strip()
        )
