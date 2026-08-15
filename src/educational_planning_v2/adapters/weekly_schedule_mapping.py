"""Inspect changeable workbooks and persist user-defined source mappings."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from io import BytesIO
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from educational_planning_v2.adapters.weekly_schedule_excel_adapter import WeeklyScheduleWorkbookSchema


@dataclass(frozen=True)
class WorkbookSheetInspection:
    name: str
    max_row: int
    max_column: int
    preview_rows: tuple[tuple[Any, ...], ...]


@dataclass(frozen=True)
class WeeklyScheduleMappingProfile:
    profile_name: str
    schema: WeeklyScheduleWorkbookSchema


class WeeklyScheduleWorkbookInspector:
    def inspect(self, content: bytes, preview_limit: int = 20) -> tuple[WorkbookSheetInspection, ...]:
        if not content:
            raise ValueError("workbook content must not be empty")
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            return tuple(
                WorkbookSheetInspection(
                    name=sheet.title,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    preview_rows=tuple(tuple(row) for row in sheet.iter_rows(max_row=preview_limit, values_only=True)),
                )
                for sheet in workbook.worksheets
            )
        finally:
            workbook.close()

    @staticmethod
    def headers(inspection: WorkbookSheetInspection, header_row: int) -> tuple[str, ...]:
        if not isinstance(header_row, int) or isinstance(header_row, bool) or header_row <= 0:
            raise ValueError("header_row must be a positive integer")
        if header_row > len(inspection.preview_rows):
            raise ValueError("header_row is outside the inspected preview")
        values = inspection.preview_rows[header_row - 1]
        headers = tuple(str(value).strip() for value in values if value not in (None, "") and str(value).strip())
        if len({item.casefold() for item in headers}) != len(headers):
            raise ValueError("header row contains duplicate column names")
        return headers


class LocalWeeklyScheduleMappingRepository:
    _SAFE_NAME = re.compile(r"^[A-Za-z0-9À-ỹ][A-Za-z0-9À-ỹ ._-]{0,99}$")

    def __init__(self, storage_root: str | Path) -> None:
        self._root = Path(storage_root)

    def save(self, profile: WeeklyScheduleMappingProfile) -> WeeklyScheduleMappingProfile:
        if not isinstance(profile, WeeklyScheduleMappingProfile):
            raise TypeError("profile must be a WeeklyScheduleMappingProfile")
        path = self._path(profile.profile_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        payload = {"schema_version": 1, "profile_name": profile.profile_name, "schema": asdict(profile.schema)}
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return profile

    def get(self, profile_name: str) -> WeeklyScheduleMappingProfile | None:
        path = self._path(profile_name)
        if not path.exists():
            return None
        payload = json.loads(path.read_text(encoding="utf-8"))
        if payload.get("schema_version") != 1:
            raise ValueError("unsupported mapping profile schema")
        return WeeklyScheduleMappingProfile(
            profile_name=payload["profile_name"],
            schema=WeeklyScheduleWorkbookSchema(**payload["schema"]),
        )

    def list_names(self) -> tuple[str, ...]:
        if not self._root.exists():
            return ()
        return tuple(sorted(path.stem for path in self._root.glob("*.json")))

    def _path(self, profile_name: str) -> Path:
        if not isinstance(profile_name, str):
            raise TypeError("profile_name must be a string")
        normalized = profile_name.strip()
        if not self._SAFE_NAME.fullmatch(normalized):
            raise ValueError("profile_name contains unsafe characters")
        return self._root / f"{normalized}.json"
