from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from models.yccd_period_map_record import (
    YCCDPeriodMapRecord,
)


class YCCDPeriodMapWriter:
    """Ghi mapping YCCĐ - tiết vào tblYCCDPeriodMap."""

    SHEET_NAME = "YCCD_PERIOD_MAP"
    TABLE_NAME = "tblYCCDPeriodMap"

    HEADERS = [
        "MAP_ID",
        "LESSON_KEY",
        "TIET_TRONG_BAI",
        "YCCD_ID",
        "VAI_TRO",
        "PHIEN_BAN",
        "TRANG_THAI",
        "NGAY_CAP_NHAT",
        "GHI_CHU",
    ]

    def append_records(
        self,
        file_path: str | Path,
        records: Iterable[YCCDPeriodMapRecord],
    ) -> int:

        normalized_path = Path(file_path)
        records_list = list(records)

        if not normalized_path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy workbook: {normalized_path}"
            )

        if not records_list:
            return 0

        # Validate trước khi mở workbook để ghi.
        for record in records_list:
            record.validate()

        # Không cho trùng MAP_ID ngay trong batch.
        batch_ids = [
            record.map_id
            for record in records_list
        ]

        if len(batch_ids) != len(set(batch_ids)):
            raise ValueError(
                "Có MAP_ID trùng trong dữ liệu cần ghi."
            )

        workbook = load_workbook(
            filename=normalized_path,
            read_only=False,
            data_only=False,
            keep_vba=(
                normalized_path.suffix.lower()
                == ".xlsm"
            ),
        )

        try:
            if self.SHEET_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"Không tìm thấy worksheet "
                    f"{self.SHEET_NAME}."
                )

            worksheet = workbook[self.SHEET_NAME]

            if self.TABLE_NAME not in worksheet.tables:
                raise ValueError(
                    f"Không tìm thấy table "
                    f"{self.TABLE_NAME}."
                )

            table = worksheet.tables[self.TABLE_NAME]

            # Kiểm tra schema.
            actual_headers = []

            for column_index in range(
                1,
                len(self.HEADERS) + 1,
            ):
                value = worksheet.cell(
                    row=1,
                    column=column_index,
                ).value

                actual_headers.append(
                    str(value).strip()
                    if value is not None
                    else ""
                )

            if actual_headers != self.HEADERS:
                raise ValueError(
                    "Schema tblYCCDPeriodMap "
                    "không đúng 9 cột chuẩn."
                )

            # Đọc MAP_ID hiện có.
            existing_ids: set[str] = set()

            for row_index in range(
                2,
                worksheet.max_row + 1,
            ):
                value = worksheet.cell(
                    row=row_index,
                    column=1,
                ).value

                if value not in (None, ""):
                    existing_ids.add(
                        str(value).strip()
                    )

            # Không cho ghi trùng MAP_ID.
            for record in records_list:
                if record.map_id in existing_ids:
                    raise ValueError(
                        "MAP_ID đã tồn tại: "
                        f"{record.map_id}"
                    )

            # Tìm hàng trống đầu tiên.
            next_row = 2

            while (
                worksheet.cell(
                    row=next_row,
                    column=1,
                ).value
                not in (None, "")
            ):
                next_row += 1

            # Ghi dữ liệu.
            for record in records_list:
                excel_row = record.to_excel_row()

                for column_index, header in enumerate(
                    self.HEADERS,
                    start=1,
                ):
                    worksheet.cell(
                        row=next_row,
                        column=column_index,
                        value=excel_row[header],
                    )

                next_row += 1

            # Mở rộng Excel Table.
            last_data_row = next_row - 1

            table.ref = (
                f"A1:I{last_data_row}"
            )

            # Chỉ save sau khi toàn bộ kiểm tra đã PASS.
            workbook.save(
                normalized_path
            )

            return len(records_list)

        finally:
            workbook.close()