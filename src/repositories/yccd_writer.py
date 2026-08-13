from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from models.yccd_record import YCCDRecord
from utils.yccd_provenance import validate_provenance


class YCCDWriter:
    """Ghi YCCDRecord vào tblYCCD theo cách có kiểm soát."""

    SHEET_NAME = "YCCD"
    TABLE_NAME = "tblYCCD"

    HEADERS = [
        "YCCD_ID",
        "LESSON_KEY",
        "MON",
        "KHOI",
        "BAI_ID",
        "TEN_BAI",
        "YCCD_ORDER",
        "YEU_CAU_CAN_DAT",
        "LOAI_YCCD",
        "YCCD_GOC_ID",
        "NGUON",
        "THAM_CHIEU",
        "PHIEN_BAN",
        "TRANG_THAI",
        "NGAY_CAP_NHAT",
        "GHI_CHU",
    ]

    def append_records(
        self,
        file_path: str | Path,
        records: Iterable[YCCDRecord],
    ) -> int:
        """
        Append các YCCDRecord vào tblYCCD.

        Trả về số record đã ghi.
        """

        normalized_path = Path(file_path)
        records_list = list(records)

        if not normalized_path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy workbook: {normalized_path}"
            )

        if not records_list:
            return 0

        # -----------------------------------------------------
        # 1. Validate dữ liệu trước khi mở workbook để ghi
        # -----------------------------------------------------

        validate_provenance(
            records_list
        )

        workbook = load_workbook(
            filename=normalized_path,
            read_only=False,
            data_only=False,
            keep_vba=(
                normalized_path.suffix.lower()
                == ".xlsm"
            ),
        )

        try:
            # -------------------------------------------------
            # 2. Kiểm tra worksheet
            # -------------------------------------------------

            if (
                self.SHEET_NAME
                not in workbook.sheetnames
            ):
                raise ValueError(
                    f"Không tìm thấy worksheet "
                    f"{self.SHEET_NAME}."
                )

            worksheet = workbook[
                self.SHEET_NAME
            ]

            # -------------------------------------------------
            # 3. Kiểm tra table
            # -------------------------------------------------

            if (
                self.TABLE_NAME
                not in worksheet.tables
            ):
                raise ValueError(
                    f"Không tìm thấy table "
                    f"{self.TABLE_NAME}."
                )

            table = worksheet.tables[
                self.TABLE_NAME
            ]

            # -------------------------------------------------
            # 4. Kiểm tra schema 16 cột
            # -------------------------------------------------

            actual_headers = []

            for column_index in range(
                1,
                len(self.HEADERS) + 1,
            ):
                value = worksheet.cell(
                    row=1,
                    column=column_index,
                ).value

                actual_headers.append(
                    str(value).strip()
                    if value is not None
                    else ""
                )

            if (
                actual_headers
                != self.HEADERS
            ):
                raise ValueError(
                    "Schema tblYCCD không đúng "
                    "16 cột chuẩn."
                )

            # -------------------------------------------------
            # 5. Lấy toàn bộ YCCD_ID hiện có
            # -------------------------------------------------

            existing_ids: set[str] = set()

            for row_index in range(
                2,
                worksheet.max_row + 1,
            ):
                value = worksheet.cell(
                    row=row_index,
                    column=1,
                ).value

                if value not in (
                    None,
                    "",
                ):
                    existing_ids.add(
                        str(value).strip()
                    )

            # -------------------------------------------------
            # 6. Không cho ghi trùng ID
            # -------------------------------------------------

            for record in records_list:
                if (
                    record.yccd_id
                    in existing_ids
                ):
                    raise ValueError(
                        "YCCD_ID đã tồn tại "
                        "trong workbook: "
                        f"{record.yccd_id}"
                    )

            # -------------------------------------------------
            # 7. Xác định hàng bắt đầu ghi
            # -------------------------------------------------

            next_row = 2

            while (
                worksheet.cell(
                    row=next_row,
                    column=1,
                ).value
                not in (
                    None,
                    "",
                )
            ):
                next_row += 1

            # -------------------------------------------------
            # 8. Ghi dữ liệu
            # -------------------------------------------------

            for record in records_list:
                excel_row = (
                    record.to_excel_row()
                )

                for (
                    column_index,
                    header,
                ) in enumerate(
                    self.HEADERS,
                    start=1,
                ):
                    worksheet.cell(
                        row=next_row,
                        column=column_index,
                        value=excel_row[
                            header
                        ],
                    )

                next_row += 1

            # -------------------------------------------------
            # 9. Mở rộng phạm vi tblYCCD
            # -------------------------------------------------

            last_data_row = (
                next_row - 1
            )

            table.ref = (
                f"A1:P{last_data_row}"
            )

            # -------------------------------------------------
            # 10. Chỉ save sau khi mọi bước đã thành công
            # -------------------------------------------------

            workbook.save(
                normalized_path
            )

            return len(
                records_list
            )

        finally:
            workbook.close()