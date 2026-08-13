from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from models.yccd_period_map_record import (
    YCCDPeriodMapRecord,
)


class YCCDPeriodMapRepository:
    """Đọc dữ liệu từ tblYCCDPeriodMap."""

    SHEET_NAME = "YCCD_PERIOD_MAP"

    REQUIRED_HEADERS = {
        "MAP_ID",
        "LESSON_KEY",
        "TIET_TRONG_BAI",
        "YCCD_ID",
        "VAI_TRO",
        "PHIEN_BAN",
        "TRANG_THAI",
        "NGAY_CAP_NHAT",
        "GHI_CHU",
    }

    def load_rows(
        self,
        file_path: str | Path,
    ) -> list[dict[str, Any]]:
        """Đọc toàn bộ mapping có dữ liệu."""

        normalized_path = Path(
            file_path
        )

        if not normalized_path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy workbook: "
                f"{normalized_path}"
            )

        workbook = load_workbook(
            filename=normalized_path,
            read_only=False,
            data_only=True,
            keep_vba=(
                normalized_path.suffix.lower()
                == ".xlsm"
            ),
        )

        try:
            if (
                self.SHEET_NAME
                not in workbook.sheetnames
            ):
                raise ValueError(
                    "Không tìm thấy worksheet "
                    f"{self.SHEET_NAME}."
                )

            worksheet = workbook[
                self.SHEET_NAME
            ]

            headers = self._read_headers(
                worksheet
            )

            missing_headers = (
                self.REQUIRED_HEADERS
                - set(
                    headers.values()
                )
            )

            if missing_headers:
                raise ValueError(
                    "Thiếu các cột "
                    "YCCD_PERIOD_MAP: "
                    + ", ".join(
                        sorted(
                            missing_headers
                        )
                    )
                )

            rows: list[
                dict[str, Any]
            ] = []

            for row_index in range(
                2,
                worksheet.max_row + 1,
            ):
                row_data: dict[
                    str,
                    Any,
                ] = {}

                has_data = False

                for (
                    column_index,
                    header,
                ) in headers.items():
                    value = worksheet.cell(
                        row=row_index,
                        column=column_index,
                    ).value

                    row_data[
                        header
                    ] = value

                    if value not in (
                        None,
                        "",
                    ):
                        has_data = True

                if has_data:
                    rows.append(
                        row_data
                    )

            return rows

        finally:
            workbook.close()

    def find_by_lesson_key(
        self,
        file_path: str | Path,
        lesson_key: str,
        status: str = "draft",
    ) -> list[YCCDPeriodMapRecord]:
        """Lấy mapping của toàn bài."""

        rows = self.load_rows(
            file_path
        )

        normalized_key = (
            self._normalize(
                lesson_key
            )
        )

        normalized_status = (
            self._normalize(
                status
            )
        )

        records: list[
            YCCDPeriodMapRecord
        ] = []

        for row in rows:
            if (
                self._normalize(
                    row.get(
                        "LESSON_KEY"
                    )
                )
                != normalized_key
            ):
                continue

            if (
                self._normalize(
                    row.get(
                        "TRANG_THAI"
                    )
                )
                != normalized_status
            ):
                continue

            record = (
                self._row_to_record(
                    row
                )
            )

            record.validate()

            records.append(
                record
            )

        records.sort(
            key=lambda item: (
                int(
                    item.period_in_lesson
                ),
                item.map_id,
            )
        )

        return records

    def find_by_period(
        self,
        file_path: str | Path,
        lesson_key: str,
        period_in_lesson: int,
        status: str = "draft",
    ) -> list[YCCDPeriodMapRecord]:
        """Lấy mapping của một tiết cụ thể."""

        all_records = (
            self.find_by_lesson_key(
                file_path=file_path,
                lesson_key=lesson_key,
                status=status,
            )
        )

        try:
            period_number = int(
                period_in_lesson
            )
        except (
            TypeError,
            ValueError,
        ) as error:
            raise ValueError(
                "period_in_lesson "
                "phải là số nguyên."
            ) from error

        if period_number <= 0:
            raise ValueError(
                "period_in_lesson "
                "phải lớn hơn 0."
            )

        return [
            record
            for record in all_records
            if (
                int(
                    record.period_in_lesson
                )
                == period_number
            )
        ]

    def _row_to_record(
        self,
        row: dict[str, Any],
    ) -> YCCDPeriodMapRecord:
        return YCCDPeriodMapRecord(
            map_id=self._text(
                row.get("MAP_ID")
            ),
            lesson_key=self._text(
                row.get(
                    "LESSON_KEY"
                )
            ),
            period_in_lesson=(
                self._to_integer(
                    row.get(
                        "TIET_TRONG_BAI"
                    )
                )
            ),
            yccd_id=self._text(
                row.get(
                    "YCCD_ID"
                )
            ),
            role=self._text(
                row.get(
                    "VAI_TRO"
                )
            ),
            version=(
                self._optional_text(
                    row.get(
                        "PHIEN_BAN"
                    )
                )
                or "1.0"
            ),
            status=(
                self._optional_text(
                    row.get(
                        "TRANG_THAI"
                    )
                )
                or "draft"
            ),
            updated_at=row.get(
                "NGAY_CAP_NHAT"
            ),
            note=self._optional_text(
                row.get(
                    "GHI_CHU"
                )
            ),
        )

    @staticmethod
    def _read_headers(
        worksheet,
    ) -> dict[int, str]:
        headers: dict[
            int,
            str,
        ] = {}

        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):
            value = worksheet.cell(
                row=1,
                column=column_index,
            ).value

            if value is None:
                continue

            header = (
                str(value)
                .strip()
                .upper()
            )

            if header:
                headers[
                    column_index
                ] = header

        return headers

    @staticmethod
    def _normalize(
        value: Any,
    ) -> str:
        if value is None:
            return ""

        return " ".join(
            str(value)
            .strip()
            .lower()
            .split()
        )

    @staticmethod
    def _text(
        value: Any,
    ) -> str:
        if value is None:
            return ""

        return " ".join(
            str(value)
            .strip()
            .split()
        )

    @classmethod
    def _optional_text(
        cls,
        value: Any,
    ) -> str | None:
        text = cls._text(
            value
        )

        return (
            text
            if text
            else None
        )

    @staticmethod
    def _to_integer(
        value: Any,
    ) -> int:
        try:
            return int(
                value
            )
        except (
            TypeError,
            ValueError,
        ) as error:
            raise ValueError(
                "TIET_TRONG_BAI "
                "phải là số nguyên."
            ) from error