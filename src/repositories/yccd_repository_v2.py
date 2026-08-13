from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from models.yccd_record import YCCDRecord


class YCCDRepositoryV2:
    """Repository YCCD theo schema 16 cột và LESSON_KEY."""

    SHEET_NAME = "YCCD"

    REQUIRED_HEADERS = {
        "YCCD_ID",
        "LESSON_KEY",
        "MON",
        "KHOI",
        "BAI_ID",
        "TEN_BAI",
        "YCCD_ORDER",
        "YEU_CAU_CAN_DAT",
        "LOAI_YCCD",
        "YCCD_GOC_ID",
        "NGUON",
        "THAM_CHIEU",
        "PHIEN_BAN",
        "TRANG_THAI",
        "NGAY_CAP_NHAT",
        "GHI_CHU",
    }

    def load_rows(
        self,
        file_path: str | Path,
    ) -> list[dict[str, Any]]:
        """Đọc các dòng dữ liệu từ sheet YCCD."""

        normalized_path = Path(file_path)

        if not normalized_path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy file Excel: {normalized_path}"
            )

        workbook = load_workbook(
            filename=normalized_path,
            read_only=False,
            data_only=True,
            keep_vba=(
                normalized_path.suffix.lower()
                == ".xlsm"
            ),
        )

        try:
            if (
                self.SHEET_NAME
                not in workbook.sheetnames
            ):
                raise ValueError(
                    f"Không tìm thấy worksheet: "
                    f"{self.SHEET_NAME}"
                )

            worksheet = workbook[
                self.SHEET_NAME
            ]

            headers = self._read_headers(
                worksheet
            )

            missing_headers = (
                self.REQUIRED_HEADERS
                - set(headers.values())
            )

            if missing_headers:
                raise ValueError(
                    "Thiếu các cột YCCD: "
                    + ", ".join(
                        sorted(
                            missing_headers
                        )
                    )
                )

            rows: list[
                dict[str, Any]
            ] = []

            for row_index in range(
                2,
                worksheet.max_row + 1,
            ):
                row_data: dict[
                    str,
                    Any,
                ] = {}

                has_data = False

                for (
                    column_index,
                    header,
                ) in headers.items():
                    value = worksheet.cell(
                        row=row_index,
                        column=column_index,
                    ).value

                    row_data[
                        header
                    ] = value

                    if value not in (
                        None,
                        "",
                    ):
                        has_data = True

                if has_data:
                    rows.append(
                        row_data
                    )

            return rows

        finally:
            workbook.close()

    def find_by_lesson_key(
        self,
        file_path: str | Path,
        lesson_key: str,
        status: str = "approved",
    ) -> list[YCCDRecord]:
        """Tìm YCCĐ theo LESSON_KEY."""

        rows = self.load_rows(
            file_path
        )

        normalized_key = (
            self._normalize(
                lesson_key
            )
        )

        normalized_status = (
            self._normalize(
                status
            )
        )

        matched_records: list[
            YCCDRecord
        ] = []

        for row in rows:
            if (
                self._normalize(
                    row.get(
                        "LESSON_KEY"
                    )
                )
                != normalized_key
            ):
                continue

            if (
                self._normalize(
                    row.get(
                        "TRANG_THAI"
                    )
                )
                != normalized_status
            ):
                continue

            record = (
                self._row_to_record(
                    row
                )
            )

            record.validate()

            matched_records.append(
                record
            )

        matched_records.sort(
            key=lambda item: (
                int(item.order)
            )
        )

        return matched_records

    def _row_to_record(
        self,
        row: dict[str, Any],
    ) -> YCCDRecord:
        """Chuyển một dòng Excel thành YCCDRecord."""

        return YCCDRecord(
            yccd_id=self._text(
                row.get("YCCD_ID")
            ),
            lesson_key=self._text(
                row.get(
                    "LESSON_KEY"
                )
            ),
            subject=self._text(
                row.get("MON")
            ),
            grade=row.get(
                "KHOI"
            ),
            lesson_id=(
                self._optional_text(
                    row.get(
                        "BAI_ID"
                    )
                )
            ),
            lesson_name=self._text(
                row.get(
                    "TEN_BAI"
                )
            ),
            order=self._to_integer(
                row.get(
                    "YCCD_ORDER"
                )
            ),
            requirement=self._text(
                row.get(
                    "YEU_CAU_CAN_DAT"
                )
            ),
            yccd_type=(
                self._optional_text(
                    row.get(
                        "LOAI_YCCD"
                    )
                )
                or "TRIEN_KHAI"
            ),
            source_yccd_id=(
                self._optional_text(
                    row.get(
                        "YCCD_GOC_ID"
                    )
                )
            ),
            source=self._optional_text(
                row.get(
                    "NGUON"
                )
            ),
            reference=self._optional_text(
                row.get(
                    "THAM_CHIEU"
                )
            ),
            version=(
                self._optional_text(
                    row.get(
                        "PHIEN_BAN"
                    )
                )
                or "1.0"
            ),
            status=(
                self._optional_text(
                    row.get(
                        "TRANG_THAI"
                    )
                )
                or "draft"
            ),
            updated_at=row.get(
                "NGAY_CAP_NHAT"
            ),
            note=self._optional_text(
                row.get(
                    "GHI_CHU"
                )
            ),
        )

    @staticmethod
    def _read_headers(
        worksheet,
    ) -> dict[int, str]:
        """Đọc header từ hàng 1."""

        headers: dict[
            int,
            str,
        ] = {}

        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):
            value = worksheet.cell(
                row=1,
                column=column_index,
            ).value

            if value is None:
                continue

            header = (
                str(value)
                .strip()
                .upper()
            )

            if header:
                headers[
                    column_index
                ] = header

        return headers

    @staticmethod
    def _normalize(
        value: Any,
    ) -> str:
        if value is None:
            return ""

        return " ".join(
            str(value)
            .strip()
            .lower()
            .split()
        )

    @staticmethod
    def _text(
        value: Any,
    ) -> str:
        if value is None:
            return ""

        return " ".join(
            str(value)
            .strip()
            .split()
        )

    @classmethod
    def _optional_text(
        cls,
        value: Any,
    ) -> str | None:
        text = cls._text(
            value
        )

        return (
            text
            if text
            else None
        )

    @staticmethod
    def _to_integer(
        value: Any,
    ) -> int:
        try:
            return int(value)
        except (
            TypeError,
            ValueError,
        ) as error:
            raise ValueError(
                "YCCD_ORDER phải là số nguyên."
            ) from error