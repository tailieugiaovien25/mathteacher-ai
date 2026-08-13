from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class YCCDRepository:
    """Đọc và tìm kiếm Yêu cầu cần đạt từ sheet YCCD."""

    SHEET_NAME = "YCCD"

    REQUIRED_HEADERS = {
        "YCCD_ID",
        "MON",
        "KHOI",
        "BAI_ID",
        "TEN_BAI",
        "TIET",
        "YCCD_ORDER",
        "YEU_CAU_CAN_DAT",
        "NGUON",
        "PHIEN_BAN",
        "TRANG_THAI",
        "NGAY_CAP_NHAT",
        "GHI_CHU",
    }

    def load_rows(
        self,
        file_path: str | Path,
    ) -> list[dict[str, Any]]:
        """Đọc toàn bộ dữ liệu YCCD thành danh sách dictionary."""

        normalized_path = Path(file_path)

        if not normalized_path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy file Excel: {normalized_path}"
            )

        workbook = load_workbook(
            filename=normalized_path,
            read_only=False,
            data_only=True,
            keep_vba=normalized_path.suffix.lower() == ".xlsm",
        )

        try:
            if self.SHEET_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"Không tìm thấy worksheet: {self.SHEET_NAME}"
                )

            worksheet = workbook[self.SHEET_NAME]

            headers = self._read_headers(worksheet)

            missing_headers = (
                self.REQUIRED_HEADERS - set(headers.values())
            )

            if missing_headers:
                raise ValueError(
                    "Thiếu các cột YCCD: "
                    + ", ".join(sorted(missing_headers))
                )

            rows: list[dict[str, Any]] = []

            for row_index in range(
                2,
                worksheet.max_row + 1,
            ):
                row_data: dict[str, Any] = {}

                has_data = False

                for column_index, header in headers.items():
                    value = worksheet.cell(
                        row=row_index,
                        column=column_index,
                    ).value

                    row_data[header] = value

                    if value not in (None, ""):
                        has_data = True

                if has_data:
                    rows.append(row_data)

            return rows

        finally:
            workbook.close()

    def find(
        self,
        file_path: str | Path,
        subject: str,
        grade: int | str,
        lesson_name: str,
        period: int | None = None,
        status: str = "approved",
    ) -> list[dict[str, Any]]:
        """Tìm các YCCD phù hợp với bài học."""

        rows = self.load_rows(file_path)

        normalized_subject = self._normalize(subject)
        normalized_grade = self._normalize(grade)
        normalized_lesson_name = self._normalize(lesson_name)
        normalized_status = self._normalize(status)

        matched_rows: list[dict[str, Any]] = []

        for row in rows:
            if self._normalize(row.get("MON")) != normalized_subject:
                continue

            if self._normalize(row.get("KHOI")) != normalized_grade:
                continue

            if (
                self._normalize(row.get("TEN_BAI"))
                != normalized_lesson_name
            ):
                continue

            if period is not None:
                if self._to_integer(row.get("TIET")) != period:
                    continue

            if self._normalize(
                row.get("TRANG_THAI")
            ) != normalized_status:
                continue

            matched_rows.append(row)

        matched_rows.sort(
            key=lambda item: (
                self._to_integer(
                    item.get("YCCD_ORDER")
                )
                or 0
            )
        )

        return matched_rows

    @staticmethod
    def _read_headers(
        worksheet,
    ) -> dict[int, str]:
        """Đọc tên cột tại hàng 1."""

        headers: dict[int, str] = {}

        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):
            value = worksheet.cell(
                row=1,
                column=column_index,
            ).value

            if value is None:
                continue

            header = str(value).strip().upper()

            if header:
                headers[column_index] = header

        return headers

    @staticmethod
    def _normalize(value: Any) -> str:
        if value is None:
            return ""

        return " ".join(
            str(value).strip().lower().split()
        )

    @staticmethod
    def _to_integer(
        value: Any,
    ) -> int | None:
        if value is None or value == "":
            return None

        try:
            return int(value)
        except (TypeError, ValueError):
            return None