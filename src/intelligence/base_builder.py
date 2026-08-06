from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.base_model import BaseModel


class BaseBuilder(ABC):
    """Lớp nền cho các Builder của AI Teacher Platform."""

    def validate_file(
        self,
        file_path: str | Path,
    ) -> Path:
        """Kiểm tra tệp tồn tại và trả về Path đã chuẩn hóa."""
        normalized_path = Path(file_path)

        if not normalized_path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy tệp: {normalized_path}"
            )

        if not normalized_path.is_file():
            raise ValueError(
                f"Đường dẫn không phải là tệp: {normalized_path}"
            )

        return normalized_path

    def open_workbook(
        self,
        file_path: str | Path,
        data_only: bool = True,
    ) -> Workbook:
        """Mở Workbook Excel ở chế độ an toàn."""
        normalized_path = self.validate_file(file_path)

        return load_workbook(
            filename=normalized_path,
            read_only=False,
            data_only=data_only,
            keep_vba=normalized_path.suffix.lower() == ".xlsm",
        )

    @staticmethod
    def get_worksheet(
        workbook: Workbook,
        sheet_name: str,
    ) -> Worksheet:
        """Kiểm tra và lấy Worksheet theo tên."""
        normalized_name = sheet_name.strip()

        if not normalized_name:
            raise ValueError(
                "Tên worksheet không được để trống."
            )

        if normalized_name not in workbook.sheetnames:
            raise ValueError(
                f"Không tìm thấy worksheet: {normalized_name}"
            )

        return workbook[normalized_name]

    @staticmethod
    def clean_text(value: Any) -> str:
        """Chuẩn hóa giá trị thành chuỗi."""
        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def to_integer(
        value: Any,
        default: int | None = None,
    ) -> int | None:
        """Chuyển giá trị thành số nguyên nếu có thể."""
        if value is None or value == "":
            return default

        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def to_string_list(
        value: Any,
    ) -> list[str]:
        """Chuyển dữ liệu thành danh sách chuỗi đã làm sạch."""
        if value is None:
            return []

        if isinstance(value, str):
            items = value.replace(
                "\r",
                "\n",
            ).split("\n")

            return [
                item.strip(" -•\t")
                for item in items
                if item.strip(" -•\t")
            ]

        if isinstance(value, (list, tuple, set)):
            return [
                str(item).strip()
                for item in value
                if str(item).strip()
            ]

        return [str(value).strip()]

    @staticmethod
    def is_empty(value: Any) -> bool:
        """Kiểm tra giá trị có được xem là rỗng hay không."""
        if value is None:
            return True

        if isinstance(value, str):
            return not value.strip()

        if isinstance(value, (list, tuple, set, dict)):
            return len(value) == 0

        return False

    @abstractmethod
    def build(
        self,
        *args: Any,
        **kwargs: Any,
    ) -> BaseModel:
        """Tạo model chuyên biệt từ dữ liệu đầu vào."""
        raise NotImplementedError