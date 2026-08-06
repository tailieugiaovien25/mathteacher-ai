import re
from typing import Any

from openpyxl import load_workbook

from models.lesson_model import LessonModel


class LessonModelBuilder:
    """Chuyển dữ liệu Excel thành LessonModel."""

    FIELD_ALIASES: dict[str, tuple[str, ...]] = {
        "subject": (
            "môn",
            "môn học",
            "subject",
        ),
        "grade": (
            "lớp",
            "khối",
            "grade",
        ),
        "lesson_name": (
            "tên bài",
            "bài học",
            "tên bài học",
            "lesson name",
        ),
        "lesson_number": (
            "bài",
            "số bài",
            "lesson number",
        ),
        "period_count": (
            "số tiết",
            "tiết",
            "period count",
        ),
        "learning_requirements": (
            "yêu cầu cần đạt",
            "yccd",
            "learning requirements",
        ),
        "objectives": (
            "mục tiêu",
            "objectives",
        ),
        "registered_equipment": (
            "thiết bị",
            "thiết bị dạy học",
            "registered equipment",
        ),
        "learning_resources": (
            "học liệu",
            "tài liệu",
            "learning resources",
        ),
    }

    def build_from_excel_row(
        self,
        file_path: str,
        sheet_name: str,
        header_row: int,
        data_row: int,
    ) -> LessonModel:
        """Đọc bảng Excel có hàng tiêu đề chuẩn."""
        workbook = load_workbook(
            filename=file_path,
            read_only=False,
            data_only=True,
            keep_vba=file_path.lower().endswith(".xlsm"),
        )

        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Không tìm thấy worksheet: {sheet_name}"
                )

            worksheet = workbook[sheet_name]

            row_data = self._read_row_as_mapping(
                worksheet=worksheet,
                header_row=header_row,
                data_row=data_row,
            )

            lesson = LessonModel(
                subject=self._get_text(
                    row_data,
                    "subject",
                ),
                grade=self._get_text(
                    row_data,
                    "grade",
                ),
                lesson_name=self._get_text(
                    row_data,
                    "lesson_name",
                ),
                lesson_number=self._get_text(
                    row_data,
                    "lesson_number",
                ),
                period_count=self._get_integer(
                    row_data,
                    "period_count",
                ),
                learning_requirements=self._get_list(
                    row_data,
                    "learning_requirements",
                ),
                objectives=self._get_list(
                    row_data,
                    "objectives",
                ),
                registered_equipment=self._get_list(
                    row_data,
                    "registered_equipment",
                ),
                learning_resources=self._get_list(
                    row_data,
                    "learning_resources",
                ),
                source_file=file_path,
                source_sheet=sheet_name,
                source_row=data_row,
                metadata={
                    "schema": "header_mapping",
                    "header_row": header_row,
                    "raw_row": row_data,
                },
            )

            self._validate_required_fields(lesson)

            return lesson
        finally:
            workbook.close()

    def build_from_luubg_row(
        self,
        file_path: str,
        sheet_name: str,
        data_row: int,
    ) -> LessonModel:
        """Đọc một hàng theo cấu trúc thực tế của worksheet LuuBG."""
        workbook = load_workbook(
            filename=file_path,
            read_only=False,
            data_only=True,
            keep_vba=file_path.lower().endswith(".xlsm"),
        )

        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Không tìm thấy worksheet: {sheet_name}"
                )

            worksheet = workbook[sheet_name]

            record_number = worksheet.cell(
                row=data_row,
                column=1,
            ).value

            category_code = worksheet.cell(
                row=data_row,
                column=2,
            ).value

            lesson_sequence = worksheet.cell(
                row=data_row,
                column=4,
            ).value

            plan_index = worksheet.cell(
                row=data_row,
                column=5,
            ).value

            grade_code = worksheet.cell(
                row=data_row,
                column=6,
            ).value

            subject_area = worksheet.cell(
                row=data_row,
                column=7,
            ).value

            raw_lesson_name = worksheet.cell(
                row=data_row,
                column=8,
            ).value

            duration_code = worksheet.cell(
                row=data_row,
                column=10,
            ).value

            lesson_name = self._clean_text(
                raw_lesson_name
            )

            period_number = self._extract_period_number(
                lesson_name
            )

            total_periods = self._extract_total_periods(
                duration_code
            )

            lesson = LessonModel(
                subject="Toán",
                grade=self._clean_text(
                    grade_code
                ),
                lesson_name=lesson_name,
                lesson_number=self._clean_text(
                    lesson_sequence
                ),
                period_count=total_periods,
                source_file=file_path,
                source_sheet=sheet_name,
                source_row=data_row,
                metadata={
                    "schema": "LuuBG",
                    "record_number": record_number,
                    "category_code": self._clean_text(
                        category_code
                    ),
                    "plan_index": plan_index,
                    "subject_area": self._clean_text(
                        subject_area
                    ),
                    "period_number": period_number,
                    "duration_code": self._clean_text(
                        duration_code
                    ),
                },
            )

            self._validate_luubg_fields(lesson)

            return lesson
        finally:
            workbook.close()

    def _read_row_as_mapping(
        self,
        worksheet,
        header_row: int,
        data_row: int,
    ) -> dict[str, Any]:
        """Ghép tên cột với giá trị của một hàng dữ liệu."""
        result: dict[str, Any] = {}

        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):
            header_value = worksheet.cell(
                row=header_row,
                column=column_index,
            ).value

            if header_value is None:
                continue

            normalized_header = self._normalize(
                header_value
            )

            if not normalized_header:
                continue

            result[normalized_header] = worksheet.cell(
                row=data_row,
                column=column_index,
            ).value

        return result

    def _find_value(
        self,
        row_data: dict[str, Any],
        field_name: str,
    ) -> Any:
        aliases = self.FIELD_ALIASES.get(
            field_name,
            (),
        )

        for alias in aliases:
            normalized_alias = self._normalize(alias)

            if normalized_alias in row_data:
                return row_data[normalized_alias]

        return None

    def _get_text(
        self,
        row_data: dict[str, Any],
        field_name: str,
    ) -> str:
        value = self._find_value(
            row_data,
            field_name,
        )

        if value is None:
            return ""

        return str(value).strip()

    def _get_integer(
        self,
        row_data: dict[str, Any],
        field_name: str,
    ) -> int | None:
        value = self._find_value(
            row_data,
            field_name,
        )

        if value is None or value == "":
            return None

        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _get_list(
        self,
        row_data: dict[str, Any],
        field_name: str,
    ) -> list[str]:
        value = self._find_value(
            row_data,
            field_name,
        )

        if value is None:
            return []

        if isinstance(value, str):
            items = value.replace(
                "\r",
                "\n",
            ).split("\n")

            return [
                item.strip(" -•\t")
                for item in items
                if item.strip(" -•\t")
            ]

        return [str(value).strip()]

    @staticmethod
    def _normalize(value: Any) -> str:
        return " ".join(
            str(value).strip().lower().split()
        )

    @staticmethod
    def _clean_text(value: Any) -> str:
        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def _extract_period_number(
        lesson_name: str,
    ) -> int | None:
        match = re.search(
            r"\(\s*tiết\s+(\d+)\s*\)",
            lesson_name,
            flags=re.IGNORECASE,
        )

        if match is None:
            return None

        return int(match.group(1))

    @staticmethod
    def _extract_total_periods(
        duration_code: Any,
    ) -> int | None:
        if duration_code is None:
            return None

        match = re.search(
            r"(\d+)",
            str(duration_code),
        )

        if match is None:
            return None

        return int(match.group(1))

    @staticmethod
    def _validate_luubg_fields(
        lesson: LessonModel,
    ) -> None:
        if not lesson.lesson_name:
            lesson.warnings.append(
                "Hàng LuuBG chưa có tên bài học."
            )

        if not lesson.grade:
            lesson.warnings.append(
                "Hàng LuuBG chưa có mã lớp hoặc khối."
            )

        if lesson.period_count is None:
            lesson.warnings.append(
                "Chưa xác định được tổng số tiết từ cột J."
            )

        if lesson.metadata.get(
            "period_number"
        ) is None:
            lesson.warnings.append(
                "Chưa xác định được số tiết thành phần trong tên bài."
            )

        if not lesson.metadata.get(
            "subject_area"
        ):
            lesson.warnings.append(
                "Chưa xác định được phân môn Đại hoặc Hình."
            )

    @staticmethod
    def _validate_required_fields(
        lesson: LessonModel,
    ) -> None:
        if not lesson.lesson_name:
            lesson.warnings.append(
                "Chưa xác định được tên bài học."
            )

        if not lesson.grade:
            lesson.warnings.append(
                "Chưa xác định được lớp hoặc khối."
            )

        if lesson.period_count is None:
            lesson.warnings.append(
                "Chưa xác định được số tiết."
            )

        if not lesson.learning_requirements:
            lesson.warnings.append(
                "Chưa xác định được yêu cầu cần đạt."
            )