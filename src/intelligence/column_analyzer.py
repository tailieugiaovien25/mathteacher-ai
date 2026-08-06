from collections import Counter
from datetime import date, datetime
from typing import Any

from openpyxl.utils import get_column_letter

from models.column_info import ColumnInfo
from models.header_info import HeaderInfo
from models.used_range import UsedRange


class ColumnAnalyzer:
    """Phân tích tên và kiểu dữ liệu của từng cột."""

    SAMPLE_SIZE = 100

    @staticmethod
    def _detect_value_type(value: Any) -> str:
        if value is None or value == "":
            return "empty"

        if isinstance(value, bool):
            return "boolean"

        if isinstance(value, (datetime, date)):
            return "date"

        if isinstance(value, int):
            return "integer"

        if isinstance(value, float):
            return "number"

        if isinstance(value, str):
            return "text"

        return "other"

    def analyze(
        self,
        worksheet,
        used_range: UsedRange,
        header_info: HeaderInfo,
    ) -> list[ColumnInfo]:
        columns: list[ColumnInfo] = []

        if header_info.row_index == 0:
            return columns

        first_data_row = header_info.row_index + 1
        last_sample_row = min(
            used_range.last_row,
            first_data_row + self.SAMPLE_SIZE - 1,
        )

        for column_number in range(
            used_range.first_column,
            used_range.last_column + 1,
        ):
            type_counter: Counter[str] = Counter()

            for row_number in range(
                first_data_row,
                last_sample_row + 1,
            ):
                value = worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value

                value_type = self._detect_value_type(value)

                if value_type != "empty":
                    type_counter[value_type] += 1

            data_type = (
                type_counter.most_common(1)[0][0]
                if type_counter
                else "empty"
            )

            header_offset = (
                column_number - used_range.first_column
            )

            header = (
                header_info.headers[header_offset]
                if header_offset < len(header_info.headers)
                else f"Column_{get_column_letter(column_number)}"
            )

            columns.append(
                ColumnInfo(
                    column_letter=get_column_letter(
                        column_number
                    ),
                    column_index=column_number,
                    header=header,
                    data_type=data_type,
                )
            )

        return columns