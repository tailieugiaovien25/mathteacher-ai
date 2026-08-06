from typing import Any

from openpyxl.utils import get_column_letter

from models.header_info import HeaderInfo
from models.used_range import UsedRange


class HeaderDetector:
    """Tự động tìm hàng có khả năng là hàng tiêu đề."""

    MAX_SCAN_ROWS = 50

    @staticmethod
    def _normalize(value: Any) -> str:
        if value is None:
            return ""

        if isinstance(value, str):
            return value.strip()

        return str(value).strip()

    def detect(
        self,
        worksheet,
        used_range: UsedRange,
    ) -> HeaderInfo:
        if used_range.last_row == 0:
            return HeaderInfo(
                row_index=0,
                headers=[],
            )

        scan_last_row = min(
            used_range.last_row,
            used_range.first_row + self.MAX_SCAN_ROWS - 1,
        )

        best_row = 0
        best_score = -1
        best_headers: list[str] = []

        for row_number in range(
            used_range.first_row,
            scan_last_row + 1,
        ):
            headers: list[str] = []
            non_empty_count = 0
            text_count = 0
            unique_values: set[str] = set()

            for column_number in range(
                used_range.first_column,
                used_range.last_column + 1,
            ):
                value = worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value

                normalized = self._normalize(value)
                headers.append(normalized)

                if normalized:
                    non_empty_count += 1
                    unique_values.add(normalized.lower())

                    if isinstance(value, str):
                        text_count += 1

            score = (
                non_empty_count * 3
                + text_count * 2
                + len(unique_values)
            )

            if non_empty_count >= 2 and score > best_score:
                best_score = score
                best_row = row_number
                best_headers = headers

        if best_row == 0:
            return HeaderInfo(
                row_index=0,
                headers=[],
            )

        final_headers: list[str] = []

        for offset, header in enumerate(best_headers):
            column_number = used_range.first_column + offset
            column_letter = get_column_letter(column_number)

            final_headers.append(
                header if header else f"Column_{column_letter}"
            )

        return HeaderInfo(
            row_index=best_row,
            headers=final_headers,
        )