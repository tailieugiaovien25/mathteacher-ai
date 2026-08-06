from dataclasses import asdict
from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import load_workbook

from excel_engine.range_detector import RangeDetector
from intelligence.column_analyzer import ColumnAnalyzer
from intelligence.header_detector import HeaderDetector
from intelligence.table_detector import TableDetector


class ExcelIntelligenceEngine:
    """Điều phối quá trình phân tích cấu trúc Worksheet."""

    def __init__(self) -> None:
        self.range_detector = RangeDetector()
        self.header_detector = HeaderDetector()
        self.column_analyzer = ColumnAnalyzer()
        self.table_detector = TableDetector()

    def analyze(
        self,
        file_path: str,
        sheet_name: str,
    ) -> dict[str, Any]:
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(
                f"Không tìm thấy file: {file_path}"
            )

        total_start = perf_counter()

        print("  - Đang mở Workbook...", flush=True)
        step_start = perf_counter()

        workbook = load_workbook(
            filename=path,
            read_only=False,
            data_only=True,
            keep_vba=path.suffix.lower() == ".xlsm",
        )

        print(
            f"    Hoàn thành: "
            f"{perf_counter() - step_start:.2f} giây",
            flush=True,
        )

        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Không tìm thấy worksheet: {sheet_name}"
                )

            worksheet = workbook[sheet_name]

            print(
                "  - Đang xác định vùng dữ liệu...",
                flush=True,
            )
            step_start = perf_counter()

            used_range = self.range_detector.detect(
                worksheet
            )

            range_time = perf_counter() - step_start
            print(
                f"    Hoàn thành: {range_time:.2f} giây",
                flush=True,
            )

            print(
                "  - Đang nhận diện hàng tiêu đề...",
                flush=True,
            )
            step_start = perf_counter()

            header_info = self.header_detector.detect(
                worksheet,
                used_range,
            )

            header_time = perf_counter() - step_start
            print(
                f"    Hoàn thành: {header_time:.2f} giây",
                flush=True,
            )

            print(
                "  - Đang phân tích các cột...",
                flush=True,
            )
            step_start = perf_counter()

            columns = self.column_analyzer.analyze(
                worksheet,
                used_range,
                header_info,
            )

            column_time = perf_counter() - step_start
            print(
                f"    Hoàn thành: {column_time:.2f} giây",
                flush=True,
            )

            print(
                "  - Đang phát hiện các bảng dữ liệu...",
                flush=True,
            )
            step_start = perf_counter()

            tables = self.table_detector.detect(
                worksheet,
                used_range,
                header_info,
            )

            table_time = perf_counter() - step_start
            print(
                f"    Hoàn thành: {table_time:.2f} giây",
                flush=True,
            )

            total_time = perf_counter() - total_start

            return {
                "file_name": path.name,
                "sheet_name": sheet_name,
                "used_range": asdict(used_range),
                "header": asdict(header_info),
                "columns": [
                    asdict(column)
                    for column in columns
                ],
                "tables": [
                    asdict(table)
                    for table in tables
                ],
                "performance": {
                    "range_detector_seconds": round(
                        range_time,
                        4,
                    ),
                    "header_detector_seconds": round(
                        header_time,
                        4,
                    ),
                    "column_analyzer_seconds": round(
                        column_time,
                        4,
                    ),
                    "table_detector_seconds": round(
                        table_time,
                        4,
                    ),
                    "total_seconds": round(
                        total_time,
                        4,
                    ),
                },
            }
        finally:
            workbook.close()