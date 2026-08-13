import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    if value is None:
        return None

    match = re.search(
        r"([6-9])",
        str(value),
    )

    if not match:
        return None

    return int(match.group(1))


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        index: dict[
            tuple[int, int],
            list[dict[str, object]],
        ] = defaultdict(list)

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if (
                subject_grade is None
                or period is None
                or lesson_name is None
            ):
                continue

            grade = detect_grade(
                subject_grade
            )

            if grade is None:
                continue

            try:
                period_number = int(
                    period
                )
            except (TypeError, ValueError):
                continue

            key = (
                grade,
                period_number,
            )

            index[key].append(
                {
                    "row": row_number,
                    "subject_grade": (
                        str(subject_grade)
                    ),
                    "period": period_number,
                    "lesson_name": (
                        str(lesson_name)
                    ),
                }
            )

        duplicates = {
            key: rows
            for key, rows in index.items()
            if len(rows) > 1
        }

        print("=" * 72)
        print(
            "LP-03C.5B - "
            "GRADE + PERIOD UNIQUENESS TEST"
        )
        print("=" * 72)

        print(
            "Tổng khóa KHOI + TIET: "
            f"{len(index)}"
        )

        print(
            "Số khóa bị trùng: "
            f"{len(duplicates)}"
        )

        if duplicates:
            print(
                "\nCÁC KHÓA BỊ TRÙNG"
            )

            for (
                grade,
                period_number,
            ), rows in sorted(
                duplicates.items()
            ):
                print(
                    f"\n- Khối {grade} | "
                    f"Tiết {period_number}"
                )

                for item in rows:
                    print(
                        f"    Row {item['row']} | "
                        f"{item['subject_grade']} | "
                        f"{item['lesson_name']}"
                    )
        else:
            print(
                "\nKhông phát hiện khóa trùng."
            )

        print(
            "\nKẾT QUẢ: "
            "UNIQUENESS INSPECTION COMPLETE"
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()