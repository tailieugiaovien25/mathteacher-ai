import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, "src")

from repositories.yccd_repository_v2 import (
    YCCDRepositoryV2,
)


TEST_FILE = Path(
    "output/test_yccd_repository_v2_provenance.xlsx"
)


HEADERS = [
    "YCCD_ID",
    "LESSON_KEY",
    "MON",
    "KHOI",
    "BAI_ID",
    "TEN_BAI",
    "YCCD_ORDER",
    "YEU_CAU_CAN_DAT",
    "LOAI_YCCD",
    "YCCD_GOC_ID",
    "NGUON",
    "THAM_CHIEU",
    "PHIEN_BAN",
    "TRANG_THAI",
    "NGAY_CAP_NHAT",
    "GHI_CHU",
]


def create_test_workbook() -> None:
    TEST_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "YCCD"

    for column_index, header in enumerate(
        HEADERS,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

    rows = [
        [
            "T7_DAI_B03_Y00",
            "T7_DAI_B03",
            "Toán",
            7,
            "B03",
            "Bài 3. Lũy thừa",
            1,
            "YCCĐ chính thức của chương trình.",
            "CHINH_THUC",
            None,
            "CTGDPT_2018",
            "Yêu cầu cần đạt tương ứng",
            "1.0",
            "approved",
            date(2026, 8, 8),
            "",
        ],
        [
            "T7_DAI_B03_Y01",
            "T7_DAI_B03",
            "Toán",
            7,
            "B03",
            "Bài 3. Lũy thừa",
            2,
            "YCCĐ triển khai số 1.",
            "TRIEN_KHAI",
            "T7_DAI_B03_Y00",
            "TONG_HOP",
            "CTGDPT_2018 + SGK Kết nối tri thức",
            "1.0",
            "approved",
            date(2026, 8, 8),
            "",
        ],
        [
            "T7_DAI_B03_Y02",
            "T7_DAI_B03",
            "Toán",
            7,
            "B03",
            "Bài 3. Lũy thừa",
            3,
            "YCCĐ triển khai số 2.",
            "TRIEN_KHAI",
            "T7_DAI_B03_Y00",
            "TONG_HOP",
            "CTGDPT_2018 + SGK Kết nối tri thức",
            "1.0",
            "draft",
            date(2026, 8, 8),
            "",
        ],
    ]

    for row_index, row_values in enumerate(
        rows,
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    workbook.save(TEST_FILE)
    workbook.close()


def main() -> None:
    create_test_workbook()

    repository = YCCDRepositoryV2()

    try:
        rows = repository.load_rows(
            TEST_FILE
        )

        assert len(rows) == 3

        approved = (
            repository.find_by_lesson_key(
                file_path=TEST_FILE,
                lesson_key="T7_DAI_B03",
            )
        )

        assert len(approved) == 2

        official = approved[0]
        derived = approved[1]

        assert (
            official.yccd_type
            == "CHINH_THUC"
        )

        assert (
            official.source_yccd_id
            is None
        )

        assert (
            derived.yccd_type
            == "TRIEN_KHAI"
        )

        assert (
            derived.source_yccd_id
            == "T7_DAI_B03_Y00"
        )

        assert (
            derived.source
            == "TONG_HOP"
        )

        assert (
            derived.reference
            == "CTGDPT_2018 + SGK Kết nối tri thức"
        )

        draft = (
            repository.find_by_lesson_key(
                file_path=TEST_FILE,
                lesson_key="T7_DAI_B03",
                status="draft",
            )
        )

        assert len(draft) == 1

        assert (
            draft[0].yccd_id
            == "T7_DAI_B03_Y02"
        )

        print("=" * 70)
        print(
            "LP-03D.3D.3B - "
            "YCCD REPOSITORY V2 PROVENANCE TEST"
        )
        print("=" * 70)

        print("- Đọc schema 16 cột: PASS")
        print("- Đọc CHINH_THUC: PASS")
        print("- CHINH_THUC không có YCCD_GOC_ID: PASS")
        print("- Đọc TRIEN_KHAI: PASS")
        print("- TRIEN_KHAI truy vết YCCD_GOC_ID: PASS")
        print("- Đọc NGUON: PASS")
        print("- Đọc THAM_CHIEU: PASS")
        print("- Lọc approved: PASS")
        print("- Lọc draft: PASS")

        print(
            "\nKẾT QUẢ: 9/9 TEST PASS"
        )

    finally:
        if TEST_FILE.exists():
            TEST_FILE.unlink()


if __name__ == "__main__":
    main()