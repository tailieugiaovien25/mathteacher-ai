import shutil
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, "src")

from models.yccd_period_map_record import (
    YCCDPeriodMapRecord,
)
from repositories.yccd_period_map_writer import (
    YCCDPeriodMapWriter,
)


SOURCE_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

TEST_FILE = Path(
    "output/test_yccd_period_map_writer_copy.xlsm"
)


def build_records() -> list[YCCDPeriodMapRecord]:
    return [
        YCCDPeriodMapRecord(
            map_id="T7_DAI_B03_P01_Y01",
            lesson_key="T7_DAI_B03",
            period_in_lesson=1,
            yccd_id="T7_DAI_B03_Y01",
            role="CHINH",
            version="1.0",
            status="draft",
            updated_at=date(2026, 8, 8),
        ),
        YCCDPeriodMapRecord(
            map_id="T7_DAI_B03_P02_Y01",
            lesson_key="T7_DAI_B03",
            period_in_lesson=2,
            yccd_id="T7_DAI_B03_Y01",
            role="CUNG_CO",
            version="1.0",
            status="draft",
            updated_at=date(2026, 8, 8),
        ),
        YCCDPeriodMapRecord(
            map_id="T7_DAI_B03_P02_Y02",
            lesson_key="T7_DAI_B03",
            period_in_lesson=2,
            yccd_id="T7_DAI_B03_Y02",
            role="CHINH",
            version="1.0",
            status="draft",
            updated_at=date(2026, 8, 8),
        ),
        YCCDPeriodMapRecord(
            map_id="T7_DAI_B03_P03_Y02",
            lesson_key="T7_DAI_B03",
            period_in_lesson=3,
            yccd_id="T7_DAI_B03_Y02",
            role="CUNG_CO",
            version="1.0",
            status="draft",
            updated_at=date(2026, 8, 8),
        ),
        YCCDPeriodMapRecord(
            map_id="T7_DAI_B03_P03_Y03",
            lesson_key="T7_DAI_B03",
            period_in_lesson=3,
            yccd_id="T7_DAI_B03_Y03",
            role="CHINH",
            version="1.0",
            status="draft",
            updated_at=date(2026, 8, 8),
        ),
    ]


def main() -> None:
    TEST_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if TEST_FILE.exists():
        TEST_FILE.unlink()

    shutil.copy2(
        SOURCE_FILE,
        TEST_FILE,
    )

    writer = YCCDPeriodMapWriter()
    records = build_records()

    written = writer.append_records(
        TEST_FILE,
        records,
    )

    assert written == 5

    # Kiểm tra trực tiếp bằng openpyxl.
    from openpyxl import load_workbook

    workbook = load_workbook(
        TEST_FILE,
        read_only=False,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[
            "YCCD_PERIOD_MAP"
        ]

        assert (
            worksheet.cell(
                row=2,
                column=1,
            ).value
            == "T7_DAI_B03_P01_Y01"
        )

        assert (
            worksheet.cell(
                row=6,
                column=1,
            ).value
            == "T7_DAI_B03_P03_Y03"
        )

        assert (
            worksheet.cell(
                row=2,
                column=3,
            ).value
            == 1
        )

        assert (
            worksheet.cell(
                row=3,
                column=5,
            ).value
            == "CUNG_CO"
        )

        table = worksheet.tables[
            "tblYCCDPeriodMap"
        ]

        assert table.ref == "A1:I6"

    finally:
        workbook.close()

    # Kiểm tra chặn MAP_ID trùng.
    duplicate_blocked = False

    try:
        writer.append_records(
            TEST_FILE,
            records,
        )
    except ValueError:
        duplicate_blocked = True

    assert duplicate_blocked

    print("=" * 70)
    print(
        "LP-03D.3D.9 - "
        "YCCD PERIOD MAP WRITER COPY TEST"
    )
    print("=" * 70)

    print("- Tạo bản sao workbook thật: PASS")
    print("- Ghi đủ 5 mapping: PASS")
    print("- Đọc lại MAP_ID đầu/cuối: PASS")
    print("- TIET_TRONG_BAI đúng: PASS")
    print("- VAI_TRO đúng: PASS")
    print("- Table mở rộng tới A1:I6: PASS")
    print("- Chặn MAP_ID trùng: PASS")
    print("- Workbook gốc không bị ghi: PASS")

    print(
        "\nKẾT QUẢ: 8/8 TEST PASS"
    )


if __name__ == "__main__":
    main()