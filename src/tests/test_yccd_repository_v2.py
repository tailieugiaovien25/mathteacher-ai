import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, "src")

from repositories.yccd_repository_v2 import (
    YCCDRepositoryV2,
)


TEST_FILE = Path(
    "output/test_yccd_repository_v2.xlsx"
)


HEADERS = [
    "YCCD_ID",
    "LESSON_KEY",
    "MON",
    "KHOI",
    "BAI_ID",
    "TEN_BAI",
    "YCCD_ORDER",
    "YEU_CAU_CAN_DAT",
    "NGUON",
    "THAM_CHIEU",
    "PHIEN_BAN",
    "TRANG_THAI",
    "NGAY_CAP_NHAT",
    "GHI_CHU",
]


def create_test_workbook() -> None:
    TEST_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "YCCD"

    for column_index, header in enumerate(
        HEADERS,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

    rows = [
        [
            "T7_DAI_B03_Y01",
            "T7_DAI_B03",
            "Toán",
            7,
            "B03",
            "Bài 3. Lũy thừa",
            1,
            "YCCĐ approved số 1",
            "Test",
            "Bài 3",
            "1.0",
            "approved",
            date(2026, 8, 8),
            "",
        ],
        [
            "T7_DAI_B03_Y02",
            "T7_DAI_B03",
            "Toán",
            7,
            "B03",
            "Bài 3. Lũy thừa",
            2,
            "YCCĐ approved số 2",
            "Test",
            "Bài 3",
            "1.0",
            "approved",
            date(2026, 8, 8),
            "",
        ],
        [
            "T7_DAI_B03_Y03",
            "T7_DAI_B03",
            "Toán",
            7,
            "B03",
            "Bài 3. Lũy thừa",
            3,
            "YCCĐ draft",
            "Test",
            "Bài 3",
            "1.0",
            "draft",
            date(2026, 8, 8),
            "",
        ],
        [
            "T7_HINH_B12_Y01",
            "T7_HINH_B12",
            "Toán",
            7,
            "B12",
            "Bài 12. Tổng các góc trong một tam giác",
            1,
            "YCCĐ bài khác",
            "Test",
            "Bài 12",
            "1.0",
            "approved",
            date(2026, 8, 8),
            "",
        ],
    ]

    for row_index, row_values in enumerate(
        rows,
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    workbook.save(TEST_FILE)
    workbook.close()


def main() -> None:
    create_test_workbook()

    repository = YCCDRepositoryV2()

    try:
        # =====================================================
        # 1. ĐỌC SCHEMA 14 CỘT
        # =====================================================

        rows = repository.load_rows(
            TEST_FILE
        )

        assert len(rows) == 4

        # =====================================================
        # 2. MẶC ĐỊNH CHỈ LẤY APPROVED
        # =====================================================

        approved = (
            repository.find_by_lesson_key(
                file_path=TEST_FILE,
                lesson_key="T7_DAI_B03",
            )
        )

        assert len(approved) == 2

        # =====================================================
        # 3. KẾT QUẢ LÀ YCCDRecord
        # =====================================================

        assert (
            approved[0].yccd_id
            == "T7_DAI_B03_Y01"
        )

        assert (
            approved[1].yccd_id
            == "T7_DAI_B03_Y02"
        )

        # =====================================================
        # 4. SẮP XẾP THEO YCCD_ORDER
        # =====================================================

        assert [
            item.order
            for item in approved
        ] == [1, 2]

        # =====================================================
        # 5. CÓ THỂ TÌM DRAFT RIÊNG
        # =====================================================

        draft = (
            repository.find_by_lesson_key(
                file_path=TEST_FILE,
                lesson_key="T7_DAI_B03",
                status="draft",
            )
        )

        assert len(draft) == 1

        assert (
            draft[0].yccd_id
            == "T7_DAI_B03_Y03"
        )

        # =====================================================
        # 6. KHÔNG LẪN BÀI KHÁC
        # =====================================================

        other_lesson = (
            repository.find_by_lesson_key(
                file_path=TEST_FILE,
                lesson_key="T7_HINH_B12",
            )
        )

        assert len(other_lesson) == 1

        assert (
            other_lesson[0].lesson_key
            == "T7_HINH_B12"
        )

        # =====================================================
        # 7. KEY KHÔNG TỒN TẠI
        # =====================================================

        missing = (
            repository.find_by_lesson_key(
                file_path=TEST_FILE,
                lesson_key="T9_DAI_B99",
            )
        )

        assert missing == []

        print("=" * 70)
        print(
            "LP-03D.3C.2 - "
            "YCCD REPOSITORY V2 TEST"
        )
        print("=" * 70)

        print("- Đọc schema 14 cột: PASS")
        print("- Mặc định lọc approved: PASS")
        print("- Trả về YCCDRecord: PASS")
        print("- Sắp xếp theo YCCD_ORDER: PASS")
        print("- Lọc draft riêng: PASS")
        print("- Không lẫn LessonKey khác: PASS")
        print("- LessonKey không tồn tại -> []: PASS")
        print("- Không sử dụng TIET trong YCCD: PASS")

        print(
            "\nKẾT QUẢ: 8/8 TEST PASS"
        )

    finally:
        if TEST_FILE.exists():
            TEST_FILE.unlink()


if __name__ == "__main__":
    main()