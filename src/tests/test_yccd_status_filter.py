import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, "src")

from repositories.yccd_repository import YCCDRepository


TEST_FILE = Path(
    "output/test_yccd_status_filter.xlsx"
)


HEADERS = [
    "YCCD_ID",
    "MON",
    "KHOI",
    "BAI_ID",
    "TEN_BAI",
    "TIET",
    "YCCD_ORDER",
    "YEU_CAU_CAN_DAT",
    "NGUON",
    "PHIEN_BAN",
    "TRANG_THAI",
    "NGAY_CAP_NHAT",
    "GHI_CHU",
]


def create_test_workbook() -> None:
    TEST_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "YCCD"

    for column_index, header in enumerate(
        HEADERS,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

    rows = [
        [
            "TEST_DRAFT",
            "Toán",
            8,
            "B02",
            "Bài 2. Đa thức",
            1,
            1,
            "YCCĐ trạng thái draft",
            "Test",
            "1.0",
            "draft",
            None,
            "",
        ],
        [
            "TEST_APPROVED",
            "Toán",
            8,
            "Bài 2",
            "Bài 2. Đa thức",
            1,
            2,
            "YCCĐ trạng thái approved",
            "Test",
            "1.0",
            "approved",
            None,
            "",
        ],
        [
            "TEST_DEPRECATED",
            "Toán",
            8,
            "B02",
            "Bài 2. Đa thức",
            1,
            3,
            "YCCĐ trạng thái deprecated",
            "Test",
            "1.0",
            "deprecated",
            None,
            "",
        ],
    ]

    for row_index, row_values in enumerate(
        rows,
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    workbook.save(TEST_FILE)
    workbook.close()


def main() -> None:
    create_test_workbook()

    repository = YCCDRepository()

    try:
        all_rows = repository.load_rows(
            TEST_FILE
        )

        assert len(all_rows) == 3

        approved_rows = repository.find(
            file_path=TEST_FILE,
            subject="Toán",
            grade=8,
            lesson_name="Bài 2. Đa thức",
            period=1,
        )

        assert len(approved_rows) == 1
        assert (
            approved_rows[0]["YCCD_ID"]
            == "TEST_APPROVED"
        )

        draft_rows = repository.find(
            file_path=TEST_FILE,
            subject="Toán",
            grade=8,
            lesson_name="Bài 2. Đa thức",
            period=1,
            status="draft",
        )

        assert len(draft_rows) == 1
        assert (
            draft_rows[0]["YCCD_ID"]
            == "TEST_DRAFT"
        )

        deprecated_rows = repository.find(
            file_path=TEST_FILE,
            subject="Toán",
            grade=8,
            lesson_name="Bài 2. Đa thức",
            period=1,
            status="deprecated",
        )

        assert len(deprecated_rows) == 1
        assert (
            deprecated_rows[0]["YCCD_ID"]
            == "TEST_DEPRECATED"
        )

        print("=" * 70)
        print("LP-03B.4 - YCCD STATUS FILTER TEST")
        print("=" * 70)

        print("- Đọc 3 trạng thái: PASS")
        print("- draft filter: PASS")
        print("- approved filter: PASS")
        print("- deprecated filter: PASS")
        print("- approved mặc định: PASS")
        print("- File test độc lập với workbook thật: PASS")

        print("\nKẾT QUẢ: 6/6 TEST PASS")

    finally:
        if TEST_FILE.exists():
            TEST_FILE.unlink()


if __name__ == "__main__":
    main()