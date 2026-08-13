import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "src")

from utils.lesson_key import build_lesson_key


EXCEL_FILE = Path(
    "data/input/LBG-TUYEN_chuan_VBA_macro.xlsm"
)

SHEET_NAME = "PPCT"


def detect_grade(value) -> int | None:
    """Lấy khối 6-9 từ giá trị Môn-lớp."""

    if value is None:
        return None

    match = re.search(
        r"([6-9])",
        str(value),
    )

    if not match:
        return None

    return int(match.group(1))


def is_fallback_key(
    lesson_key: str | None,
) -> bool:
    """Nhận diện LessonKey fallback dạng ..._Pxxx."""

    if not lesson_key:
        return False

    return bool(
        re.search(
            r"_P\d{3}$",
            lesson_key,
        )
    )


def main() -> None:
    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    try:
        worksheet = workbook[SHEET_NAME]

        records = []

        # -----------------------------------------------------
        # BƯỚC 1: ĐỌC PPCT VÀ TẠO LESSONKEY
        # -----------------------------------------------------

        for row_number in range(
            4,
            worksheet.max_row + 1,
        ):
            subject_grade = worksheet.cell(
                row=row_number,
                column=2,
            ).value

            period = worksheet.cell(
                row=row_number,
                column=3,
            ).value

            lesson_name = worksheet.cell(
                row=row_number,
                column=4,
            ).value

            if (
                subject_grade is None
                or period is None
                or lesson_name is None
            ):
                continue

            grade = detect_grade(
                subject_grade
            )

            if grade is None:
                continue

            lesson_key = build_lesson_key(
                grade=grade,
                subject_grade=str(
                    subject_grade
                ),
                lesson_name=str(
                    lesson_name
                ),
                period=period,
            )

            if lesson_key is None:
                continue

            records.append(
                {
                    "row": row_number,
                    "subject_grade": str(
                        subject_grade
                    ),
                    "period": period,
                    "lesson_name": str(
                        lesson_name
                    ),
                    "lesson_key": lesson_key,
                    "period_in_lesson": None,
                }
            )

        # -----------------------------------------------------
        # BƯỚC 2: NHÓM THEO LESSONKEY
        # -----------------------------------------------------

        groups = defaultdict(list)

        for item in records:
            groups[
                item["lesson_key"]
            ].append(item)

        # -----------------------------------------------------
        # BƯỚC 3: TÍNH TIET_TRONG_BAI
        # -----------------------------------------------------

        for lesson_key, items in groups.items():

            # Fallback là một nội dung gắn với một tiết riêng.
            if is_fallback_key(
                lesson_key
            ):
                for item in items:
                    item[
                        "period_in_lesson"
                    ] = 1

                continue

            # Bài có số:
            # sắp xếp theo tiết PPCT rồi đánh số 1, 2, 3...
            items.sort(
                key=lambda x: (
                    int(x["period"])
                    if str(
                        x["period"]
                    ).isdigit()
                    else 999999
                )
            )

            for index, item in enumerate(
                items,
                start=1,
            ):
                item[
                    "period_in_lesson"
                ] = index

        # -----------------------------------------------------
        # BƯỚC 4: KIỂM TRA
        # -----------------------------------------------------

        invalid = []

        for lesson_key, items in groups.items():

            values = [
                item[
                    "period_in_lesson"
                ]
                for item in items
            ]

            if is_fallback_key(
                lesson_key
            ):
                if any(
                    value != 1
                    for value in values
                ):
                    invalid.append(
                        lesson_key
                    )

                continue

            expected = list(
                range(
                    1,
                    len(items) + 1,
                )
            )

            if sorted(values) != expected:
                invalid.append(
                    lesson_key
                )

        # -----------------------------------------------------
        # BÁO CÁO
        # -----------------------------------------------------

        multi_period = {
            key: items
            for key, items in groups.items()
            if (
                not is_fallback_key(key)
                and len(items) > 1
            )
        }

        fallback_groups = {
            key: items
            for key, items in groups.items()
            if is_fallback_key(key)
        }

        print("=" * 72)
        print(
            "LP-03D.2C - "
            "PERIOD IN LESSON INSPECTION"
        )
        print("=" * 72)

        print(
            f"Tổng dòng PPCT: "
            f"{len(records)}"
        )

        print(
            f"Tổng LessonKey: "
            f"{len(groups)}"
        )

        print(
            f"Bài có nhiều tiết: "
            f"{len(multi_period)}"
        )

        print(
            f"Fallback theo tiết: "
            f"{len(fallback_groups)}"
        )

        print(
            f"LessonKey tính sai "
            f"TIET_TRONG_BAI: "
            f"{len(invalid)}"
        )

        print(
            "\n10 BÀI NHIỀU TIẾT ĐẦU TIÊN"
        )

        for lesson_key, items in list(
            multi_period.items()
        )[:10]:

            print(
                f"\n- {lesson_key}"
            )

            items_sorted = sorted(
                items,
                key=lambda x: (
                    x["period_in_lesson"]
                ),
            )

            for item in items_sorted:
                print(
                    f"    Tiết PPCT "
                    f"{item['period']} "
                    f"-> Tiết trong bài "
                    f"{item['period_in_lesson']} "
                    f"| "
                    f"{item['lesson_name']}"
                )

        if invalid:
            print(
                "\nCÁC LESSONKEY CÓ LỖI"
            )

            for lesson_key in invalid:
                print(
                    f"- {lesson_key}"
                )
        else:
            print(
                "\nKhông phát hiện lỗi "
                "TIET_TRONG_BAI."
            )

        accepted = (
            len(records) == 571
            and len(invalid) == 0
        )

        print(
            "\nTIÊU CHUẨN CHẤP NHẬN"
        )

        print(
            "- 571 dòng PPCT: "
            + (
                "PASS"
                if len(records) == 571
                else "FAIL"
            )
        )

        print(
            "- TIET_TRONG_BAI hợp lệ: "
            + (
                "PASS"
                if len(invalid) == 0
                else "FAIL"
            )
        )

        print(
            "\nKẾT QUẢ: "
            + (
                "PERIOD MAPPING ACCEPTED"
                if accepted
                else "PERIOD MAPPING NOT ACCEPTED"
            )
        )

    finally:
        workbook.close()


if __name__ == "__main__":
    main()