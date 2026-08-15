from pathlib import Path

import pytest

from scripts.weekly_schedule.app import (
    academic_year_options,
    build_mapping_schema,
    build_weekly_schedule,
    export_weekly_schedule,
    load_uploaded_workbook,
    inspect_uploaded_workbook,
    load_saved_schedule,
    save_weekly_schedule,
    saved_schedule_options,
    supabase_settings,
    authenticate_supabase,
    comma_separated_values,
    save_teacher_profile,
    schedule_rows,
    source_table_rows,
    teacher_options,
    week_options,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = (
    PROJECT_ROOT
    / "templates"
    / "weekly_schedule"
    / "mau_du_lieu_lich_bao_giang_v2.xlsx"
)


def _data():
    return load_uploaded_workbook(TEMPLATE.read_bytes(), TEMPLATE.name)


def test_upload_reads_template_and_returns_selection_options():
    data = _data()

    assert academic_year_options(data) == ("2026-2027",)
    assert week_options(data, "2026-2027") == (5, 6)
    assert teacher_options(data) == ("GV001",)


def test_app_builds_week_five_schedule_from_uploaded_template():
    schedule = build_weekly_schedule(
        data=_data(),
        teacher_id="GV001",
        academic_year="2026-2027",
        week_number=5,
    )

    rows = schedule_rows(schedule)
    assert len(rows) == 3
    assert [row["Tiết PPCT"] for row in rows] == [3, 2, 4]
    assert [row["Thứ"] for row in rows] == ["Thứ 2", "Thứ 3", "Thứ 4"]

    exported = export_weekly_schedule(schedule)
    assert exported.file_name.endswith("tuan-05.xlsx")
    assert exported.content.startswith(b"PK")


def test_source_tables_are_available_to_the_user_interface():
    tables = source_table_rows(_data())

    assert tuple(tables) == ("Tuần học", "Thời khóa biểu", "PPCT", "Tiết đã dạy")
    assert len(tables["Tuần học"]) == 2
    assert len(tables["Thời khóa biểu"]) == 3
    assert len(tables["PPCT"]) == 6
    assert len(tables["Tiết đã dạy"]) == 3


@pytest.mark.parametrize("name", ("data.xls", "data.xlsm", "data.csv"))
def test_upload_rejects_non_xlsx_files(name):
    with pytest.raises(ValueError, match=".xlsx"):
        load_uploaded_workbook(b"content", name)


def test_upload_rejects_empty_file():
    with pytest.raises(ValueError, match="rỗng"):
        load_uploaded_workbook(b"", "data.xlsx")


def test_ui_module_keeps_excel_details_in_the_adapter():
    import inspect
    import scripts.weekly_schedule.app as app

    source = inspect.getsource(app).lower()
    assert "openpyxl" not in source
    assert "load_workbook" not in source


def test_ui_builds_mapping_schema_without_excel_library_details():
    inspections = inspect_uploaded_workbook(TEMPLATE.read_bytes())

    sheet_names = {item.name for item in inspections}
    assert {
        "Tuan_hoc",
        "Thoi_khoa_bieu",
        "PPCT",
        "Tiet_da_day",
    }.issubset(sheet_names)

    selection = {
        "academic_weeks": {
            "sheet": "Tuan_hoc",
            "header_row": 1,
            "columns": {
                "academic_year": "nam_hoc",
                "week_number": "tuan",
                "start_date": "tu_ngay",
                "end_date": "den_ngay",
            },
        },
        "timetable": {
            "sheet": "Thoi_khoa_bieu",
            "header_row": 1,
            "columns": {
                "teacher_id": "ma_giao_vien",
                "class_id": "lop",
                "subject_ref": "mon_hoc",
                "component_ref": "phan_mon",
                "weekday": "thu",
                "timetable_period": "tiet_hoc",
                "effective_from": "hieu_luc_tu",
                "effective_to": "hieu_luc_den",
            },
        },
        "curriculum": {
            "sheet": "PPCT",
            "header_row": 1,
            "columns": {
                "class_id": "lop",
                "subject_ref": "mon_hoc",
                "component_ref": "phan_mon",
                "period_number": "tiet_ppct",
                "lesson_id": "ma_bai_hoc",
                "lesson_title": "ten_bai_hoc",
                "period_in_lesson": "tiet_trong_bai",
                "total_lesson_periods": "tong_tiet_cua_bai",
                "teaching_equipment": "thiet_bi_day_hoc",
            },
        },
        "executions": {
            "sheet": "Tiet_da_day",
            "header_row": 1,
            "columns": {
                "teacher_id": "ma_giao_vien",
                "class_id": "lop",
                "subject_ref": "mon_hoc",
                "component_ref": "phan_mon",
                "teaching_date": "ngay_day",
                "curriculum_period": "tiet_ppct",
                "status": "trang_thai",
            },
        },
    }

    schema = build_mapping_schema(selection)

    assert load_uploaded_workbook(
        TEMPLATE.read_bytes(),
        TEMPLATE.name,
        schema,
    ) == _data()


def test_app_can_save_list_and_reopen_schedule(tmp_path):
    schedule = build_weekly_schedule(
        data=_data(), teacher_id="GV001", academic_year="2026-2027", week_number=5
    )
    saved = save_weekly_schedule(schedule, tmp_path)
    assert saved.schedule_id == schedule.schedule_id
    assert saved_schedule_options("GV001", tmp_path)[0].week_number == 5
    assert load_saved_schedule(schedule.schedule_id, tmp_path) == schedule


def test_app_storage_helpers_keep_persistence_details_outside_core():
    import inspect
    from educational_planning_v2.services import WeeklyTeachingScheduleService

    source = inspect.getsource(WeeklyTeachingScheduleService).lower()
    assert "json" not in source
    assert "supabase" not in source
    assert "storage_root" not in source


def test_supabase_settings_requires_both_public_values():
    assert supabase_settings({}) is None
    assert supabase_settings({"SUPABASE_URL": "https://example.supabase.co"}) is None
    assert supabase_settings({
        "SUPABASE_URL": "https://example.supabase.co",
        "SUPABASE_PUBLISHABLE_KEY": "sb_publishable_test",
    }) == ("https://example.supabase.co", "sb_publishable_test")


def test_authentication_scopes_repository_to_returned_user():
    class User:
        id = "user-123"

    class Auth:
        def sign_in_with_password(self, credentials):
            assert credentials == {"email": "teacher@example.com", "password": "safe-pass"}
            return type("AuthResponse", (), {"user": User()})()

    client = type("Client", (), {"auth": Auth()})()
    repository = authenticate_supabase(client, " teacher@example.com ", "safe-pass")
    assert repository._user_id == "user-123"


def test_teacher_profile_is_entered_in_system_and_added_to_schedule_metadata():
    class Repository:
        def save(self, profile):
            return profile

    profile = save_teacher_profile(
        Repository(), teacher_code="GV001", full_name="Nguyễn Văn A",
        school_name="THCS Mẫu", subjects="Toán, Tin học, Toán",
        grade_levels="6, 7", default_academic_year="2026-2027",
        show_teacher_name=True, show_school_name=True,
    )
    schedule = build_weekly_schedule(
        data=_data(), teacher_id="GV001", academic_year="2026-2027",
        week_number=5, teacher_profile=profile,
    )
    assert comma_separated_values("Toán, Tin học, Toán") == ("Toán", "Tin học")
    assert schedule.metadata["teacher_profile"]["full_name"] == "Nguyễn Văn A"
    assert schedule.metadata["teacher_profile"]["school_name"] == "THCS Mẫu"


def test_export_uses_profile_display_preferences():
    from io import BytesIO
    from openpyxl import load_workbook
    from educational_planning_v2.models import TeacherProfile

    profile = TeacherProfile(
        "GV001", "Nguyễn Văn A", "THCS Mẫu", ("Toán",), ("6",), "2026-2027"
    )
    schedule = build_weekly_schedule(
        data=_data(), teacher_id="GV001", academic_year="2026-2027",
        week_number=5, teacher_profile=profile,
    )
    exported = export_weekly_schedule(schedule)
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    heading = workbook["Lich_bao_giang"]["A2"].value
    workbook.close()
    assert "Nguyễn Văn A" in heading
    assert "THCS Mẫu" in heading
